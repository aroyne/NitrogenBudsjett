"""
One-off utility for splicing a new flow row into an already-exported
Report.xlsx copy, between two existing marker rows in a sheet, without
disturbing merged cells or cell styling. Used when a flow gets added to the
model after a report has already been generated, so it doesn't have to be
regenerated from scratch just to include one new row.

requires: pip install openpyxl
"""
from openpyxl import load_workbook
from openpyxl.utils import (
    column_index_from_string, get_column_letter, range_boundaries
)
from copy import copy
from datetime import datetime
import re

def col_to_idx(c):
    return column_index_from_string(c) if isinstance(c, str) else int(c)

def find_next_index(ws, start_row, col_idx, value):
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v == value:
            return r
    return None

def find_all_ranges(ws, col_idx, start_marker, end_marker):
    ranges = []
    r = 1
    while r <= ws.max_row:
        start = find_next_index(ws, r, col_idx, start_marker)
        if start is None:
            break
        end = find_next_index(ws, start + 1, col_idx, end_marker)
        if end is None:
            end = ws.max_row
            ranges.append((start, end))
            break
        else:
            ranges.append((start, end))
            r = end + 1
    return ranges

def exists_between(ws, col_idx, start_row, end_row, target):
    for r in range(start_row + 1, end_row):
        v = ws.cell(row=r, column=col_idx).value
        if v == target:
            return True
    return False

def save_merged_info(ws):
    """
    Returns a list of dicts describing each merged range:
    {'coord': 'A1:A3', 'min_col':.., 'min_row':.., 'max_col':.., 'max_row':..,
     'top_left_val': ..., 'top_left_style': {...}}
    """
    out = []
    for m in list(ws.merged_cells.ranges):
        coord = m.coord
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        top_cell = ws.cell(row=min_row, column=min_col)
        # Save the top-left cell's value and a copy of its style.
        style = {}
        if top_cell.has_style:
            style['font'] = copy(top_cell.font)
            style['border'] = copy(top_cell.border)
            style['fill'] = copy(top_cell.fill)
            style['number_format'] = copy(top_cell.number_format)
            style['protection'] = copy(top_cell.protection)
            style['alignment'] = copy(top_cell.alignment)
        out.append({
            'coord': coord,
            'min_col': min_col,
            'min_row': min_row,
            'max_col': max_col,
            'max_row': max_row,
            'top_left_val': top_cell.value,
            'top_left_style': style
        })
    return out

def restore_merged_info(ws, merged_info):
    """
    Restores merged ranges from a list produced by save_merged_info (with
    already-updated row/column coordinates). Also sets the top-left cell's
    value and style.
    """
    # Remove any existing merges first, to avoid overlap.
    for m in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass

    for info in merged_info:
        coord = f"{get_column_letter(info['min_col'])}{info['min_row']}:{get_column_letter(info['max_col'])}{info['max_row']}"
        try:
            ws.merge_cells(coord)
            tl = ws.cell(row=info['min_row'], column=info['min_col'])
            tl.value = info['top_left_val']
            style = info.get('top_left_style', {})
            if style:
                tl.font = copy(style.get('font'))
                tl.border = copy(style.get('border'))
                tl.fill = copy(style.get('fill'))
                tl.number_format = copy(style.get('number_format'))
                tl.protection = copy(style.get('protection'))
                tl.alignment = copy(style.get('alignment'))
        except Exception as e:
            print("Advarsel: kunne ikke re-merge", coord, e)

def adjust_and_insert_row(ws, insert_at, amount=1):
    """
    Inserts rows without breaking merged cells:
    - saves the current merges and removes them,
    - shifts each merge's coordinates relative to insert_at,
    - inserts the new rows,
    - restores the merges at their shifted coordinates, with the top-left
      cell's content/style intact.
    Modifies ws in place; returns nothing.
    """
    merged_info = save_merged_info(ws)

    # Remove existing merges.
    for m in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass

    # Shift merge coordinates to account for the rows about to be inserted.
    new_infos = []
    for info in merged_info:
        min_col = info['min_col']
        max_col = info['max_col']
        min_row = info['min_row']
        max_row = info['max_row']

        if min_row >= insert_at:
            min_row += amount
            max_row += amount
        elif min_row < insert_at <= max_row:
            # Ranges spanning insert_at grow downward to keep covering it.
            max_row += amount
        # Ranges entirely before insert_at are unaffected.

        ni = info.copy()
        ni.update({'min_row': min_row, 'max_row': max_row})
        new_infos.append(ni)

    # Insert the rows.
    ws.insert_rows(insert_at, amount=amount)

    # Restore the merges at their shifted coordinates.
    restore_merged_info(ws, new_infos)

def copy_style_from_row_above(ws, target_row, max_col):
    src_row = target_row - 1
    if src_row < 1:
        return
    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        t = ws.cell(row=target_row, column=c)
        if s.has_style:
            t.font = copy(s.font)
            t.border = copy(s.border)
            t.fill = copy(s.fill)
            t.number_format = copy(s.number_format)
            t.protection = copy(s.protection)
            t.alignment = copy(s.alignment)

def set_year_from_above(ws, new_row, col_idx=16):
    above = ws.cell(row=new_row - 1, column=col_idx).value
    year = None
    if isinstance(above, datetime):
        year = above.year
    elif isinstance(above, int) and 1000 <= above <= 9999:
        year = above
    elif isinstance(above, str):
        s = above.strip()
        if re.fullmatch(r"\d{4}", s):
            year = int(s)
        else:
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    year = dt.year
                    break
                except Exception:
                    continue
    if year is not None:
        ws.cell(row=new_row, column=col_idx).value = int(year)
    else:
        ws.cell(row=new_row, column=col_idx).value = above

def fill_row_after_insert(ws, new_row, values_by_col_idx):
    max_col = max(values_by_col_idx.keys()) if values_by_col_idx else ws.max_column
    # Copy the row above's style for visual consistency.
    copy_style_from_row_above(ws, new_row, max_col)

    for col_idx, template in values_by_col_idx.items():
        cell = ws.cell(row=new_row, column=col_idx)
        if isinstance(template, str):
            if template == "{year_from_above}":
                set_year_from_above(ws, new_row, col_idx)
                continue
            if ("{r" in template) or ("{r}" in template):
                ctx = {'r': new_row, 'r_prev': new_row - 1, 'r_next': new_row + 1}
                try:
                    formatted = template.format(**ctx)
                except Exception as e:
                    print("Formateringsfeil:", template, e)
                    formatted = template
                cell.value = formatted
            elif template.startswith('='):
                cell.value = template
            else:
                cell.value = template
        else:
            cell.value = template

def ensure_target_between_markers(
    filename, sheetname,
    col_letter='C',
    start_marker='START',
    end_marker='END',
    target='TARGET',
    insert_position='after_start',
    values_to_set=None,
    out_filename=None
):
    """
    For every (start_marker, end_marker) range found in col_letter, inserts a
    row containing values_to_set immediately after start_marker (or before
    end_marker), unless target already appears somewhere in that range.
    Saves the result to out_filename (or filename with a '_modified' suffix)
    and returns that path.
    """
    col_idx = col_to_idx(col_letter)
    wb = load_workbook(filename)
    ws = wb[sheetname]

    ranges = find_all_ranges(ws, col_idx, start_marker, end_marker)
    ops = []
    for start_row, end_row in ranges:
        if not exists_between(ws, col_idx, start_row, end_row, target):
            if insert_position == 'after_start':
                insert_at = start_row + 1
            elif insert_position == 'before_end':
                insert_at = end_row
            else:
                insert_at = start_row + 1

            values_by_col_idx = {}
            if values_to_set:
                for k, v in values_to_set.items():
                    idx = col_to_idx(k)
                    values_by_col_idx[idx] = v
            else:
                values_by_col_idx[col_idx] = target

            ops.append((insert_at, values_by_col_idx))

    # Sort descending so earlier insertions don't shift later rows' target positions.
    ops.sort(key=lambda t: t[0], reverse=True)

    for insert_at, values_by_col_idx in ops:
        # Insert while preserving merged cells.
        adjust_and_insert_row(ws, insert_at, amount=1)
        # Fill the topmost of the newly-inserted rows.
        new_row = insert_at
        fill_row_after_insert(ws, new_row, values_by_col_idx)

    out_fn = out_filename or filename.replace('.xlsx', '_modified.xlsx')
    wb.save(out_fn)
    return out_fn

# -------------------- MAIN --------------------
if __name__ == "__main__":
    input_file = "Report-kopi.xlsx"
    sheet = "2a. Database N flows"
    start_marker = "FS.FO-PR.SO-Mixed waste-Nmix"
    end_marker = "FS.OL-AT.AT-Emissions-N2"
    flow_code = "FS.OL-AG.MM-Grazing-Nmix"
    flow = 'FS.OL-AG.MM'
    pool_out = 'FS'
    subpool_out = 'OL'
    pool_in = 'AG'
    subpool_in = 'MM'
    flow_name = 'Grazing'
    STAN_flow_name = 'grazing'
    react = 'reactive'
    species = 'Nmix'
    uncertainty = 0.2
    flow_type = 'useful output'

    values = {
        'B': flow,
        'C': flow_code,
        'D': pool_out,
        'E': subpool_out,
        'F': 0,
        'G': pool_in,
        'H': subpool_in,
        'I': 0,
        'J': flow_name,
        'K': STAN_flow_name,
        'L': react,
        'M': species,
        'N': 0,
        'O': uncertainty,
        'P': "{year_from_above}",
        'R': 'text',
        'S': 'text',
        'T': 'Norway',
        'U': flow_type,
        'V': "=(O{r}*N{r})^2"
    }

    out = ensure_target_between_markers(
        input_file, sheet,
        col_letter='C',
        start_marker=start_marker,
        end_marker=end_marker,
        target=flow_code,
        insert_position='after_start',
        values_to_set=values,
        out_filename="Report-kopi_med_innsatte_rader.xlsx"
    )
    print("Lagret:", out)
