#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CLI entry point and Monte Carlo driver for the nitrogen budget model: loads
source data and parameters once, runs the requested number of MC iterations
across the selected pools, then hands the collected results off to
utils_stat.py for statistical aggregation, to report_generator.py for the
GitHub Pages report, and optionally writes the aggregated medians/CVs back
into the official Report.xlsx template.
"""
import os
import sys
import time
import argparse
import numpy as np
import pandas as pd
import openpyxl

# Ensures the repo root is on the Python path regardless of the caller's cwd.
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from data_loader import load_all_data
from calculations.n_params import NParameters
from report_generator import generate_github_pages_report
from utils_stat import process_and_export_mc_results

REPORT_PATH = "Report.xlsx"
SHEET_NAME = "2a. Database N flows"

class YearNotFoundError(Exception): pass


def parse_arguments():
    """Handles command-line arguments for the MC run."""
    parser = argparse.ArgumentParser(description="Monte Carlo framework for the nitrogen model.")
    parser.add_argument(
        '--pool',
        type=str,
        required=True,
        help="Which pool(s) to run (e.g. 'at', 'rw', 'all')"
    )
    parser.add_argument(
        '--nsim',
        type=int,
        default=100,
        help="Number of Monte Carlo iterations to run"
    )
    parser.add_argument(
        '--no-excel',
        action='store_true',
        help="Skip writing results to the official Excel report"
    )
    return parser.parse_args()


def draw_from_pert(low, likely, high):
    """Draws one sample from a PERT distribution."""
    range_val = high - low
    if range_val == 0:
        return likely
    alpha = 1 + 4 * (likely - low) / range_val
    beta = 1 + 4 * (high - likely) / range_val
    return low + np.random.beta(alpha, beta) * range_val


def _draw_perturbed_value(val, low_b, upp_b, unc_type, dist_type):
    """
    Draws one Monte Carlo perturbed value for a parameter given its base
    value, uncertainty bounds and distribution type - the shared logic behind
    every perturbed table in generate_mc_parameters_fast (global parameters,
    waste fractions, dataset noise, trade parameters, animal weights).

    A blank (NaN) low_b/upp_b, or low_b == upp_b == 0, means the parameter is
    not perturbed; val is returned unchanged and neither the RNG nor
    low_b/upp_b are touched.

    For unc_type == 'perc', low_b/upp_b are +/- percentages of val. For any
    other unc_type ('abs'), low_b/upp_b are literal absolute bounds on the
    drawn value, not offsets from val.
    """
    if pd.isna(low_b) or pd.isna(upp_b) or (low_b == 0 and upp_b == 0):
        return val

    low_b = float(low_b)
    upp_b = float(upp_b)
    unc_type = str(unc_type).lower().strip()
    dist_type = str(dist_type).lower().strip()

    if unc_type == 'perc':
        abs_min = val * (1 - low_b / 100.0)
        abs_max = val * (1 + upp_b / 100.0)
        std_dev = ((low_b + upp_b) / 2.0 / 100.0) * val
    else:
        abs_min = low_b
        abs_max = upp_b
        std_dev = (low_b + upp_b) / 2.0 / 1.96

    if 'pert' in dist_type:
        chosen_val = draw_from_pert(abs_min, val, abs_max)
    elif 'log' in dist_type:
        cv = std_dev / val if val > 0 else 0.1
        sigma_log = np.sqrt(np.log(1 + cv**2))
        mu_log = np.log(val) - (sigma_log ** 2) / 2
        chosen_val = np.random.lognormal(mu_log, sigma_log)
    else:
        chosen_val = np.random.normal(val, std_dev)

    if val >= 0 and chosen_val < 0:
        chosen_val = 0.0

    return chosen_val


def generate_mc_parameters_fast(base_params, df_global, df_datasets, df_animal_products_static, df_trade_params=None, df_animal_weights=None, is_deterministic=False):
    """
    Draws global parameters, generates per-dataset noise factors, and draws
    perturbed N factors for trade goods, animal weights and animal products.
    Everything is collected as flat keys in custom_dict, which is pushed onto
    base_params via override_global_params.
    """
    # --- IDENTIFY THE TRADE DATA'S ID COLUMN ---
    pid_col = None
    df_trade_local = None
    if df_trade_params is not None:
        df_trade_local = df_trade_params.copy()
        idx_name = str(df_trade_local.index.name).lower().strip() if df_trade_local.index.name else ''
        if idx_name in ['param_id', 'parameter_id', 'konv', 'id']:
            df_trade_local = df_trade_local.reset_index()
            
        clean_cols = {str(c).lower().strip(): c for c in df_trade_local.columns}
        for variant in ['param_id', 'parameter_id', 'konv', 'id']:
            if variant in clean_cols:
                pid_col = clean_cols[variant]
                break
                
        if pid_col is None and len(df_trade_local) > 0:
            if isinstance(df_trade_local[df_trade_local.columns[0]].iloc[0], str):
                pid_col = df_trade_local.columns[0]

    # --- DETERMINISTIC BRANCH (round i=0) ---
    if is_deterministic:
        static_trade = {}
        if df_trade_local is not None and pid_col is not None:
            keys = df_trade_local[pid_col].astype(str).str.strip()
            static_trade = dict(zip(keys, df_trade_local['value']))
            
        custom_dict = {}
        
        if df_animal_weights is not None:
            for _, row in df_animal_weights.iterrows():
                t_id = f"weight_{str(row.name).strip()}" if df_animal_weights.index.name == 'item_name' else f"weight_{str(row['item_name']).strip()}"
                custom_dict[t_id] = float(row['avg_weight_kg'])
                
        if df_animal_products_static is not None:
            for _, row in df_animal_products_static.iterrows():
                p_id = f"prod_{str(row['item']).strip()}"
                custom_dict[p_id] = float(row['N_content_percent'])
                
        df_waste_static = base_params.get_table('waste_fractions')
        for _, row in df_waste_static.iterrows():
            cat_id = str(row['waste_category']).strip()
            custom_dict[cat_id] = float(row['N_frac'])

        # --- Protein food items (no noise in round 0) ---
        df_food_static = base_params.get_table('protein_food_items')
        for _, row in df_food_static.iterrows():
            f_id = f"food_protein_{str(row['food_group']).strip()}"
            custom_dict[f_id] = float(row['protein_content'])
        
        base_params.override_global_params(custom_dict)
        return base_params, {}, static_trade
        
    # --- 1. GLOBAL PARAMETERS ---
    custom_dict = {}
    df_perturbed = df_global.copy()

    global_idx_name = str(df_perturbed.index.name).lower().strip() if df_perturbed.index.name else ''
    if global_idx_name in ['param_id', 'parameter_id', 'id'] or 'param_id' not in df_perturbed.columns:
        df_perturbed = df_perturbed.reset_index()

    global_pid_col = 'param_id'
    for c in df_perturbed.columns:
        if str(c).lower().strip() in ['param_id', 'parameter_id', 'id']:
            global_pid_col = c
            break

    for idx, row in df_perturbed.iterrows():
        pid = row[global_pid_col]
        val = float(row['value'])
        chosen_val = _draw_perturbed_value(val, row['lower_bound'], row['upper_bound'], row['uncertainty_type'], row['distribution_type'])
        custom_dict[pid] = chosen_val
        df_perturbed.at[idx, 'value'] = chosen_val
        
    if hasattr(base_params, '_tables') and 'global_parameters' in base_params._tables:
        base_params._tables['global_parameters'] = df_perturbed
    elif hasattr(base_params, 'tables') and 'global_parameters' in base_params.tables:
        base_params.tables['global_parameters'] = df_perturbed

    # --- STEP 1.5: PERTURB WASTE_FRACTIONS ---
    df_waste = base_params.get_table('waste_fractions')
    for idx, row in df_waste.iterrows():
        cat_id = str(row['waste_category']).strip()
        val = float(row['N_frac'])
        custom_dict[cat_id] = _draw_perturbed_value(val, row['lower_bound'], row['upper_bound'], row['uncertainty_type'], row['distribution_type'])

    # --- 2. PER-DATASET NOISE FACTORS ---
    dataset_noise_dict = {}
    for _, row in df_datasets.iterrows():
        ds_id = str(row['dataset_name']).strip()
        dataset_noise_dict[ds_id] = float(_draw_perturbed_value(1.0, row['lower_bound'], row['upper_bound'], row['uncertainty_type'], row['distribution_type']))

    # --- 3. PERTURB TRADE_PARAMETERS ---
    trade_noise_dict = {}
    if df_trade_local is not None and pid_col is not None:
        for _, row in df_trade_local.iterrows():
            t_id = str(row[pid_col]).strip()
            val = float(row['value'])
            trade_noise_dict[t_id] = _draw_perturbed_value(val, row['lower_bound'], row['upper_bound'], row['uncertainty_type'], row['distribution_type'])

    # --- 4. PERTURB ANIMAL WEIGHTS ---
    if df_animal_weights is not None:
        for _, row in df_animal_weights.iterrows():
            t_id = f"weight_{str(row.name).strip()}" if df_animal_weights.index.name == 'item_name' else f"weight_{str(row['item_name']).strip()}"
            val = float(row['avg_weight_kg'])
            custom_dict[t_id] = _draw_perturbed_value(val, row['lower_bound'], row['upper_bound'], row['uncertainty_type'], row['distribution_type'])

    # --- 5. PERTURB ANIMAL PRODUCTS ---
    if df_animal_products_static is not None:
        for _, row in df_animal_products_static.iterrows():
            p_id = f"prod_{str(row['item']).strip()}"
            base_val = float(row['N_content_percent'])
            dist_type = str(row['distribution_type']).strip().lower()
            u_val = float(row['upper_bound']) / 100.0
            
            if u_val > 0:
                if dist_type == 'unif':
                    perturbed_val = base_val * np.random.uniform(1.0 - u_val, 1.0 + u_val)
                else:
                    perturbed_val = base_val * np.random.normal(1.0, u_val)
            else:
                perturbed_val = base_val
            
            if perturbed_val < 0: 
                perturbed_val = 0.0
            
            custom_dict[p_id] = perturbed_val
            
    # --- 6. PERTURB FOOD ITEM PROTEIN CONTENT ---
    df_food_items = base_params.get_table('protein_food_items')
    for idx, row in df_food_items.iterrows():
        f_group = str(row['food_group']).strip()
        f_id = f"food_protein_{f_group}"
        base_val = float(row['protein_content'])

        # uncertainty is a percentage (e.g. 10 means +/-10%)
        u_val = float(row['uncertainty']) / 100.0 if 'uncertainty' in row else 0.0

        if u_val > 0:
            perturbed_val = base_val * np.random.normal(1.0, u_val)
        else:
            perturbed_val = base_val
            
        if perturbed_val < 0:
            perturbed_val = 0.0
            
        custom_dict[f_id] = perturbed_val

    base_params.override_global_params(custom_dict)
    return base_params, dataset_noise_dict, trade_noise_dict

def write_mc_flows_to_international_report(summary_df):
    """
    Writes aggregated results directly into the official Excel template,
    using the exact column layout of the '2a. Database N flows' sheet.
    """
    if not os.path.exists(REPORT_PATH):
        print(f"[INFO] Fant ikke Excel-malen på '{REPORT_PATH}'. Hopper over offisiell rapportering.")
        return

    print(f"[EXCEL] Åpner offisiell rapporteringsmal: {REPORT_PATH}...")
    workbook = openpyxl.load_workbook(REPORT_PATH)

    if SHEET_NAME not in workbook.sheetnames:
        print(f"[ALARM] Fant ikke fanen '{SHEET_NAME}' i Excel-arket. Avbryter skriving.")
        return

    sheet = workbook[SHEET_NAME]
    print("[EXCEL] Skriver oppdaterte MC-resultater (Median og CV%) til Excel-databasen...")

    # Column layout of the '2a. Database N flows' sheet.
    FIRST_DATA_ROW = 3     # data starts on row 3, after the two header rows

    CODE_COL = 3           # column C: 'Flow Code' (the full unique name)
    VALUE_COL = 14         # column N: 'Value' (kt N)
    UNCERTAINTY_COL = 15   # column O: 'Uncertainty' (%)
    YEAR_COL = 16          # column P: 'Year'
    DATASOURCE_COL = 17    # column Q: 'Data sources'
    COMMENT_COL = 18       # column R: 'Comment'

    for _, row_data in summary_df.iterrows():
        flow_name = str(row_data["flow_name"]).strip()
        year = int(row_data["year"]) 
        value = row_data["median"]
        cv_percent = row_data["cv_percent"]
        comment = row_data.get("comment", "")
        data_sources = row_data.get("data_sources", "")

        year_found = False
        flow_found = False

        # Scan the sheet row by row.
        for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=YEAR_COL).value

            if cell_value is None:
                continue

            try:
                # Column P's year is sometimes stored as a string or a date.
                year_in_row = int(float(str(cell_value).strip()))
            except (ValueError, TypeError):
                continue

            # If the year matches, check whether the flow code in column C also matches.
            if year_in_row == year:
                year_found = True
                name_in_row = sheet.cell(row=row, column=CODE_COL).value or ""
                
                if str(name_in_row).strip() == flow_name:
                    # Write the values into their exact columns.
                    sheet.cell(row=row, column=VALUE_COL, value=value)

                    # CV% is stored as a real fraction in Excel (e.g. 0.20 for 20%).
                    sheet.cell(row=row, column=UNCERTAINTY_COL, value=cv_percent / 100.0)
                    
                    if data_sources:
                        sheet.cell(row=row, column=DATASOURCE_COL, value=data_sources)
                    if comment:
                        sheet.cell(row=row, column=COMMENT_COL, value=comment)
                        
                    flow_found = True
                    break

        if year_found and not flow_found:
            # Logged rather than raised: not every flow is necessarily reported in this template.
            print(f"[INFO] Strømkode '{flow_name}' ble ikke funnet for året {year} i Excel-malen. Hopper over.")
        if not year_found:
            raise YearNotFoundError(
                f"Året {year} mangler helt eller har feil format i kolonne P (16) i Excel-malen! "
                f"Sjekk rad {FIRST_DATA_ROW} og nedover."
            )

    workbook.save(REPORT_PATH)
    print("[SUCCESS] Det offisielle Excel-dokumentet er oppdatert på en strukturert måte!")

    
def main():
    args = parse_arguments()
    
    pool_input = args.pool.lower().strip()
    if pool_input == 'all':
        selected_pools = ['ag','at','ef', 'fs', 'hs', 'hy', 'mp', 'pr', 'rw']
    else:
        selected_pools = [p.strip() for p in pool_input.split(',')]

    print("="*60)
    print("[INFO] Starter MC-rammeverk.")
    print(f"[INFO] Aktiverte pooler: {', '.join(selected_pools).upper()}")
    print(f"[INFO] Antall ønskede iterasjoner: {args.nsim}")
    print("="*60)

    # 1. PRE-LOAD DATA AND PARAMETERS ONCE
    preloaded_data = load_all_data(selected_pools)

    print("[INFO] Pre-loader N_parameters.xlsx inn i RAM...")
    base_params = NParameters("parameters/N_parameters.xlsx")
    df_global_static = base_params.get_table('global_parameters')
    original_clean_dict = dict(zip(df_global_static['parameter_id'], df_global_static['value']))
    df_dataset_uncertainties = base_params.get_table('dataset_uncertainties')

    print("[INFO] Henter animal_weights tabell fra base_params...")
    df_animal_weights = base_params.get_table('animal_weights')
        
    print("[INFO] Henter animal_products tabell fra base_params...")
    df_animal_products_static = base_params.get_table('animal_products')
        
    if 'trade_params' not in preloaded_data:
        preloaded_data['trade_params'] = base_params.get_trade_params()


    all_mc_records = []
    
    print(f"\n[INFO] Starter simuleringsløkke: Kjører {args.nsim} iterasjoner...")
    start_time = time.time()

    for i in range(args.nsim):
        if args.nsim <= 10:
            print(i)
        elif args.nsim <= 100:
            if i/10-int(i/10)==0:
                print(['Starter iterasjon ',i,' av ', args.nsim])
        else:
            if i/50-int(i/50)==0:
                print(['Starter iterasjon ',i,' av ', args.nsim])
        # Reset tables to their static values before each round.
        if hasattr(base_params, '_tables'):
            base_params._tables['global_parameters'] = df_global_static.copy()
        elif hasattr(base_params, 'tables'):
            base_params.tables['global_parameters'] = df_global_static.copy()
        base_params.override_global_params(original_clean_dict)

        if i == 0:
            # Round 0: fully deterministic baseline.
            df_trade_params = preloaded_data['trade_params']
            current_params, _, current_trade_factors = generate_mc_parameters_fast(
                base_params, 
                df_global_static, 
                df_dataset_uncertainties,
                df_animal_products_static,
                df_trade_params=df_trade_params,
                df_animal_weights=df_animal_weights,
                is_deterministic=True
            )
            
            # Round 0 uses noise=1.0 for every registered dataset (no perturbation).
            dataset_noise = {}
            for _, row in df_dataset_uncertainties.iterrows():
                ds_id = str(row['dataset_name']).strip()
                dataset_noise[ds_id] = 1.0

        else:
            # Subsequent rounds: draw stochastic noise.
            df_trade_params = preloaded_data['trade_params']
            current_params, dataset_noise, current_trade_factors = generate_mc_parameters_fast(
                base_params, 
                df_global_static, 
                df_dataset_uncertainties,
                df_animal_products_static,
                df_trade_params=df_trade_params,
                df_animal_weights=df_animal_weights,
                is_deterministic=False
            )            
        iteration_output = {}
        
        # 2. RUN CALCULATIONS FOR THE SELECTED POOLS
        if 'at' in selected_pools:
            from calculations.at_mc import execute_calculations_at
            iteration_output['at'] = execute_calculations_at(preloaded_data, current_params, dataset_noise, current_trade_factors)
            
        if 'rw' in selected_pools:
            from calculations.rw_mc import execute_calculations_rw
            iteration_output['rw'] = execute_calculations_rw(preloaded_data, current_params, dataset_noise, current_trade_factors)

        if 'ag' in selected_pools:
            from calculations.ag_mc import execute_calculations_ag
            iteration_output['ag'] = execute_calculations_ag(preloaded_data, current_params, dataset_noise, current_trade_factors)

        if 'hy' in selected_pools:
            from calculations.hy_mc import execute_calculations_hy
            iteration_output['hy'] = execute_calculations_hy(preloaded_data, current_params, dataset_noise)

        if 'fs' in selected_pools:
            from calculations.fs_mc import execute_calculations_fs
            iteration_output['fs'] = execute_calculations_fs(preloaded_data, current_params, dataset_noise)

        if 'ef' in selected_pools:
            from calculations.ef_mc import execute_calculations_ef
            iteration_output['ef'] = execute_calculations_ef(preloaded_data, current_params, dataset_noise, current_trade_factors)

        if 'hs' in selected_pools:
            from calculations.hs_mc import execute_calculations_hs
            iteration_output['hs'] = execute_calculations_hs(preloaded_data, current_params, dataset_noise, current_trade_factors)

        if 'mp' in selected_pools:
            from calculations.mp_mc import execute_calculations_mp
            iteration_output['mp'] = execute_calculations_mp(preloaded_data, current_params, dataset_noise, current_trade_factors)

        if 'pr' in selected_pools:
            from calculations.pr_mc import execute_calculations_pr
            iteration_output['pr'] = execute_calculations_pr(preloaded_data, current_params, dataset_noise, current_trade_factors)
            
        # --- COLLECT RESULTS AND TAG THEM WITH SIM_ID ---
        for pool_name, pool_results in iteration_output.items():
            for row in pool_results:
                row_copy = row.copy()
                row_copy['sim_id'] = i
                all_mc_records.append(row_copy)
                
    elapsed_time = time.time() - start_time
    print(f"[SUKSESS] Simulering av {args.nsim} runder fullført på {elapsed_time:.4f} sekunder.")

    # 3. STATISTICAL ANALYSIS, EXCEL EXPORT AND PLOTTING
    summary_df = process_and_export_mc_results(all_mc_records)

    if args.no_excel:
        print("[INFO] Kjører UTEN å skrive til offisiell Excel-mal (--no-excel er aktiv).")
    else:
        write_mc_flows_to_international_report(summary_df)

    # 4. WEBSITE GENERATION
    generate_github_pages_report()


if __name__ == '__main__':
    main()