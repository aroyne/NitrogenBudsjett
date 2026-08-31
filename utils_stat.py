"""
Aggregates raw Monte Carlo iteration records into per-flow statistics
(process_and_export_mc_results: medians, 95% CIs, CV%), exports them to
MC_Reporting_Statistics.xlsx, and generates the time-series, pool-balance and
Sankey plots used in the GitHub Pages report.
"""
import os
import shutil
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime
import plotly.graph_objects as go

def plot_pool_balance_interactive(df_flows, pool_code, output_dir="output_files/plots"):
    """
    Genererer et INTERAKTIVT og RESPONSIVT balansediagram (HTML) for en spesifikk pool eller subpool.
    Inngående strømmer stables oppover (positive), utgående strømmer stables nedover (negative).
    Viser KUN info for strømmen musen er nærmest.
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Filtrer ut inngående og utgående strømmer for denne poolen
    df_in = df_flows[df_flows['target'].str.startswith(pool_code, na=False)]
    df_out = df_flows[df_flows['source'].str.startswith(pool_code, na=False)]
    
    if df_in.empty and df_out.empty:
        print(f"[WARN] Ingen data funnet for pool-balanse: {pool_code}")
        return None

    # Sørg for at alle årstall er synkronisert
    all_years = sorted(list(set(df_flows['year'])))
    
    # Grupper per år og fullt flomnavn
    df_in_grouped = df_in.groupby(['year', 'flow_name'])['value'].sum().unstack(fill_value=0).reindex(all_years, fill_value=0)
    df_out_grouped = df_out.groupby(['year', 'flow_name'])['value'].sum().unstack(fill_value=0).reindex(all_years, fill_value=0)
    
    # Beregn netto balanse og akkumulert usikkerhet
    df_in_total_unc = df_in.groupby('year')['uncertainty'].apply(lambda x: np.sqrt((x**2).sum())).reindex(all_years, fill_value=0)
    df_out_total_unc = df_out.groupby('year')['uncertainty'].apply(lambda x: np.sqrt((x**2).sum())).reindex(all_years, fill_value=0)
    
    net_balance = df_in_grouped.sum(axis=1) - df_out_grouped.sum(axis=1)
    combined_unc = np.sqrt(df_in_total_unc**2 + df_out_total_unc**2)

    # Opprett tom Plotly-figur
    fig = go.Figure()

    # --- 2. STACK INNGÅENDE STRØMMER (Positive) ---
    for col in df_in_grouped.columns:
        fig.add_trace(go.Scatter(
            x=all_years,
            y=df_in_grouped[col],
            mode='lines',
            name=f"IN: {col}",
            stackgroup='one',  
            groupnorm='',      
            hovertemplate=(
                f"<b>IN: {col}</b><br>" +
                "År: %{x}<br>" +
                "Verdi: %{y:.3f} kt N/year<br>" +
                "<extra></extra>"  
            ),
            legendgroup="Inngående",
            legendgrouptitle_text="══ SYSTEM INFLOW ══"
        ))

    # --- 3. STACK UTGÅENDE STRØMMER (Negative) ---
    for col in df_out_grouped.columns:
        fig.add_trace(go.Scatter(
            x=all_years,
            y=-df_out_grouped[col], 
            mode='lines',
            name=f"OUT: {col}",
            stackgroup='two',  
            hovertemplate=(
                f"<b>OUT: {col}</b><br>" +
                "År: %{x}<br>" +
                "Verdi: %{text:.3f} kt N/year<br>" + 
                "<extra></extra>"
            ),
            text=df_out_grouped[col], 
            legendgroup="Utgående",
            legendgrouptitle_text="══ SYSTEM OUTFLOW ══"
        ))

    # --- 4. USIKKERHETSBÅND ---
    fig.add_trace(go.Scatter(
        x=all_years,
        y=net_balance + combined_unc,
        mode='lines',
        line=dict(width=0),
        showlegend=False,
        hoverinfo='skip'
    ))
    fig.add_trace(go.Scatter(
        x=all_years,
        y=net_balance - combined_unc,
        mode='lines',
        line=dict(width=0),
        fill='tonexty',
        fillcolor='rgba(0, 0, 0, 0.12)',
        name='Uncertainty (±1σ)',
        hoverinfo='skip',
        legendgroup="Netto",
        legendgrouptitle_text="══ NET BALANCE ══"
    ))

    # --- 5. NETTO BALANSELINJE ---
    fig.add_trace(go.Scatter(
        x=all_years,
        y=net_balance,
        mode='lines',
        line=dict(color='black', width=3),
        name='Net Balance (Inn - Ut)',
        hovertemplate=(
            "<b>Net Balance</b><br>" +
            "År: %{x}<br>" +
            "Netto: %{y:.3f} kt N/year<br>" +
            "Usikkerhet: ±%{text:.3f}<br>" +
            "<extra></extra>"
        ),
        text=combined_unc,
        legendgroup="Netto"
    ))

    # --- 6. LAYOUT OG STYLING ---
    fig.update_layout(
        title=dict(
            text=f"Mass Balance Overview: {pool_code}",
            font=dict(size=16, family="Arial, sans-serif", color="black")
        ),
        xaxis=dict(
            title="Year",
            range=[1990, 2023],
            tickmode='array',
            tickvals=list(np.arange(1990, 2021, 5)) + [2023],
            gridcolor='rgba(200, 200, 200, 0.4)',
            showspikes=True,      
            spikethickness=1,
            spikedash="dot",
            spikemode="across"
        ),
        yaxis=dict(
            title="Nitrogen Flow (kt N / year)",
            gridcolor='rgba(200, 200, 200, 0.4)',
            zeroline=True,
            zerolinecolor='gray',
            zerolinewidth=1
        ),
        # "closest" isolates the hover tooltip to only the trace under the cursor.
        hovermode="closest",
        plot_bgcolor='white',
        paper_bgcolor='white',
        legend=dict(
            x=1.02, # Plassert rett utenfor plottet
            y=1.0,
            xanchor='left',
            yanchor='top',
            font=dict(size=10),
            traceorder="grouped" 
        ),
        # Wide right margin (r=180) makes room for the legend.
        margin=dict(l=60, r=180, t=60, b=50),
        height=600
    )

    # Lagre med default_width='100%' for full responsivitet i iframes
    plot_filename = f"balance_{pool_code.replace('.', '_')}.html"
    filepath = os.path.join(output_dir, plot_filename)
    fig.write_html(filepath, include_plotlyjs='cdn', default_width='100%')
    
    print(f"[INFO] Interaktivt balanseplott generert for {pool_code} -> {filepath}")
    return plot_filename


def plot_pool_balance(df_flows, pool_code, output_dir="output_files/plots"):
    """
    Genererer et balansediagram for en spesifikk pool eller subpool.
    Deler legenden inn i to ryddige blokker: "Inngående strømmer" og "Utgående strømmer",
    og viser den fulle flomkoden (flow_name) for hver strøm.
    Rekkefølgen i legendene matcher stablingen i plottet (visuelt ovenfra og ned).
    """
    os.makedirs(output_dir, exist_ok=True)

    # 1. Filtrer ut inngående og utgående strømmer for denne poolen
    df_in = df_flows[df_flows['target'].str.startswith(pool_code, na=False)]
    df_out = df_flows[df_flows['source'].str.startswith(pool_code, na=False)]
    
    if df_in.empty and df_out.empty:
        print(f"[WARN] Ingen data funnet for pool-balanse: {pool_code}")
        return None

    # 2. Grupper per år og fullt flomnavn (flow_name) for stacking
    df_in_grouped = df_in.groupby(['year', 'flow_name'])['value'].sum().unstack(fill_value=0)
    df_out_grouped = df_out.groupby(['year', 'flow_name'])['value'].sum().unstack(fill_value=0)
    
    # Hent usikkerhetene (kvadratrot av summen av kvadrater)
    df_in_unc = df_in.groupby('year')['uncertainty'].apply(lambda x: np.sqrt((x**2).sum()))
    df_out_unc = df_out.groupby('year')['uncertainty'].apply(lambda x: np.sqrt((x**2).sum()))
    
    # Sørg for at alle årstall er synkronisert (1984-2025)
    all_years = sorted(list(set(df_flows['year'])))
    df_in_grouped = df_in_grouped.reindex(all_years, fill_value=0)
    df_out_grouped = df_out_grouped.reindex(all_years, fill_value=0)
    df_in_unc = df_in_unc.reindex(all_years, fill_value=0)
    df_out_unc = df_out_unc.reindex(all_years, fill_value=0)

    # 3. Beregn netto balanse og akkumulert usikkerhet
    total_in = df_in_grouped.sum(axis=1)
    total_out = df_out_grouped.sum(axis=1)
    net_balance = total_in - total_out
    
    combined_unc = np.sqrt(df_in_unc**2 + df_out_unc**2)

    # 4. Plottingen
    fig, ax = plt.subplots(figsize=(11, 6), dpi=300)
    
    # Lagre plot-objekter (handles) for å kunne bygge legendene manuelt etterpå
    in_handles = []
    out_handles = []
    
    # Stack inngående (Positive verdier)
    if not df_in_grouped.empty:
        # stackplot returnerer en liste med PolyCollection-objekter (ett for hvert lag)
        polys_in = ax.stackplot(all_years, df_in_grouped.values.T, alpha=0.7)
        in_handles = list(polys_in)
    
    # Stack utgående (Negative verdier)
    if not df_out_grouped.empty:
        polys_out = ax.stackplot(all_years, (-df_out_grouped).values.T, alpha=0.7)
        out_handles = list(polys_out)
    
    # Tegn den svarte balanselinjen og usikkerhetsbåndet
    line_balance, = ax.plot(all_years, net_balance, color='black', linewidth=2, zorder=5)
    poly_unc = ax.fill_between(all_years, net_balance - combined_unc, net_balance + combined_unc, 
                               color='black', alpha=0.12, zorder=4, linestyle='--')
    
    # Styling av akser
    ax.axhline(0, color='gray', linestyle='-', linewidth=0.8, zorder=3)
    ax.set_title(f"Mass Balance Overview: {pool_code}", fontsize=12, fontweight='bold', loc='left')
    ax.set_xlabel("Year", fontsize=10)
    ax.set_ylabel("Nitrogen Flow (kt N / year)", fontsize=10)
    
    # --- AKSEBEGRENSNINGER ---
    ax.set_xlim(1990, 2023)
    custom_ticks = list(np.arange(1990, 2021, 5)) + [2023]
    ax.set_xticks(custom_ticks)
    ax.grid(True, linestyle='--', alpha=0.4)

    # ========================================================
    # AVANSERT LEGEND-HÅNDTERING MED KORREKT VISUELL REKKEFØLGE
    # ========================================================
    
    # 1. Hent de opprinnelige merkelappene kronologisk fra DataFrame-kolonnene
    in_labels = list(df_in_grouped.columns)
    out_labels = list(df_out_grouped.columns)
    
    # Sortering for Inngående (Positive):
    # Siste kolonne ligger øverst i plottet, så vi reverserer listene 
    # for å få den øverst i den øverste legend-blokken.
    in_handles_sorted = list(reversed(in_handles))
    in_labels_sorted = list(reversed(in_labels))
    
    # Sortering for Utgående (Negative):
    # Siste kolonne pushes lengst NED i plottet (bort fra 0). 
    # Ved å beholde den opprinnelige rekkefølgen havner den også nederst i legend-blokken.
    out_handles_sorted = out_handles
    out_labels_sorted = out_labels
    
    # 2. Opprett den første legenden for INNGÅENDE (øverst til høyre)
    # Vi inkluderer Net Balance og Uncertainty øverst i denne blokken
    top_handles = [line_balance, poly_unc] + in_handles_sorted
    top_labels = ['Net Balance (Inn - Ut)', 'Uncertainty (±1σ)'] + in_labels_sorted
    
    legend_in = ax.legend(
        top_handles, 
        top_labels, 
        bbox_to_anchor=(1.05, 1.0), 
        loc='upper left', 
        fontsize=8, 
        title="══ SYSTEM INFLOW & NET ══",
        title_fontsize=9,
        frameon=True
    )
    legend_in._legend_box.align = "left"
    ax.add_artist(legend_in)
    
    # 3. Opprett den andre legenden for UTGÅENDE (plassert under den første)
    if out_handles_sorted:
        # Dynamisk plassering basert på antall inngående strømmer
        approx_offset = max(0.0, 0.55 - (len(in_labels) * 0.03))
        
        legend_out = ax.legend(
            out_handles_sorted, 
            out_labels_sorted, 
            bbox_to_anchor=(1.05, approx_offset), 
            loc='upper left', 
            fontsize=8, 
            title="══ SYSTEM OUTFLOW ══",
            title_fontsize=9,
            frameon=True
        )
        legend_out._legend_box.align = "left"

    plt.tight_layout()
    
    # Lagre filen
    plot_filename = f"balance_{pool_code.replace('.', '_')}.png"
    filepath = os.path.join(output_dir, plot_filename)
    plt.savefig(filepath, bbox_inches='tight')
    plt.close()
    
    print(f"[INFO] Balanseplott generert for {pool_code} -> {filepath}")
    return plot_filename

SANKEY_SPECIES_COLORS = {
    'N2O': "rgba(156, 39, 176, 0.55)",
    'N2 (excl. NOx)': "rgba(140, 140, 140, 0.55)",
    'NH3 / RDN': "rgba(255, 152, 0, 0.55)",
    'NOx / OXN': "rgba(244, 67, 54, 0.55)",
    'Nmix': "rgba(76, 175, 80, 0.55)",
    'Other': "rgba(33, 150, 243, 0.55)",
}

SANKEY_CATEGORY_COLORS = {
    'useful output': "rgba(76, 175, 80, 0.55)",
    'N wasted': "rgba(244, 67, 54, 0.55)",
    'import': "rgba(33, 150, 243, 0.55)",
    'recycling': "rgba(255, 152, 0, 0.55)",
    'other/unclassified': "rgba(160, 160, 160, 0.55)",
}

SANKEY_NODE_COLORS = {
    "AT": "rgba(26, 54, 93, 0.85)",
    "HY": "rgba(43, 108, 176, 0.85)",
    "RW": "rgba(74, 85, 104, 0.85)",
    "AG": "#2f855a", "EF": "#c53030", "FS": "#2c7a7b", "HS": "#744210", "PR": "#97266d", "MP": "#d69e2e",
}
SANKEY_NODE_X = {
    "AT": 0.02, "AG": 0.40, "FS": 0.40, "EF": 0.65, "PR": 0.65, "MP": 0.65, "HS": 0.65, "HY": 0.98, "RW": 0.98,
}


def _get_species_color(flow_name):
    """Colors a link by which N species it carries."""
    fn = flow_name.upper()
    if "N2O" in fn:
        return SANKEY_SPECIES_COLORS['N2O']
    elif "N2" in fn and "NOX" not in fn:
        return SANKEY_SPECIES_COLORS['N2 (excl. NOx)']
    elif "NH3" in fn or "AMMONIA" in fn or "RDN" in fn:
        return SANKEY_SPECIES_COLORS['NH3 / RDN']
    elif "NOX" in fn or "OXN" in fn or "NITRITE" in fn or "NITRATE" in fn:
        return SANKEY_SPECIES_COLORS['NOx / OXN']
    elif "NMIX" in fn:
        return SANKEY_SPECIES_COLORS['Nmix']
    else:
        return SANKEY_SPECIES_COLORS['Other']


def _load_flow_categories(report_path='Report.xlsx', sheet_name='2a. Database N flows'):
    """
    Loads the manually-maintained flow -> overview-category mapping from the
    official Report.xlsx template ('2a. Database N flows', column C = flow
    code, column U = overViewType: useful output / N wasted / import /
    recycling / -), keyed by flow code. The two spellings of 'N wasted' found
    in the template ('N wasted' and 'N wastedd') are merged into one category.
    """
    category_map = {}
    if not os.path.exists(report_path):
        return category_map
    wb_report = openpyxl.load_workbook(report_path, data_only=True)
    if sheet_name not in wb_report.sheetnames:
        return category_map
    sheet = wb_report[sheet_name]
    for row in range(3, sheet.max_row + 1):
        code = sheet.cell(row=row, column=3).value
        raw_cat = sheet.cell(row=row, column=21).value
        if code is None or raw_cat is None:
            continue
        cat = str(raw_cat).strip()
        if cat.lower() in ('n wasted', 'n wastedd'):
            cat = 'N wasted'
        elif cat == '-':
            cat = 'other/unclassified'
        category_map[str(code).strip()] = cat
    return category_map


def _get_category_color(flow_name, category_map):
    cat = category_map.get(flow_name, 'other/unclassified')
    return SANKEY_CATEGORY_COLORS.get(cat, SANKEY_CATEGORY_COLORS['other/unclassified'])


def plot_global_sankey_interactive(df_flows, output_dir="output_files/plots"):
    """
    Generates two interactive Sankey diagrams at the main-pool level (AT, EF,
    AG, ...), limited to 1990-2023: one with every flow, one with the
    dominant trade flows (fertilizer/ammonia trade and crude oil export)
    hidden so the smaller internal flows stay legible. Both let the viewer
    scrub through years via a slider
    and switch link coloring between N species and Report.xlsx's
    useful-output/waste/import/recycling categorization.

    Locking the scale across years: Plotly's Sankey layout scales each
    column (the nodes sharing an x-position) to fit the plot height based on
    that column's total throughput, recomputed independently for every
    frame. A single global buffer link therefore isn't enough to keep the
    scale constant if different columns dominate in different years, which
    is the case here (e.g. the AT column's total throughput varies more than
    2x across 1990-2023). Instead, each column gets one dedicated,
    disconnected filler link (own source and sink node, invisible, touching
    no real node) padded up to that column's fixed historical peak
    footprint every year, so every column's total - and thus its
    value-to-pixel scale - never changes. The filler is deliberately kept
    off the real nodes themselves (rather than attached to their own
    inflow/outflow) so a real node's displayed size is always its true value
    for that year, not inflated toward its own historical peak.
    """
    os.makedirs(output_dir, exist_ok=True)

    df_base = df_flows.copy()
    df_base = df_base[(df_base['year'] >= 1990) & (df_base['year'] <= 2023)]
    if df_base.empty:
        print("[WARN] Ingen data funnet for tidsperioden 1990-2023.")
        return None

    res = df_base['flow_name'].apply(extract_source_target)
    df_base['source_pool'] = [r[0].split('.')[0] for r in res]
    df_base['target_pool'] = [r[1].split('.')[0] for r in res]

    df_base = df_base[(df_base['source_pool'] != "Unknown") & (df_base['target_pool'] != "Unknown")]
    df_base = df_base[df_base['source_pool'] != df_base['target_pool']]
    if df_base.empty:
        print("[WARN] Ingen gyldige strømmer funnet til å generere Sankey-diagram.")
        return None

    df_base = df_base.groupby(['year', 'source_pool', 'target_pool', 'flow_name'], as_index=False).agg({
        'value': 'sum',
        'uncertainty': lambda x: np.sqrt((x**2).sum())
    })

    category_map = _load_flow_categories()
    df_base['color_species'] = df_base['flow_name'].apply(_get_species_color)
    df_base['color_category'] = df_base['flow_name'].apply(lambda fn: _get_category_color(fn, category_map))
    # Hover label: the descriptive part of the flow code only (e.g. "Emissions-N2O"),
    # since the pool codes are already identified by which nodes the link connects.
    df_base['hover_label'] = df_base['flow_name'].str.split('-').str[2:].str.join('-')

    hidden_keywords = ["AMMONIA IMPORT", "AMMONIA EXPORT", "AMMONIA SYNTHESIS", "FERTILIZER EXPORT", "FUEL EXPORT"]
    filter_regex = "|".join(hidden_keywords)
    df_filtered = df_base[~df_base['flow_name'].str.upper().str.contains(filter_regex, na=False)].copy()

    def build_sankey_figure(df_data, title_suffix, filename):
        base_nodes = sorted(list(set(df_data['source_pool'].unique()) | set(df_data['target_pool'].unique())))

        # Padding used to be attached directly to each real node's own
        # inflow/outflow, which does lock the scale, but it also means a
        # real node's displayed box height is its *padded* total (its own
        # historical peak) rather than that year's true value - a small pool
        # in a light year would show as tall as its heaviest year ever. A
        # node's box height is simply the sum of its own connected link
        # values, so there is no way to pad a real node without inflating
        # it. A later per-column redesign (one disconnected filler link per
        # x-position) turned out not to match Plotly's actual behavior
        # either: measuring real nodes' rendered pixel height against their
        # true kt value confirmed empirically (via a headless-browser probe
        # reading Plotly's own trace/link data back out of the page) that
        # the value-to-pixel scale is ONE ratio for the *entire* figure, not
        # set independently per column, and that scale is driven by the sum
        # of every node's own max(inflow, outflow) - counting each filler
        # node separately, since a disconnected filler is still two distinct
        # node boxes (source and sink) that each contribute their own value
        # to that sum. So a single global filler pair - not one per column -
        # sized to keep that whole-figure sum constant is what actually
        # locks the scale, and it can sit anywhere since the scale isn't
        # column-local.
        fill_src, fill_sink = "__FILL_SRC", "__FILL_SINK"
        all_nodes = base_nodes + [fill_src, fill_sink]
        node_indices = {node: i for i, node in enumerate(all_nodes)}
        n_fill = 2

        node_colors = [SANKEY_NODE_COLORS.get(n, "#bdc3c7") for n in base_nodes] + ["rgba(0,0,0,0)"] * n_fill
        node_x = [SANKEY_NODE_X.get(n, 0.5) for n in base_nodes] + [0.5, 0.5]
        # Filler nodes are pinned near the top so they never push the real
        # nodes around; None lets Plotly place real nodes freely.
        node_y = [None] * len(base_nodes) + [0.001] * n_fill

        # No node.hovertemplate: sankey 'hoverinfo' is a single enum for the
        # whole trace (not arrayable per element), and setting even one
        # array entry of node.hovertemplate turns out to blank out the
        # hover text for EVERY node on the trace, including the ones left as
        # None - the same CDN Plotly.js (3.7.0) bug as the one on link
        # hovertemplate below. The filler nodes don't need a suppressed
        # tooltip of their own: pointer-events is set to none on them (see
        # fixPadElements below), so the mouse can't hover them at all, and
        # the real nodes are left to Plotly's own default hover (their pool
        # code plus total value).
        #
        # Node/link borders aren't set here via the `line` attribute either:
        # Plotly's sankey renderer doesn't apply an arrayed node.line/
        # link.line per element (verified by inspecting the rendered SVG -
        # every node and link, including the fully transparent filler ones,
        # came out with the same hardcoded opaque black 1px border
        # regardless of what was requested here). Borders are instead
        # corrected after rendering, in the post-script below, by reading
        # each element's actual paint back out of the DOM.
        static_node_config = dict(
            pad=20,
            thickness=25,
            label=list(base_nodes) + [""] * n_fill,
            color=node_colors,
            x=node_x,
            y=node_y,
        )

        all_years = sorted(list(df_data['year'].unique()))

        def node_max_through(df_yr, node):
            out_total = df_yr.loc[df_yr['source_pool'] == node, 'value'].sum()
            in_total = df_yr.loc[df_yr['target_pool'] == node, 'value'].sum()
            return max(out_total, in_total)

        def real_total(df_yr):
            return sum(node_max_through(df_yr, n) for n in base_nodes)

        # 5% margin, same as the earlier per-column design.
        global_ceiling = max(real_total(df_data[df_data['year'] == yr]) for yr in all_years) * 1.05

        def get_sankey_components(df_yr, color_col):
            sources = [node_indices[s] for s in df_yr['source_pool']]
            targets = [node_indices[t] for t in df_yr['target_pool']]
            values = df_yr['value'].tolist()
            colors = df_yr[color_col].tolist()
            # No explicit link.hovertemplate: the CDN-pulled Plotly.js
            # version (3.7.0) has a bug where any hovertemplate on a sankey
            # LINK renders a completely empty tooltip (confirmed in
            # isolation, down to a 3-node/2-link minimal reproduction, with
            # both token-based and plain-literal templates - the highlight-
            # on-hover effect still works, only the text is silently
            # missing). Plotly's own default link hover already reads
            # link.label - set here from hover_label, the flow's descriptive
            # name only (e.g. "Emissions-N2O"), not the full
            # source.target-coded flow_name - alongside the value, so
            # dropping hovertemplate keeps the wanted name+value on hover
            # without hitting the bug. The filler link doesn't need a
            # suppressed tooltip of its own: pointer-events is set to none
            # on it (see fixPadElements below) so the mouse can't hover it
            # at all.
            labels = df_yr['hover_label'].tolist()

            # The filler link is present in every single frame, even when
            # the padding needed is zero. Sankey's d3 rendering joins
            # consecutive frames' links by position/identity to animate
            # between them; letting the link disappear in some frames breaks
            # that identity match and leaves stray elements on screen with a
            # stale, pre-transition style (usually an opaque black border)
            # instead of the requested transparent one.
            #
            # Divided by 2, not subtracted directly: the filler source and
            # sink are two separate node boxes, each independently
            # contributing its own value to the figure-wide scale sum, so a
            # single filler link of value v adds 2v to that sum, not v.
            pad_value = max(0.0, (global_ceiling - real_total(df_yr)) / 2.0)
            sources.append(node_indices[fill_src])
            targets.append(node_indices[fill_sink])
            values.append(pad_value); colors.append("rgba(0,0,0,0)"); labels.append("")

            return dict(
                source=sources, target=targets, value=values, color=colors, label=labels,
            )

        color_modes = {'species': 'color_species', 'category': 'color_category'}
        frames = []
        for yr in all_years:
            df_yr = df_data[df_data['year'] == yr]
            for mode_name, color_col in color_modes.items():
                frames.append(go.Frame(
                    data=[go.Sankey(node=static_node_config, link=get_sankey_components(df_yr, color_col), arrangement='fixed')],
                    name=f"{yr}|{mode_name}",
                ))

        first_year = all_years[0]
        initial_sankey = go.Sankey(
            node=static_node_config,
            link=get_sankey_components(df_data[df_data['year'] == first_year], 'color_species'),
            arrangement='fixed',
        )

        slider_steps = [dict(
            method="animate",
            args=[[f"{yr}|species"], dict(mode="immediate", frame=dict(duration=200, redraw=True), transition=dict(duration=0))],
            label=str(yr),
        ) for yr in all_years]

        # The legend is built as plain HTML/CSS in the post-script below, not
        # as Plotly traces: adding go.Scatter "dummy" traces alongside a
        # go.Sankey trace forces Plotly to create a real cartesian x/y axis
        # for the Scatter traces (Sankey doesn't use one), which then shows
        # up as a visible, oddly-scaled axis fighting the Sankey for space.

        fig = go.Figure(data=[initial_sankey], frames=frames)
        fig.update_layout(
            title=dict(text=f"Global Nitrogen Flow Evolution (1990-2023) - {title_suffix}", font=dict(size=18, family="Arial")),
            height=750,
            margin=dict(l=20, r=180, t=60, b=110),
            sliders=[dict(
                active=0, steps=slider_steps, x=0.08, y=-0.05,
                currentvalue=dict(font=dict(size=14, color="navy"), prefix="Year: ", visible=True),
                len=0.75,
            )],
        )

        # Legend swatches: same colors as the link coloring, built as plain
        # HTML rather than Plotly traces (see note above on why).
        def _swatch_rows(color_dict):
            return "".join(
                f'<div style="margin-bottom:3px;"><span style="display:inline-block;width:12px;height:12px;'
                f'background:{c.replace("0.55", "1.0")};margin-right:6px;border-radius:2px;"></span>{label}</div>'
                for label, c in color_dict.items()
            )
        species_legend_html = _swatch_rows(SANKEY_SPECIES_COLORS)
        category_legend_html = _swatch_rows(SANKEY_CATEGORY_COLORS)

        # Custom JS: native `updatemenus` buttons can't read the figure's
        # current animation state, so a button-driven color-mode toggle can't
        # combine with "whatever year is currently showing" using the
        # declarative updatemenus/sliders spec alone. We track the current
        # year ourselves (from the slider's change event) and drive real HTML
        # buttons that jump to the frame named "{year}|{mode}", and a
        # Play/Pause pair that steps through the current mode's frames.
        #
        # write_html() itself appends a `Plotly.animate(gd, null)` call right
        # after the initial Plotly.newPlot() - Plotly.js's built-in shorthand
        # for "play every frame in sequence" - which is why the diagram used
        # to auto-play through every year/mode combination on open. The first
        # line below runs in a later .then() callback and immediately jumps
        # to a single, fixed frame, which cancels that queued auto-play.
        all_years_js = "[" + ",".join(f"'{y}'" for y in all_years) + "]"
        post_script = f"""
        var gd = document.getElementsByClassName('plotly-graph-div')[0];
        Plotly.animate(gd, ['{first_year}|species'], {{frame: {{duration: 0, redraw: true}}, transition: {{duration: 0}}, mode: 'immediate'}});

        // Plotly's own sankey hover box splits into two side-by-side boxes -
        // a colored one (the link/node's own color) holding just the value,
        // and a white one holding the label/source/target text - which
        // reads as two disconnected tooltips rather than one. Since
        // hoverinfo can only be 'all'/'none'/'skip' for sankey (no way to
        // keep just the label half) and hovertemplate is unusable (see the
        // note on link.hovertemplate below), Plotly's own hover box is
        // hidden outright and replaced with one plain custom tooltip built
        // from the same plotly_hover event data, styled to match the rest
        // of this page's controls.
        var style = document.createElement('style');
        style.textContent = '.hoverlayer {{ display: none !important; }}';
        document.head.appendChild(style);

        var tooltip = document.createElement('div');
        tooltip.style.position = 'fixed';
        tooltip.style.pointerEvents = 'none';
        tooltip.style.background = 'rgba(255,255,255,0.95)';
        tooltip.style.border = '1px solid #888';
        tooltip.style.borderRadius = '4px';
        tooltip.style.padding = '5px 9px';
        tooltip.style.fontFamily = 'Arial, sans-serif';
        tooltip.style.fontSize = '13px';
        tooltip.style.display = 'none';
        tooltip.style.zIndex = '10000';
        document.body.appendChild(tooltip);

        gd.on('plotly_hover', function(evt) {{
            var pt = evt.points[0];
            tooltip.textContent = pt.label + ': ' + pt.value.toFixed(2) + ' kt N';
            tooltip.style.display = 'block';
        }});
        gd.on('plotly_unhover', function() {{ tooltip.style.display = 'none'; }});
        gd.addEventListener('mousemove', function(e) {{
            tooltip.style.left = (e.clientX + 14) + 'px';
            tooltip.style.top = (e.clientY + 14) + 'px';
        }});

        // Plotly's sankey renderer paints every node-rect/sankey-link with a
        // hardcoded opaque black 1px border, ignoring the node.line/link.line
        // trace attributes entirely (confirmed by inspecting the rendered
        // SVG - both real and fully-transparent filler elements came out
        // with identical black borders no matter what those attributes
        // said), and it doesn't stop the filler links/node from lighting up
        // and showing an (empty) tooltip on hover either. Both are corrected
        // after the fact here: a filler element is identified by its own
        // fill being fully transparent (fill-opacity 0), since only the
        // invisible filler link/nodes are ever colored that way.
        function fixPadElements() {{
            document.querySelectorAll('rect.node-rect, path.sankey-link').forEach(function(el) {{
                var isPad = parseFloat(getComputedStyle(el).fillOpacity) === 0;
                if (isPad) {{
                    el.style.stroke = 'none';
                    el.style.pointerEvents = 'none';
                }} else if (el.tagName.toLowerCase() === 'rect') {{
                    el.style.stroke = 'rgba(0,0,0,0.6)';
                    el.style.strokeWidth = '0.5px';
                }} else {{
                    el.style.stroke = 'rgba(50,50,50,0.3)';
                    el.style.strokeWidth = '0.5px';
                }}
            }});
        }}

        // The figure's own vertical extent always includes room for the
        // filler node near the top (needed to keep the value-to-pixel scale
        // constant across years - see the note on the filler construction
        // above), which shows up as blank space above the real diagram
        // whenever that year doesn't need much filler. Rather than trying
        // to precompute how much room that takes (it changes every year),
        // this measures the real, visible content's own bounding box after
        // each render and crops a wrapper div around exactly that, so the
        // reserved filler space is clipped away instead of shown as padding.
        var plotWrapper = document.createElement('div');
        plotWrapper.style.overflow = 'hidden';
        gd.parentNode.insertBefore(plotWrapper, gd);
        plotWrapper.appendChild(gd);

        function cropToContent() {{
            gd.style.transform = 'none';
            // Real links can bulge above/below the node boxes they connect
            // (a loop-back curve dips well past both endpoints), so the
            // content bounds need every real node AND real link, not just
            // the nodes.
            var realEls = Array.from(document.querySelectorAll('rect.node-rect, path.sankey-link')).filter(function(el) {{
                return parseFloat(getComputedStyle(el).fillOpacity) > 0;
            }});
            var titleEl = document.querySelector('.g-gtitle');
            var sliderEl = document.querySelector('g.slider-container');
            var minTop = Infinity, maxBottom = -Infinity;
            if (titleEl) {{
                var tb = titleEl.getBoundingClientRect();
                minTop = Math.min(minTop, tb.top); maxBottom = Math.max(maxBottom, tb.bottom);
            }}
            if (sliderEl) {{
                // The slider sits at a fixed position regardless of frame,
                // but the real content's own bounding box shrinks in low-
                // value years - without this, a small enough year crops the
                // wrapper shorter than the (unmoved) slider, clipping its
                // year tick labels.
                var sb = sliderEl.getBoundingClientRect();
                minTop = Math.min(minTop, sb.top); maxBottom = Math.max(maxBottom, sb.bottom);
            }}
            realEls.forEach(function(el) {{
                var b = el.getBoundingClientRect();
                minTop = Math.min(minTop, b.top); maxBottom = Math.max(maxBottom, b.bottom);
            }});
            if (!isFinite(minTop)) return;
            var margin = 30;
            var wrapperTop = plotWrapper.getBoundingClientRect().top;
            var shiftUp = Math.max(0, (minTop - wrapperTop) - margin);
            plotWrapper.style.height = ((maxBottom - minTop) + 2 * margin) + 'px';
            gd.style.transform = 'translateY(-' + shiftUp + 'px)';
        }}

        // A MutationObserver (rather than a fixed setTimeout delay) reacts
        // to Plotly's own redraw synchronously, in the same tick, so there's
        // no gap where an unfixed frame - with its default black border or
        // its untrimmed height - gets painted and briefly flashes on screen
        // (which happened during Play's rapid automatic frame changes with
        // the previous, delay-based version of this fix). fixPadElements()
        // and cropToContent() both write style attributes themselves, which
        // would otherwise make the observer trigger itself forever; it is
        // disconnected for the duration of every fix pass to prevent that.
        var padObserver = new MutationObserver(fixAndCrop);
        function fixAndCrop() {{
            padObserver.disconnect();
            fixPadElements();
            cropToContent();
            padObserver.observe(gd, {{
                attributes: true, subtree: true, attributeFilter: ['style', 'fill', 'd', 'transform'],
            }});
        }}
        fixAndCrop();

        var allYears = {all_years_js};
        var currentYear = '{first_year}';
        var currentMode = 'species';
        var playTimer = null;

        gd.on('plotly_sliderchange', function(e) {{ currentYear = e.step.label; }});
        gd.on('plotly_animatingframe', function(e) {{
            if (e.name) {{ currentYear = e.name.split('|')[0]; }}
        }});

        function showYear(year) {{
            Plotly.animate(gd, [year + '|' + currentMode], {{
                frame: {{duration: 0, redraw: true}}, transition: {{duration: 0}}, mode: 'immediate'
            }});
        }}
        function setMode(mode) {{
            currentMode = mode;
            showYear(currentYear);
            document.getElementById('legend-species').style.display = (mode === 'species') ? 'block' : 'none';
            document.getElementById('legend-category').style.display = (mode === 'category') ? 'block' : 'none';
            document.getElementById('btn-species').style.fontWeight = (mode === 'species') ? 'bold' : 'normal';
            document.getElementById('btn-category').style.fontWeight = (mode === 'category') ? 'bold' : 'normal';
        }}
        function play() {{
            if (playTimer) return;
            playTimer = setInterval(function() {{
                var idx = (allYears.indexOf(currentYear) + 1) % allYears.length;
                currentYear = allYears[idx];
                showYear(currentYear);
            }}, 500);
        }}
        function pause() {{
            clearInterval(playTimer);
            playTimer = null;
        }}

        var ctrlDiv = document.createElement('div');
        ctrlDiv.style.textAlign = 'center';
        ctrlDiv.style.marginTop = '10px';
        ctrlDiv.style.fontFamily = 'Arial, sans-serif';
        var BTN_STYLE = 'font-family:Arial,sans-serif;font-size:14px;padding:6px 14px;';
        ctrlDiv.innerHTML =
            '<button id="btn-play" style="' + BTN_STYLE + '">&#9654; Play</button> ' +
            '<button id="btn-pause" style="' + BTN_STYLE + '">&#10074;&#10074; Pause</button>' +
            '<span style="margin:0 10px;font-size:14px;">|</span>' +
            '<span style="margin-right:8px;font-size:14px;">Color by:</span>' +
            '<button id="btn-species" style="' + BTN_STYLE + 'font-weight:bold;margin-right:6px;">N species</button>' +
            '<button id="btn-category" style="' + BTN_STYLE + '">Category (useful output / waste / ...)</button>';
        // Inserted relative to plotWrapper, not gd: gd now lives inside the
        // cropping wrapper (see cropToContent above), and that wrapper clips
        // anything placed inside it via overflow:hidden, so these controls
        // and the legend have to sit alongside the wrapper instead.
        plotWrapper.parentNode.insertBefore(ctrlDiv, plotWrapper.nextSibling);
        document.getElementById('btn-play').onclick = play;
        document.getElementById('btn-pause').onclick = pause;
        document.getElementById('btn-species').onclick = function() {{ setMode('species'); }};
        document.getElementById('btn-category').onclick = function() {{ setMode('category'); }};

        plotWrapper.parentNode.style.position = 'relative';
        var legendDiv = document.createElement('div');
        legendDiv.style.position = 'absolute';
        legendDiv.style.right = '15px';
        legendDiv.style.top = '70px';
        legendDiv.style.fontFamily = 'Arial, sans-serif';
        legendDiv.style.fontSize = '12px';
        legendDiv.style.background = 'rgba(255,255,255,0.85)';
        legendDiv.style.padding = '8px 10px';
        legendDiv.style.border = '1px solid #ccc';
        legendDiv.style.borderRadius = '4px';
        legendDiv.innerHTML =
            '<div id="legend-species" style="display:block;"><b>N species</b>{species_legend_html}</div>' +
            '<div id="legend-category" style="display:none;"><b>Category</b>{category_legend_html}</div>';
        plotWrapper.parentNode.insertBefore(legendDiv, plotWrapper.nextSibling);
        """

        fig.write_html(
            os.path.join(output_dir, filename),
            include_plotlyjs='cdn',
            default_width='100%',
            post_script=[post_script],
        )

    build_sankey_figure(df_base, "All Flows", "global_nitrogen_sankey.html")
    print(f"[SUCCESS] Komplett Sankey med låst skalering per node generert -> {os.path.join(output_dir, 'global_nitrogen_sankey.html')}")

    build_sankey_figure(df_filtered, "Fertilizer & Crude Oil Trade Hidden", "global_nitrogen_sankey_no_fertilizer.html")
    print(f"[SUCCESS] Filtrert Sankey med låst skalering per node generert -> {os.path.join(output_dir, 'global_nitrogen_sankey_no_fertilizer.html')}")

    return "global_nitrogen_sankey.html"


def extract_source_target(flow_name):
    """
    Splits a flow_code like 'AG.MM-AT.AT-Emissions-N2O' into its source and
    target sub-pool codes ('AG.MM', 'AT.AT'). Every flow_name produced by the
    calculations/ modules uses this hyphenated format.
    """
    fn = flow_name.upper().strip()
    parts = fn.split('-')
    return parts[0].strip(), parts[1].strip()


def process_and_export_mc_results(all_records):
    """
    Receives a list of dictionaries from ALL MC iterations.
    Calculates statistics, exports to Excel, and generates plots.
    """
    if not all_records:
        print("[WARNING] No records available to process.")
        return

    current_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print("\n" + "="*60)
    print("[STATISTICS] Starting statistical analysis of MC results...")
    print(f"[DEBUG] Processing run at: {current_time_str}")
    
    # 1. Convert to DataFrame and clean data types
    df_all = pd.DataFrame(all_records)
    df_all['value'] = pd.to_numeric(df_all['value'], errors='coerce')
    
    total_simulations = df_all['sim_id'].nunique()
    print(f"[STATISTICS] Detected {total_simulations} unique Monte Carlo iterations.")
    
    trimmed_chunks = []
    unique_flows = df_all['flow_name'].unique()
    
    print("[STATISTICS] Analyzing and trimming time intervals per flow...")
    
    for flow in unique_flows:
        df_flow = df_all[df_all['flow_name'] == flow]

        years_with_data = df_flow[df_flow['value'].notna()]['year'].unique().tolist()
        
        if not years_with_data:
            print(f"  [INFO] Flow '{flow}' has no data in the entire period. Skipping.")
            continue
            
        start_year = min(years_with_data)
        end_year = max(years_with_data)

        expected_years = set(range(start_year, end_year + 1))
        missing_years = expected_years - set(years_with_data)

        if missing_years:
            raise ValueError(
                f"Flow '{flow}' has missing data gaps within its own active range "
                f"({start_year}-{end_year}): {sorted(missing_years)}."
            )

        df_trimmed = df_flow[(df_flow['year'] >= start_year) & (df_flow['year'] <= end_year)].copy()
        trimmed_chunks.append(df_trimmed)

    if not trimmed_chunks:
        print("[ABORTED] No flows contained valid data after trimming intervals.")
        return
        
    df_all_trimmed = pd.concat(trimmed_chunks, ignore_index=True)

    # 2. Calculate statistics across all sim_ids
    print("[STATISTICS] Calculating medians and 95% confidence intervals...")
    summary_df = df_all_trimmed.groupby(['flow_name', 'year']).agg(
        median=('value', np.median),
        mean=('value', np.mean),
        p2_5=('value', lambda x: np.percentile(x, 2.5)),
        p97_5=('value', lambda x: np.percentile(x, 97.5)),
        std=('value', np.std),
        comment=('comment', 'first'),
        data_sources=('data_sources', 'first')
    ).reset_index()

    # 3. Calculate uncertainty metrics
    summary_df['unc_down_percent'] = np.where(
        summary_df['median'] > 0, 
        ((summary_df['median'] - summary_df['p2_5']) / summary_df['median']) * 100, 
        0.0
    )
    summary_df['unc_up_percent'] = np.where(
        summary_df['median'] > 0, 
        ((summary_df['p97_5'] - summary_df['median']) / summary_df['median']) * 100, 
        0.0
    )
    summary_df['cv_percent'] = np.where(
        summary_df['mean'] > 0, 
        (summary_df['std'] / summary_df['mean']) * 100, 
        0.0
    )

    # 4. Export to Excel
    output_dir = 'output_files'
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, 'MC_Reporting_Statistics.xlsx')
    
    summary_df_rounded = summary_df.round({
        'median': 4, 'mean': 4, 'p2_5': 4, 'p97_5': 4, 'std': 4,
        'unc_down_percent': 2, 'unc_up_percent': 2, 'cv_percent': 2
    })
    summary_df_rounded.to_excel(excel_path, index=False)
    print(f"[SUCCESS] Statistical report saved to: {excel_path}")

    # 5. GENERATE TIME-SERIES PLOTS
    plot_dir = os.path.join(output_dir, 'plots')
    
    if os.path.exists(plot_dir):
        print(f"[PLOTTING] Cleaning old directory '{plot_dir}' to avoid stale generation leaks...")
        shutil.rmtree(plot_dir)
    os.makedirs(plot_dir, exist_ok=True)

    print("[PLOTTING] Generating fresh time-series plots for each nitrogen flow...")

    for flow in summary_df['flow_name'].unique():
        df_flow = summary_df[summary_df['flow_name'] == flow].sort_values('year')
        
        plt.figure(figsize=(10, 3.6))
        
        plt.fill_between(
            df_flow['year'], 
            df_flow['p2_5'], 
            df_flow['p97_5'], 
            color='skyblue', 
            alpha=0.4, 
            label='95% Confidence Interval (MC)'
        )
        plt.plot(df_flow['year'], df_flow['p2_5'], color='steelblue', linestyle=':', linewidth=1, alpha=0.7)
        plt.plot(df_flow['year'], df_flow['p97_5'], color='steelblue', linestyle=':', linewidth=1, alpha=0.7)
        
        plt.plot(
            df_flow['year'], 
            df_flow['median'], 
            color='navy', 
            linewidth=2.5, 
            label='Median (50th percentile)'
        )
        
        flow_start = df_flow['year'].min()
        flow_end = df_flow['year'].max()
        
        plt.title(f"{flow}", fontsize=11, fontweight='bold', loc='left')        
        plt.xlim(1984, 2025)
        plt.xticks(np.arange(1984, 2026, 5))
        
        plt.ylim(bottom=0)
        current_ymax = plt.ylim()[1]
        plt.ylim(top=current_ymax * 1.15 if current_ymax > 0 else 10)
        
        plt.text(1985, plt.ylim()[1] * 0.05, 
                 f"Data Range: {flow_start}-{flow_end}", 
                 fontsize=9, style='italic', bbox=dict(facecolor='white', alpha=0.8, edgecolor='none'))
        
        plt.text(2024.5, plt.ylim()[1] * 0.05, 
                 f"Updated: {current_time_str}", 
                 fontsize=8, color='gray', ha='right', style='italic')
        
        plt.xlabel("Year", fontsize=10)
        plt.ylabel("Nitrogen Flow (kt N / year)", fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.4)
        plt.legend(loc='upper left')
        
        safe_filename = flow.replace('.', '_').replace('-', '_').replace(' ', '_') + '.png'        
        plt.savefig(os.path.join(plot_dir, safe_filename), dpi=150, bbox_inches='tight')        
        plt.close()

    # ========================================================
    # INTEGRASJON: GENERERING AV BALANSEPLOTT FOR POOLER
    # ========================================================
    print("\n[PLOTTING] Preparing mass balance datasets for pools and subpools...")
    
    df_balance_input = summary_df.copy()
    
    # Assigns source/target pools using the module-level extract_source_target helper.
    res = df_balance_input['flow_name'].apply(extract_source_target)
    df_balance_input['source'] = [r[0] for r in res]
    df_balance_input['target'] = [r[1] for r in res]
    
    # plot_pool_balance forventer kolonnene 'value' og 'uncertainty'
    df_balance_input['value'] = df_balance_input['median']
    df_balance_input['uncertainty'] = df_balance_input['std']  # Bruker standardavviket som 1σ usikkerhet
    
    # 2. Hent ut alle unike pool-koder som faktisk er til stede i dataene
    all_codes = set(df_balance_input['source'].unique()) | set(df_balance_input['target'].unique())
    all_codes.discard('Unknown')
    
    pools_to_plot = sorted(list(all_codes))
    
    # Automatisk finn og legg til overordnede hovedpooler (f.eks. 'AG' fra 'AG.MM', 'HY' fra 'HY.SW')
    main_pools = set()
    for p in pools_to_plot:
        if '.' in p:
            main_code = p.split('.')[0]
            main_pools.add(main_code)
            
    for main_code in main_pools:
        if main_code not in pools_to_plot:
            pools_to_plot.append(main_code)
            
    pools_to_plot.sort()

    print(f"[PLOTTING] Detected active pools for balance plots: {pools_to_plot}")
    
    print("[PLOTTING] Executing balance plots for active system pools...")
    for pool in pools_to_plot:
        # Lag en kopi av dataene for denne spesifikke iterasjonen
        df_temppool = df_balance_input.copy()
        
        # Hvis vi plotter en hovedpool (f.eks. 'AG' eller 'HY' uten punktum)
        if '.' not in pool:
            # Endre source/target til å bare være hovedkoden (før punktum)
            df_temppool['source_main'] = df_temppool['source'].apply(lambda x: x.split('.')[0])
            df_temppool['target_main'] = df_temppool['target'].apply(lambda x: x.split('.')[0])
            
            # Filtrer ut interne strømmer (f.eks. AG.MM til AG.SM blir intern for AG og skal bort)
            df_temppool = df_temppool[df_temppool['source_main'] != df_temppool['target_main']]
            
            # Forbered kolonner for plottefunksjonen slik at startswith(pool) fungerer
            df_temppool['source'] = df_temppool['source_main']
            df_temppool['target'] = df_temppool['target_main']
            
        plot_pool_balance(df_temppool, pool, output_dir=plot_dir)
        plot_pool_balance_interactive(df_temppool, pool, output_dir=plot_dir)

    print("\n[PLOTTING] Generating global interactive Sankey diagram across all years...")
    plot_global_sankey_interactive(df_balance_input, output_dir=plot_dir)
        
    print("\n" + "="*60)
    print("[SUCCESS] All MC iterations processed, statistics saved, and plots generated successfully.")

    return summary_df
