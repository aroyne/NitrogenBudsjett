#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import openpyxl
import warnings
from calculations.utils import read_trade_data

# Undertrykk openpyxl sin spesifikke header/footer-advarsel
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet.header_footer")

def load_all_data(selected_pools):
    
    
    """
    Sentral datalaster som sørger for at tunge I/O-operasjoner kun skjer ÉN gang.
    Konfigurert deklarativt for ekstremt enkel utvidelse av nye pools og filer.
    """
    preloaded = {}
    print(f"\n[DATA_LOADER] Kalles med selected_pools: {selected_pools}") # <-- LEGG TIL DENNE
    
    # =========================================================================
    # 1. KONFIGURASJONSKART (Legg til nye filer eller pools her!)
    # =========================================================================
    # Format: 'preloaded_nøkkel': ( {relevante_pools}, 'filbane', 'lesemetode', {ekstra_arg} )
    DATA_MAP = {
        'atm_in_out': ({'at', 'rw'}, 'data_files/atm_in_out.xlsx', 'excel', {'sheet_name': 'Ark1', 'header': None}),
        'faostat_fertilizer': ({'at'}, 'data_files/FAOSTAT_data_en_11-25-2025.csv', 'csv', {}),
        'deposition_data': ({'at', 'ag'}, 'data_files/N_per_class_period_distributed_unallocated_long.csv', 'csv', {}),
        'feed_raavarer_norsk': ({'mp'}, 'data_files/Årlig råvareforbruk.xlsx', 'excel_feed_raavarer_norsk', {}),
        'feed_raavarer_import': ({'rw'}, 'data_files/Årlig råvareforbruk.xlsx', 'excel_feed_raavarer_import', {}),        'feed_totalkalkyle': ({'rw','mp'}, 'data_files/NibioStatistics-4.xlsx', 'excel_feed_totalkalkyle', {}),
        'aqua_data': ({'hy', 'rw'}, 'data_files/A.06.002_20251111-140559.xlsx', 'excel_aquaculture', {}),
        'fao_live_animals_all': ({'ag', 'rw'}, 'data_files/FAOSTAT_data_en_11-12-2025.csv', 'csv_live_animals', {}),
        'fao_mineral_fertilizer': ({'rw'}, 'data_files/FAOSTAT_data_en_11-12-2025-2.csv', 'csv_fertilizer_import', {}),
        'hy_kyst_tilforsel': ({'hy','fs','hs'}, 'data_files/Tilførsel av nitrogen til kystområdene fordelt på kilder.xlsx', 'excel', {'sheet_name': 'Data fra Miljødirektoratet'}),
        'hy_teotil3': ({'hy','fs','hs'}, 'data_files/teotil3_n_summary.xlsx', 'openpyxl_teotil', {}),
        'hy_art_raw': ({'hy'}, 'data_files/art.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'Sheet 1'}),
        'hy_fiske_old_raw': ({'hy'}, 'data_files/fiske_1990_2000.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'Ark1'}),
        'avlop_sewage': ({'hy', 'pr'}, 'data_files/05280_20251113-113329.xlsx', 'openpyxl_sewage', {}),
        'ag_gnb': ({'ag'}, 'data_files/aei_pr_gnb__custom_18744910_spreadsheet.xlsx', 'openpyxl_gnb', {}),
        'ag_grovfor': ({'ag'}, 'grovfor_filer_samling', 'excel_grovfor', {}), # Spesialhåndtert
        'ag_crltap_raw_lines': ({'ag','ef'}, 'data_files/webdabData1863365.txt', 'text_lines', {}),
        'unfccc_ark1_raw': ({'ag'}, 'data_files/N2O_NOx_AG.xlsx', 'openpyxl_single_sheet_df', {'sheet_name': 'Ark1'}),        
        'ag_leaching_csv': ({'ag'}, 'data_files/Nr_AG--HY.csv', 'csv', {}),
        'ag_faostat_production_all': ({'ag'}, 'data_files/FAOSTAT_data_en_11-18-2025.csv', 'csv_faostat_production', {}),
        'wool_production': ({'ag'}, 'data_files/ull.xlsx', 'excel', {'skiprows': 3}),
        'ssb_sheep_numbers': ({'ag'}, 'data_files/03710_20260128-152225.xlsx', 'excel', {'skiprows': 2}),
        'fs_unfccc_emissions_raw': ({'fs'}, 'data_files/N2O_NOx_HS_FS.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'Ark1'}),
        'fs_firewood_raw': ({'fs'}, 'data_files/09702_20251120-133716.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'VedTonn'}),
        'fs_obb_grazing': ({'fs'}, 'data_files/OBB_Fylke_1970-2025.xlsx', 'openpyxl_obb_grazing', {}),
        'faostat_forestry': ({'fs'}, 'data_files/FAOSTAT_data_en_2-20-2026.csv', 'csv_forestry', {}),
        'fuel_for_industry': ({'ef'}, 'data_files/N_fuel_for_industry.csv', 'csv_ef_fuel', {}),
        'fuel_for_transport': ({'ef'}, 'data_files/N_fuel_for_transport.csv', 'csv_ef_fuel', {}),
        'fuel_for_heating': ({'ef'}, 'data_files/N_fuel_for_heating.csv', 'csv_ef_fuel', {}),
        'n2o_ec_data': ({'ef'}, 'data_files/N2O_EC.csv', 'csv', {}),
        'trade_fuels_n_content': ({'ef'}, 'data_files/N_content_fuels.xlsx', 'excel', {}),
        'ssb_energy_balance_11561': ({'ef'}, 'data_files/11561_20251113-154607.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'EnergibalansenGWh'}),
        'hs_pop_size_06913': ({'hs'}, 'data_files/06913_20251113-124117.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'Folkemengde'}),
        'hs_pop_age_groups_07459': ({'hs'}, 'data_files/07459_20251119-151434.xlsx', 'excel', {'sheet_name': 'Personer1', 'skiprows': 3, 'header': None}),
        'hs_smoking_stats_05307': ({'hs'}, 'data_files/05307_20251119-152214.xlsx', 'excel', {'sheet_name': 'Dagroyk', 'skiprows': 3, 'header': None}),
        'hs_unfccc_n2o_raw': ({'hs'}, 'data_files/N2O_NOx_HS_FS.xlsx', 'openpyxl_single_sheet', {'sheet_name': 'Ark1'}),
        'hs_luc_crltap_raw_lines': ({'hs'}, 'data_files/webdabData1863365.txt', 'text_lines', {}),
        'mp_sau_saakorn_raw': ({'mp'}, 'data_files/NibioStatistics-5.xlsx', 'excel', {'sheet_name': 'Sum innkjøpt såkorn', 'header': None}),
        'mp_oljefroe_raw': ({'mp'}, 'data_files/NibioStatistics-5.xlsx', 'excel', {'sheet_name': 'Oljefrø til modning', 'header': None}),
        'mp_erter_raw': ({'mp'}, 'data_files/NibioStatistics-5.xlsx', 'excel', {'sheet_name': 'Erter', 'header': None}),
        'mp_engfroe_raw': ({'mp'}, 'data_files/NibioStatistics-5.xlsx', 'excel', {'sheet_name': 'Sum engfrø', 'header': None}),
        'mp_rotvekst_groennsak_raw': ({'mp'}, 'data_files/NibioStatistics-5.xlsx', 'excel', {'sheet_name': 'Sum rotvekst- og grønnsakfrø', 'header': None}),
        'mildir_emissions': ({'mp'}, 'data_files/Årlig utslipp til vann - Landbasert 02-02-2026.xlsx', 'excel_mildir_emissions', {}),
        'industry_categories': ({'mp'}, 'data_files/industry_categories.xlsx', 'excel_industry_categories', {}),
        'ssb_05282': ({'hs','mp'}, 'data_files/05282_20260211-091021.xlsx', 'openpyxl_single_sheet', {'sheet_name': '05282'}),
        'ssb_06913': ({'mp'}, 'data_files/06913_20251113-124117.xlsx', 'excel_population', {}),
        'ssb_06376': ({'mp'}, 'data_files/06376_20260129-155937.xlsx', 'excel_ssb_generic', {'sheet': '06376'}),
        'ssb_10249': ({'mp'}, 'data_files/10249_20260129-155747.xlsx', 'excel_ssb_generic', {'sheet': '10249'}),
        'ssb_10514': ({'hs','mp'}, 'data_files/10514_20260211-094101.xlsx', 'openpyxl_single_sheet', {'sheet_name': '10514'}),
        'ssb_13695': ({'mp'}, 'data_files/13695_20260129-155515.xlsx', 'excel_ssb_generic', {'sheet': '13695'}),
        }

    # =========================================================================
    # 2. TUNGE SPESIALPR_LOADS (Handelsdata)
    # =========================================================================
    trade_needing_pools = {'at', 'rw', 'mp', 'pr', 'ef', 'ag'}
    if not trade_needing_pools.isdisjoint(selected_pools):
        print("[I/O] Pre-loader komplett varehandelsstatistikk...")
        try:
            df_trade_raw = read_trade_data('data_files/Tab_08801_1988_2024.csv')
            from calculations.n_params import NParameters
            df_mapping = NParameters("data_files/N_parameters.xlsx").get_trade_mapping()
            if 'konv' not in df_mapping.columns:
                df_mapping = df_mapping.reset_index()
            
            df_trade_raw['HS_code_str'] = df_trade_raw['HS_code'].astype(str).str.strip()
            v_col = 'Varenr' if 'Varenr' in df_mapping.columns else 'varenr'
            df_mapping['varenr_str'] = df_mapping[v_col].astype(str).str.strip()
            
            df_prepared_all = df_trade_raw.merge(
                df_mapping[[v_col, 'konv', 'type', 'varenr_str']],
                left_on='HS_code_str', right_on='varenr_str', how='inner'
            )
            preloaded['compressed_trade_volume'] = df_prepared_all.groupby(['year', 'impeks', 'type', 'konv'])['amount'].sum().reset_index()
        except Exception as e:
            print(f"[KRITISK FEIL] Kunne ikke pre-loade den generelle handelsdataen: {e}")

    # =========================================================================
    # 3. AUTOMATISK GENERISK INNLESING BASERT PÅ KARTET
    # =========================================================================
    for key, (pools, filepath, method, kwargs) in DATA_MAP.items():
        if pools.isdisjoint(selected_pools):
            continue  # Ingen av de valgte poolene trenger denne filen
            
        print(f"[I/O] Pre-loader data for {key} ({filepath.split('/')[-1] if '/' in filepath else filepath})...")
        try:
            if method == 'excel':
                preloaded[key] = pd.read_excel(filepath, **kwargs)
                if key == 'wool_production':
                    preloaded[key] = preloaded[key][['år', 'ull']].copy()
                elif key == 'ssb_sheep_numbers':
                    preloaded[key] = preloaded[key][['År', 'Husdyr (sau)']].copy()
                    
            elif method == 'csv':
                df = pd.read_csv(filepath, **kwargs)
                preloaded[key] = df[['Year', 'Value']].copy() if key == 'faostat_fertilizer' else df

            elif method == 'text_lines':
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    preloaded[key] = f.readlines()

            elif method == 'openpyxl_single_sheet':
                wb = openpyxl.load_workbook(filepath, data_only=True)
                preloaded[key] = pd.DataFrame(list(wb[kwargs['sheet_name']].values))

            elif method == 'openpyxl_single_sheet_df':
                wb = openpyxl.load_workbook(filepath, data_only=True)
                preloaded[key] = pd.DataFrame(list(wb[kwargs['sheet_name']].values))

            elif method == 'excel_feed_raavarer_norsk':
                df = pd.read_excel(filepath, sheet_name='Varegrupper')
                preloaded[key] = pd.DataFrame({
                    'year': df.iloc[3:28, 0].astype(int),
                    'value_carb': df.iloc[3:28, 1].astype(float),  # Kolonne B (Indeks 1) = Norsk Karbohydrat
                    'value_prot': df.iloc[3:28, 7].astype(float)   # Kolonne H (Indeks 7) = Norsk Protein
                }).reset_index(drop=True)

            elif method == 'excel_feed_raavarer_import':
                df = pd.read_excel(filepath, sheet_name='Varegrupper')
                preloaded[key] = pd.DataFrame({
                    'year': df.iloc[3:28, 0].astype(int),
                    'value_carb': df.iloc[3:28, 2].astype(float),  # Kolonne C (Indeks 2) = Importert Karbohydrat
                    'value_prot': df.iloc[3:28, 8].astype(float)   # Kolonne I (Indeks 8) = Importert Protein
                }).reset_index(drop=True)
                
            elif method == 'excel_feed_totalkalkyle':
                df = pd.read_excel(filepath, sheet_name='Sum innkjøpt kraftfôr ukorr.')
                preloaded[key] = pd.DataFrame({
                    'year': df.iloc[26:41, 0].astype(int),
                    'value': df.iloc[26:41, 1].astype(float),
                    'dom_frac': df.iloc[26:41, 4].astype(float)
                }).reset_index(drop=True)

            elif method == 'excel_aquaculture':
                df_modern = pd.read_excel(filepath, sheet_name='A.06.002', header=None)
                years_modern = df_modern.iloc[2, 2:].astype(int).tolist()
                df_cells = df_modern.iloc[4:43, 2:].replace('-', 0).astype(float)
                df_cells.columns = years_modern
                preloaded['aqua_modern'] = df_cells

                df_old = pd.read_excel('data_files/akvakultur_1984_1994.xlsx', sheet_name='Ark1', header=None)
                preloaded['aqua_old'] = pd.DataFrame({
                    'year': df_old.iloc[1:11, 0].astype(int),
                    'value': df_old.iloc[1:11, 1].astype(float)
                }).reset_index(drop=True)

            elif method == 'csv_live_animals':
                df_fao_raw = pd.read_csv(filepath)
                preloaded['fao_live_animals'] = df_fao_raw[(df_fao_raw['Element'] == 'Import quantity') & (df_fao_raw['Value'] != 0)][['Item', 'Year', 'Unit', 'Value']].copy()
                preloaded['fao_live_animals_export'] = df_fao_raw[(df_fao_raw['Element'] == 'Export quantity') & (df_fao_raw['Value'] != 0)][['Item', 'Year', 'Unit', 'Value']].copy()

            elif method == 'csv_fertilizer_import':
                df_fert = pd.read_csv(filepath)
                preloaded[key] = df_fert[(df_fert['Element'] == 'Import quantity') & (df_fert['Value'] != 0)][['Year', 'Value']].copy()

            elif method == 'openpyxl_teotil':
                wb = openpyxl.load_workbook(filepath, data_only=True)
                preloaded['hy_teotil3_to_coast'] = pd.DataFrame(list(wb['totn_to_coast'].values))
                preloaded['hy_teotil3_by_source'] = pd.DataFrame(list(wb['totn_by_source'].values))
                preloaded['hy_teotil3_retention'] = pd.DataFrame(list(wb['totn_retention'].values))

            elif method == 'openpyxl_sewage':
                wb_05280 = openpyxl.load_workbook(filepath, data_only=True)
                preloaded['hy_ssb_05280_raw'] = pd.DataFrame(list(wb_05280['Nitrogen'].values))
                wb_utslipp = openpyxl.load_workbook('data_files/utslipp_avløp.xlsx', data_only=True)
                preloaded['hy_utslipp_avlop_raw'] = pd.DataFrame(list(wb_utslipp['Ark1'].values))

            elif method == 'openpyxl_gnb':
                wb_gnb = openpyxl.load_workbook(filepath, data_only=True)
                preloaded['ag_gnb_workbook'] = wb_gnb
                preloaded['gnb_sheet30_raw'] = pd.DataFrame(list(wb_gnb['Sheet 30'].values))
                if 'Sheet 12' in wb_gnb.sheetnames:
                    preloaded['gnb_sheet12_raw'] = pd.DataFrame(list(wb_gnb['Sheet 12'].values))

            elif method == 'excel_grovfor':
                wb_13648 = openpyxl.load_workbook('data_files/13648_20251117-154625.xlsx', data_only=True)
                wb_05772 = openpyxl.load_workbook('data_files/05772_20251210-142618.xlsx', data_only=True)
                wb_old = openpyxl.load_workbook('data_files/grovfor_før_2000.xlsx', data_only=True)
                preloaded['ag_ssb_13648'] = wb_13648
                preloaded['ag_ssb_05772'] = wb_05772
                preloaded['ag_grovfor_old'] = wb_old
                preloaded['ssb_13648_raw'] = pd.DataFrame(list(wb_13648['Avling'].values))
                preloaded['ssb_05772_raw'] = pd.DataFrame(list(wb_05772['Gronfor'].values))
                preloaded['grovfor_old_raw'] = pd.DataFrame(list(wb_old['Ark1'].values))

            elif method == 'csv_faostat_production':
                df_fao = pd.read_csv(filepath)
                preloaded['ag_faostat_production'] = df_fao
                preloaded['fao_animal_production_clean'] = df_fao[(df_fao['Element'] == 'Production') & (df_fao['Value'] != 0) & (~df_fao['Item'].str.contains('hides', case=False, na=False))][['Item', 'Year', 'Value']].copy()
                preloaded['fao_hides_clean'] = df_fao[(df_fao['Element'] == 'Production') & (df_fao['Value'] != 0) & (df_fao['Item'].str.contains('hides', case=False, na=False))][['Item', 'Year', 'Value']].copy()

            elif method == 'openpyxl_obb_grazing':
                # Konverterer alle de relevante fanene fra organisert beitebruk til DataFrames med en gang
                wb_obb = openpyxl.load_workbook(filepath, data_only=True)
                preloaded['fs_obb_workbook'] = wb_obb # Beholder workbook hvis nødvendig
                
                target_sheets = [
                    'Sau1990-99', 'Sau2000-09', 'Sau2010-19', 'Sau2020-29', 
                    'Storfe og geit1993-2019', 'Storfe og geit2020-29'
                ]
                for sheet_name in target_sheets:
                    if sheet_name in wb_obb.sheetnames:
                        preloaded[f"obb_{sheet_name}_raw"] = pd.DataFrame(list(wb_obb[sheet_name].values))
                    else:
                        print(f"[ADVARSEL] Beite-fane mangler i OBB-filen: {sheet_name}")
                        
            elif method == 'csv_forestry':
                df_raw = pd.read_csv(filepath)
                # Vi tar vare på hele filen i minnet så find_industrial_round_wood kan filtrere den lynraskt
                preloaded[key] = df_raw
                
            elif method == 'csv_ef_fuel':
                df = pd.read_csv(filepath)
                preloaded[key] = df[['year', 'value']].copy()
            elif method == 'excel_ssb_generic':
                # Henter fane-navnet fra argumentene definert over
                sheet_name = kwargs.get('sheet')
                # Vi laster inn uten header for å beholde nøyaktig samme radindekser som i openpyxl
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                preloaded[key] = df
                
            elif method == 'excel_mildir_emissions':
                # Leses med vanlig header=0 slik som i koden din
                preloaded[key] = pd.read_excel(filepath, header=0)

            elif method == 'excel_industry_categories':
                preloaded[key] = pd.read_excel(filepath)
                
            elif method == 'excel_population':
                # Beholder nøyaktig samme rensing som din opprinnelige kode
                df = pd.read_excel(filepath, skiprows=2, skipfooter=42)
                df = df.set_index('Unnamed: 0')
                preloaded[key] = df

        except Exception as e:
            print(f"[KRITISK FEIL] Kunne ikke laste {key}: {e}")

    return preloaded