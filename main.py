#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 28 15:02:40 2025

@author: anja
"""

# ---------------------------------------------------------------------------
# HOW TO RUN THIS SCRIPT WITH DIFFERENT OPTIONS
#
# In Spyder (IPython console), use:
#
#   %run main.py [options...]
#
# or from a terminal:
#
#   python main.py [options...]
#
# Available options (see parse_args()):
#
#   --pool NAME          Run only selected pool(s). Can be repeated.
#                        Names: rw, ef, fs, at, hy, hs, mp, pr, ag
#                        If omitted, all pools are run in dependency order.
#
#   --no-plots           Disable the standard plots (time series + cumulative
#                        subpool plots). Does NOT automatically disable
#                        Sankey plots (see --sankey-year / --no-sankey).
#
#   --sankey-year YEAR   Create Sankey diagrams for the given YEAR
#                        (whole economy + by pool/subpool), unless
#                        --no-sankey is also given.
#
#   --no-sankey          Disable Sankey plots even if --sankey-year is given.
#
# Examples (Spyder console, using %run):
#
#   %run main.py
#       Run all pools, write results to Excel, make standard plots,
#       but no Sankey diagrams.
#
#   %run main.py --pool ef
#       Run only the 'ef' pool, write results, and make standard plots.
#
#   %run main.py --pool rw --pool ef --no-plots
#       Run only 'rw' and 'ef', write results, but do NOT create standard plots.
#
#   %run main.py --sankey-year 2021
#       Run all pools, write results, make standard plots, and also
#       create Sankey plots for year 2021.
#
#   %run main.py --pool ef --sankey-year 2020 --no-plots
#       Run only 'ef', write results, skip standard plots, and create
#       Sankey plots for year 2020.
#
#   %run main.py --sankey-year 2021 --no-sankey
#       Run all pools, write results, make standard plots, but skip
#       Sankey plots even though a year was given.
# ---------------------------------------------------------------------------

import openpyxl
from datetime import datetime
import pandas as pd
import sys
import argparse
import calculations.ef as ef
import calculations.mp as mp
import calculations.ag as ag
import calculations.at as at
import calculations.hy as hy
import calculations.rw as rw
import calculations.pr as pr
import calculations.hs as hs
import calculations.fs as fs
import plot.prepare_data_for_plotting as prep
import plot.create_line_plots as lp
import plot.sanckey_plots as sp
from plot.subpool_cumulative_time_plot import plot_subpool_balance
from pathlib import Path

REPORT_PATH = "Report.xlsx"
SHEET_NAME = "2a. Database N flows"
FIRST_DATA_ROW = 3
NAME_COL = 3
YEAR_COL = 16
VALUE_COL = 14
UNCERTAINTY_COL = 15
DATASOURCE_COL = 17
COMMENT_COL = 18

# List of calculation pools
POOLS = {
    "rw": rw,
    "ef": ef,
    "fs": fs,
    "at": at,
    "hy": hy,
    "hs": hs,
    "mp": mp,
    "pr": pr,
    "ag": ag,
}
POOL_ORDER = ["rw", "ef", "fs", "at", "hy", "hs", "mp", "pr", "ag"]
# EF needs to be after RW
# AT needs to be after EF, FS and RW
# PR needs to be after MP and HS
# AG needs to be after PR, AT, HY and RW
# MP needs to be after EF, AG, FS, HS, RW


class FlowNotFoundError(Exception):
    pass


class YearNotFoundError(Exception):
    pass


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run nitrogen flow calculations and optional plots."
    )
    parser.add_argument(
        "--pool",
        dest="pools",
        action="append",
        help=(
            "Name of a pool to run (e.g. 'rw', 'ef'). "
            "Can be given multiple times. "
            "If omitted, all pools are run."
        ),
    )
    parser.add_argument(
        "--no-plots",
        dest="plots",
        action="store_false",
        help="Disable all plotting (time series + cumulative).",
    )
    parser.add_argument(
        "--sankey-year",
        type=int,
        help="If given, create Sankey diagrams for this year.",
    )
    parser.add_argument(
        "--no-sankey",
        action="store_true",
        help="Disable Sankey plots even if --sankey-year is given.",
    )

    parser.set_defaults(plots=True)
    return parser.parse_args()


def run_pools(selected_names=None):
    """
    Run execute_calculations() for the selected pools.

    selected_names: list of names like ['rw', 'ef'], or None for all.
    Returns a list of flow dicts.
    """
    all_results = []

    if selected_names is None:
        names_to_run = POOL_ORDER
    else:
        # Keep the dependency order, but include only requested names
        names_to_run = [name for name in POOL_ORDER if name in selected_names]

        # Warn if some requested names are unknown
        unknown = [n for n in selected_names if n not in POOLS]
        if unknown:
            print("[WARNING] Unknown pools requested:", ", ".join(unknown))

    print("[INFO] Running pools:", ", ".join(names_to_run))

    for name in names_to_run:
        module = POOLS[name]
        print("[INFO] Executing pool:", name)
        results = module.execute_calculations()
        all_results.extend(results)

    return all_results


def write_nitrogen_flows(results):
    workbook = openpyxl.load_workbook(REPORT_PATH)
    sheet = workbook[SHEET_NAME]

    for flow in results:
        flow_name = flow["flow_name"]
        year = flow["year"]
        value = flow["value"]
        comment = flow["comment"]
        data_sources = flow["data_sources"]
        uncertainty = flow["uncertainty"]
        # print([flow_name, year, value])

        year_found = False
        flow_found = False

        for row in range(
            FIRST_DATA_ROW, sheet.max_row + 1
        ):  # Start from row 3 (assuming row 1 has headers and row 2 has other relevant info)
            year_in_row = sheet.cell(row=row, column=YEAR_COL).value

            if year_in_row == year:
                year_found = True

                name_in_row = sheet.cell(row=row, column=NAME_COL).value or ""
                # Now check for the flow name in the same year
                if name_in_row.strip() == flow_name.strip():
                    sheet.cell(row=row, column=VALUE_COL, value=value)
                    sheet.cell(row=row, column=UNCERTAINTY_COL, value=uncertainty/100)
                    sheet.cell(row=row, column=DATASOURCE_COL, value=data_sources)
                    sheet.cell(row=row, column=COMMENT_COL, value=comment)
                    flow_found = True
                    break  # Exit loop after update

        if year_found and not flow_found:
            raise FlowNotFoundError(
                "Flow Name '{}' not found for Year {}.".format(flow_name, year)
            )
        if not year_found:
            raise YearNotFoundError(
                "The year {} is missing from the Excel sheet!".format(year)
            )

    workbook.save(REPORT_PATH)
    print("Data written successfully!")


def run_plots(selected_pools=None):
    # Read the data for plotting (after Excel has been updated)
    df = prep.read_sheet()

    # If selected_pools is given, filter df to only flows where Pool-In or Pool-Out
    # is in the selected set (case-insensitive)
    if selected_pools:
        selected_upper = {p.upper() for p in selected_pools}
        df = df.copy()
        df["Pool-In"] = df["Pool-In"].astype(str)
        df["Pool-Out"] = df["Pool-Out"].astype(str)

        mask = (
            df["Pool-In"].str.upper().isin(selected_upper)
            | df["Pool-Out"].str.upper().isin(selected_upper)
        )
        df = df[mask].copy()

    # 1) Line plots
    lp.plot_flows_time_series(df)

    # 2) Cumulative plots per subpool
    pairs_in = (
        df[["Pool-In", "Subpool-In"]]
        .dropna(subset=["Pool-In", "Subpool-In"])
        .drop_duplicates()
        .rename(columns={"Pool-In": "Pool", "Subpool-In": "Subpool"})
    )

    pairs_out = (
        df[["Pool-Out", "Subpool-Out"]]
        .dropna(subset=["Pool-Out", "Subpool-Out"])
        .drop_duplicates()
        .rename(columns={"Pool-Out": "Pool", "Subpool-Out": "Subpool"})
    )

    pairs_all = pd.concat([pairs_in, pairs_out], ignore_index=True).drop_duplicates()

    # If selected_pools is given, only plot balances for those pools
    if selected_pools:
        selected_upper = {p.upper() for p in selected_pools}
        pairs_all = pairs_all[
            pairs_all["Pool"].astype(str).str.upper().isin(selected_upper)
        ].copy()

    for _, row in pairs_all.iterrows():
        pool = row["Pool"]
        subpool = row["Subpool"]
        plot_subpool_balance(df, pool=pool, subpool=subpool, output_dir="figurer_alle")


def plot_sanckeys(year, df, selected_pools=None):
    # create whole economy Sankey for given year:
    # first with N2 flows, then without N2 flows
    for show_N2 in (True, False):
        flows_df, center_code = sp.filter_flows_whole_economy(
            df,
            year=year,
            level="subpool",
            tol=1e-12,
            internal_label="Internal",
        )

        if not show_N2 and flows_df is not None and not flows_df.empty:
            flows_df = flows_df[
                ~(flows_df["Species"].astype(str).str.upper() == "N2")
            ].copy()

        code2desc = prep.read_definitions()
        # ensure internal whole-economy node is labeled nicely
        code2desc["NO"] = "WHOLE_ECONOMY"
        code2desc["NO.NO"] = "WHOLE_ECONOMY"

        labels, links = sp.make_sankey_split_neighbors(
            flows_df, center_code=center_code, code2desc=code2desc
        )

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if show_N2:
            title = "Whole economy " + str(year) + " — " + ts
            png_path = "whole_economy.png"
        else:
            title = "Whole economy " + str(year) + " — " + ts + ", no N2"
            png_path = "whole_economy_noN2.png"

        fig = sp.plot_sankey_by_species(
            labels, links, title=title, width=800, height=420
        )
        fig.show()
        fig.write_image(str(png_path), format="png", width=800, height=420, scale=2)
        print(f"[SAVED] {png_path}")

        # Sankey plots for all pools and subpools
        direction = "both"
        if show_N2:
            output_dir = Path("sankey_pngs_with_N2")
        else:
            output_dir = Path("sankey_pngs_no_N2")
        output_dir.mkdir(parents=True, exist_ok=True)

        # Find all pools in df
        pools = sorted(
            set(
                list(df["Pool-Out"].dropna().astype(str).unique())
                + list(df["Pool-In"].dropna().astype(str).unique())
            )
        )
        
        if selected_pools:
            selected_upper = {p.upper() for p in selected_pools}
            pools = [p for p in pools if str(p).upper() in selected_upper]

        def _process_and_save(pool, subpool=None, level="pool", overwrite=True):
            if level == "subpool":
                center_code_local = f"{pool}.{subpool}"
                fname = f"subpool_{pool}_{subpool}.png"
            else:
                center_code_local = pool
                fname = f"pool_{pool}.png"

            if show_N2:
                local_title = f"{center_code_local}_{year}" + " — " + ts
            else:
                local_title = (
                    f"{center_code_local}_{year}" + " — " + ts + ", no N2"
                )

            outpath = output_dir / fname
            if outpath.exists() and not overwrite:
                print(f"[SKIP exists] {outpath}")
                return

            flows_df_local, used_center = sp.filter_flows_by_pool_subpool(
                df,
                year=year,
                pool=pool,
                subpool=subpool,
                level=level,
                direction=direction,
            )

            if flows_df_local is None or flows_df_local.empty:
                print(f"[SKIP empty] {center_code_local} (level={level})")
                return

            if not show_N2:
                flows_df_local = flows_df_local[
                    ~(flows_df_local["Species"].astype(str).str.upper() == "N2")
                ].copy()

            labels_local, links_local = sp.make_sankey_split_neighbors(
                flows_df_local,
                center_code=used_center,
                code2desc=code2desc,
                level=level,
            )

            if links_local is None or len(links_local) == 0:
                print(f"[SKIP no links] {center_code_local} (level={level})")
                return

            fig_local = sp.plot_sankey_by_species(
                labels_local, links_local, title=local_title, width=800, height=420
            )
            if fig_local is None:
                print(f"[SKIP no fig] {center_code_local}")
                return

            png_path_local = outpath.with_suffix(".png")
            fig_local.write_image(
                str(png_path_local), format="png", width=800, height=420, scale=2
            )
            print(f"[SAVED] {png_path_local}")

        # 1) Loop through pools
        for p in pools:
            _process_and_save(pool=p, subpool=None, level="pool")

        # 2) Loop through subpools per pool
        for p in pools:
            mask = df["Pool-Out"] == p
            subs_out = (
                df.loc[mask, "Subpool-Out"].dropna().unique().tolist()
                if "Subpool-Out" in df.columns
                else []
            )
            mask = df["Pool-In"] == p
            subs_in = (
                df.loc[mask, "Subpool-In"].dropna().unique().tolist()
                if "Subpool-In" in df.columns
                else []
            )
            subpools = sorted(
                set([str(x) for x in subs_out + subs_in if pd.notna(x)])
            )

            if not subpools:
                print(f"[INFO] No subpools for pool {p}")
                continue

            for subp in subpools:
                _process_and_save(pool=p, subpool=subp, level="subpool")


def main():
    args = parse_args()
    selected_pools = args.pools
    
    # 1) Run calculations (all pools or selected ones)
    if args.pools:
        all_results = run_pools(selected_names=args.pools)
    else:
        all_results = run_pools()

    # 2) Write results to the Excel file
    try:
        write_nitrogen_flows(all_results)
    except (FlowNotFoundError, YearNotFoundError) as e:
        print("[ERROR]", e)
        sys.exit(1)

    # 3) Optional plotting (time series + cumulative subpool balances)
    if args.plots:
        run_plots(selected_pools=selected_pools)
    else:
        print("[INFO] Plots are disabled (--no-plots).")

    # 4) Sankey plots (whole economy + by pool/subpool)
    if args.sankey_year is not None and not args.no_sankey:
        print("[INFO] Creating Sankey plots for year", args.sankey_year)
        df = prep.read_sheet()
        plot_sanckeys(args.sankey_year, df, selected_pools=selected_pools)
    elif args.sankey_year is not None and args.no_sankey:
        print(
            "[INFO] --sankey-year given but --no-sankey is set; skipping Sankey plots."
        )


if __name__ == "__main__":
    main()
