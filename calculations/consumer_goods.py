#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Debug / analysis script for MP.OP-HS.HS-Consumer goods-Nmix.

Rebuilds all OP_in and OP_out components that affect the consumer goods
mass balance, year by year, and makes a stacked area plot:

- In-flows (positive, stacked upwards)
- Out-flows (negative, stacked downwards)
- Closing term from Report.xlsx
- Final mass balance (OP_in - OP_out) as a line
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

from calculations.n_params import NParameters


START_YEAR = 1984
END_YEAR = 2025
YEARS = list(range(START_YEAR, END_YEAR + 1))


def build_consumer_goods_mass_balance():
    """Recreate OP_in, OP_out and their components for consumer goods."""
    params = NParameters("data_files/N_parameters.xlsx")

    # ------------------------------------------------------------------
    # Set up containers
    # ------------------------------------------------------------------
    # OP_in / OP_out as in main script
    OP_in = pd.DataFrame(
        {"year": YEARS, "value": 0.0, "entries": 0}
    ).set_index("year")
    OP_out = pd.DataFrame(
        {"year": YEARS, "value": 0.0, "entries": 0}
    ).set_index("year")

    # Components that we want to track separately
    in_cols = [
        "in_crop_industrial_use",
        "in_non_edible_animal_products",
        "in_recycling_95_11",
        "in_recycling_12_23",
        "in_report_feedstock_wood",
        "in_other_goods_import",
    ]
    out_cols = [
        "out_other_industry_waste",
        "out_other_industry_wastewater",
        "out_emissions_NH3",
        "out_emissions_NOx",
        "out_untreated_wastewater",
        "out_other_goods_export",
    ]

    df_in = pd.DataFrame(0.0, index=YEARS, columns=in_cols)
    df_out = pd.DataFrame(0.0, index=YEARS, columns=out_cols)

    # Separate series for the “closing term” added in the Report-block
    closing_term = pd.Series(0.0, index=YEARS, name="closing_term")

    # ------------------------------------------------------------------
    # Waste N fractions (as in main script)
    # ------------------------------------------------------------------
    paper_N = params.waste_N_frac("paper")
    plastic_N = params.waste_N_frac("plastic")
    wood_N = params.waste_N_frac("wood")
    textiles_N = params.waste_N_frac("textiles")
    wet_org_N = params.waste_N_frac("wet_organic")
    park_garden_N = params.waste_N_frac("park_garden")
    other_mat_N = params.waste_N_frac("other_materials")
    hazardous_N = params.waste_N_frac("hazardous")
    contam_N = params.waste_N_frac("contaminated_masses")
    mixed_N = params.waste_N_frac("mixed_waste")
    rubber_N = params.waste_N_frac("rubber")

    # ------------------------------------------------------------------
    # 1) OP_out components that matter for consumer goods
    # ------------------------------------------------------------------
    # 1a) MP.OP-PR.SO-Other industry waste-Nmix (OP_out)
    # 1995-2011
    workbook = openpyxl.load_workbook("data_files/05282_20260211-091021.xlsx")
    sheet = workbook["05282"]
    for col in range(2, 170, 10):
        year = int(sheet.cell(row=4, column=col).value)
        if year not in YEARS:
            continue
        value = 0.0
        # Industri i col+3, byggenæring col+5, annen uspesifisert næring col+8
        # Papir
        value += float(sheet.cell(row=7, column=col + 3).value) * paper_N
        value += float(sheet.cell(row=7, column=col + 5).value) * paper_N
        value += float(sheet.cell(row=7, column=col + 8).value) * paper_N
        # Plast
        value += float(sheet.cell(row=9, column=col + 3).value) * plastic_N
        value += float(sheet.cell(row=9, column=col + 5).value) * plastic_N
        value += float(sheet.cell(row=9, column=col + 8).value) * plastic_N
        # Treavfall
        value += float(sheet.cell(row=12, column=col + 3).value) * wood_N
        value += float(sheet.cell(row=12, column=col + 5).value) * wood_N
        value += float(sheet.cell(row=12, column=col + 8).value) * wood_N
        # Tekstiler
        value += float(sheet.cell(row=13, column=col + 3).value) * textiles_N
        value += float(sheet.cell(row=13, column=col + 5).value) * textiles_N
        value += float(sheet.cell(row=13, column=col + 8).value) * textiles_N
        # Våtorganisk: bygg og annen uspesifisert
        value += float(sheet.cell(row=14, column=col + 5).value) * wet_org_N
        value += float(sheet.cell(row=14, column=col + 8).value) * wet_org_N
        # Andre materialer
        value += float(sheet.cell(row=17, column=col + 3).value) * other_mat_N
        value += float(sheet.cell(row=17, column=col + 5).value) * other_mat_N
        value += float(sheet.cell(row=17, column=col + 8).value) * other_mat_N
        # Farlig avfall
        value += float(sheet.cell(row=18, column=col + 3).value) * hazardous_N
        value += float(sheet.cell(row=18, column=col + 5).value) * hazardous_N
        value += float(sheet.cell(row=18, column=col + 8).value) * hazardous_N
        # Forurensede masser
        value += float(sheet.cell(row=19, column=col + 3).value) * contam_N
        value += float(sheet.cell(row=19, column=col + 5).value) * contam_N
        value += float(sheet.cell(row=19, column=col + 8).value) * contam_N

        OP_out.loc[year, "value"] += value
        OP_out.loc[year, "entries"] += 1
        df_out.loc[year, "out_other_industry_waste"] += value

    # 2012-2023
    workbook = openpyxl.load_workbook("data_files/10514_20260211-094101.xlsx")
    sheet = workbook["10514"]
    for col in range(2, 115, 10):
        year = int(sheet.cell(row=4, column=col).value)
        if year not in YEARS:
            continue
        value = 0.0
        # Våtorganisk: bygg + annen næring
        value += float(sheet.cell(row=7, column=col + 6).value) * wet_org_N
        value += float(sheet.cell(row=7, column=col + 8).value) * wet_org_N
        # Park- og hageavfall
        value += float(sheet.cell(row=8, column=col + 3).value) * park_garden_N
        value += float(sheet.cell(row=8, column=col + 6).value) * park_garden_N
        value += float(sheet.cell(row=8, column=col + 8).value) * park_garden_N
        # Treavfall
        value += float(sheet.cell(row=9, column=col + 3).value) * wood_N
        value += float(sheet.cell(row=9, column=col + 6).value) * wood_N
        value += float(sheet.cell(row=9, column=col + 8).value) * wood_N
        # Papir
        value += float(sheet.cell(row=11, column=col + 3).value) * paper_N
        value += float(sheet.cell(row=11, column=col + 6).value) * paper_N
        value += float(sheet.cell(row=11, column=col + 8).value) * paper_N
        # Plast
        value += float(sheet.cell(row=17, column=col + 3).value) * plastic_N
        value += float(sheet.cell(row=17, column=col + 6).value) * plastic_N
        value += float(sheet.cell(row=17, column=col + 8).value) * plastic_N
        # Tekstiler
        value += float(sheet.cell(row=19, column=col + 3).value) * textiles_N
        value += float(sheet.cell(row=19, column=col + 6).value) * textiles_N
        value += float(sheet.cell(row=19, column=col + 8).value) * textiles_N
        # Andre materialer
        value += float(sheet.cell(row=24, column=col + 3).value) * other_mat_N
        value += float(sheet.cell(row=24, column=col + 6).value) * other_mat_N
        value += float(sheet.cell(row=24, column=col + 8).value) * other_mat_N
        # Farlig avfall
        value += float(sheet.cell(row=22, column=col + 3).value) * hazardous_N
        value += float(sheet.cell(row=22, column=col + 6).value) * hazardous_N
        value += float(sheet.cell(row=22, column=col + 8).value) * hazardous_N
        # Blandet avfall
        value += float(sheet.cell(row=23, column=col + 3).value) * mixed_N
        value += float(sheet.cell(row=23, column=col + 6).value) * mixed_N
        value += float(sheet.cell(row=23, column=col + 8).value) * mixed_N
        # Forurensede masser
        value += float(sheet.cell(row=25, column=col + 3).value) * contam_N
        value += float(sheet.cell(row=25, column=col + 6).value) * contam_N
        value += float(sheet.cell(row=25, column=col + 8).value) * contam_N

        OP_out.loc[year, "value"] += value
        OP_out.loc[year, "entries"] += 1
        df_out.loc[year, "out_other_industry_waste"] += value

    # 1b) MP.OP-PR.WW-Other industry wastewater-Nmix (OP_out)
    emissions = pd.read_excel(
        "data_files/Årlig utslipp til vann - Landbasert 02-02-2026.xlsx", header=0
    )
    emissions = emissions[emissions["Komponent"] == "nitrogen, totalt"]
    categories = pd.read_excel("data_files/industry_categories.xlsx")
    categories_keep = categories[categories["kategori"] == "OP"]
    emissions_OP = emissions[emissions["AnleggNavn"].isin(categories_keep["Virksomhet"])]
    k = emissions_OP["Kommunalt nett"]
    mask_kommunalt_true = (
        (k == True)
        | k.astype(str)
        .str.strip()
        .str.lower()
        .isin(["true", "t", "1", "ja", "yes", "y"])
    )
    emissions_OP = emissions_OP[mask_kommunalt_true]
    sum_by_year = emissions_OP.groupby(["År"])["Mengde"].sum().reset_index()

    v01 = v03 = None
    for _, row in sum_by_year.iterrows():
        year = int(row["År"])
        value = row["Mengde"] / 1000.0  # t -> kt
        if year == 2001:
            v01 = value
        elif year == 2003:
            v03 = value
        if year < 2024 and value != 0 and year in YEARS:
            OP_out.loc[year, "value"] += value
            OP_out.loc[year, "entries"] += 1
            df_out.loc[year, "out_other_industry_wastewater"] += value
    # Interpolate missing year 2002
    if v01 is not None and v03 is not None and 2002 in YEARS:
        year = 2002
        value = (v01 + v03) / 2.0
        OP_out.loc[year, "value"] += value
        OP_out.loc[year, "entries"] += 1
        df_out.loc[year, "out_other_industry_wastewater"] += value

    # 1c) MP.OP-AT.AT-Emissions-NH3 (OP_out)
    data = pd.read_csv("data_files/webdabData1863365.txt", sep=";", header=None, skiprows=4)
    data = data[
        (data[2].isin(["2A", "2B", "2C", "2D", "2G", "2H"])) & (data[3] == "NH3")
    ]
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors="coerce").fillna(0)
    conv = params.get("NH3_to_N_factor")
    sums = data.groupby(1)[5].sum() * conv
    for year, val in sums.items():
        year = int(year)
        if year in YEARS:
            OP_out.loc[year, "value"] += val
            OP_out.loc[year, "entries"] += 1
            df_out.loc[year, "out_emissions_NH3"] += float(val)

    # 1d) MP.OP-AT.AT-Emissions-NOx (OP_out)
    data = pd.read_csv("data_files/webdabData1863365.txt", sep=";", header=None, skiprows=4)
    data = data[
        (data[2].isin(["2A", "2B", "2C", "2D", "2G", "2H"])) & (data[3] == "NOx")
    ]
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors="coerce").fillna(0)
    conv = params.get("NOx_to_N_factor")
    sums = data.groupby(1)[5].sum() * conv
    for year, val in sums.items():
        year = int(year)
        if year in YEARS:
            OP_out.loc[year, "value"] += val
            OP_out.loc[year, "entries"] += 1
            df_out.loc[year, "out_emissions_NOx"] += float(val)

    # 1e) MP.OP-HY.SW-Untreated wastewater-Nmix (OP_out, OP-category)
    emissions = pd.read_excel(
        "data_files/Årlig utslipp til vann - Landbasert 02-02-2026.xlsx", header=0
    )
    emissions = emissions[emissions["Komponent"] == "nitrogen, totalt"]
    categories = pd.read_excel("data_files/industry_categories.xlsx")
    categories_keep = categories[categories["kategori"] == "OP"]
    emissions_OP = emissions[emissions["AnleggNavn"].isin(categories_keep["Virksomhet"])]
    k = emissions_OP["Kommunalt nett"]
    mask_kommunalt = (
        k.isna()
        | (k == False)
        | k.astype(str)
        .str.strip()
        .str.lower()
        .isin(["false", "f", "0", "nei", "no", "n"])
    )
    emissions_OP = emissions_OP[mask_kommunalt]
    sum_by_year = emissions_OP.groupby(["År"])["Mengde"].sum().reset_index()
    for _, row in sum_by_year.iterrows():
        year = int(row["År"])
        if year >= 2024 or year not in YEARS:
            continue
        value = row["Mengde"] / 1000.0
        if value != 0:
            OP_out.loc[year, "value"] += value
            OP_out.loc[year, "entries"] += 1
            df_out.loc[year, "out_untreated_wastewater"] += value

    # ------------------------------------------------------------------
    # 2) OP_in components for consumer goods
    # ------------------------------------------------------------------
    # 2a) AG.SM-MP.OP-Crop products for industrial use-Nmix
    workbook = openpyxl.load_workbook(
        "data_files/aei_pr_gnb__custom_18744910_spreadsheet.xlsx"
    )
    sheet = workbook["Sheet 30"]
    for col in range(2, sheet.max_column + 1):
        year = sheet.cell(row=9, column=col).value
        value = sheet.cell(row=11, column=col).value
        if year is not None and value != ":":
            year = int(year)
            if year in YEARS:
                val = float(value) / 1000.0
                OP_in.loc[year, "value"] += val
                OP_in.loc[year, "entries"] += 1
                df_in.loc[year, "in_crop_industrial_use"] += val

    mean_value = OP_in["value"][OP_in["entries"] == 1].mean()
    for year in list(range(2017, 2020)) + list(range(2024, 2026)):
        if year in YEARS:
            OP_in.loc[year, "value"] += mean_value
            OP_in.loc[year, "entries"] += 1
            df_in.loc[year, "in_crop_industrial_use"] += mean_value

    # 2b) AG.MM-MP.OP-Non-edible animal products-Nmix (hides)
    data = pd.read_csv("data_files/FAOSTAT_data_en_11-18-2025.csv")
    filtered_data = data[
        (data["Element"] == "Production")
        & (data["Value"] != 0)
        & data["Item"].str.contains("hides", case=False, na=False)
    ]
    final_data = filtered_data[["Item", "Year", "Value"]].copy()
    N_content = {
        "Raw hides and skins of buffaloes": 5.2,
        "Raw hides and skins of cattle": 5.2,
        "Raw hides and skins of goats or kids": 5.2,
        "Raw hides and skins of sheep or lambs": 5.2,
    }
    final_data["N_amount"] = (
        final_data["Item"].map(N_content) * final_data["Value"] * 1e-5
    )
    total_N_per_year = final_data.groupby("Year", as_index=False)["N_amount"].sum()
    for year in range(1984, 2024):
        if year not in YEARS:
            continue
        row = total_N_per_year[total_N_per_year["Year"] == year]
        if row.empty:
            continue
        val = row["N_amount"].values[0]
        OP_in.loc[year, "value"] += val
        OP_in.loc[year, "entries"] += 1
        df_in.loc[year, "in_non_edible_animal_products"] += val

    # 2c) PR.SO-MP.OP-Recycling-Nmix
    # 1995-2011
    workbook = openpyxl.load_workbook("data_files/05281_20260121-140338.xlsx")
    sheet = workbook["Avfall"]
    for col in range(4, 21):
        year = int(sheet.cell(row=3, column=col).value)
        if year not in YEARS:
            continue
        val = 0.0
        val += float(sheet.cell(row=19, column=col).value) * paper_N
        val += float(sheet.cell(row=21, column=col).value) * plastic_N
        val += float(sheet.cell(row=24, column=col).value) * wood_N
        val += float(sheet.cell(row=25, column=col).value) * textiles_N
        val += float(sheet.cell(row=29, column=col).value) * other_mat_N
        val += float(sheet.cell(row=30, column=col).value) * hazardous_N
        val += float(sheet.cell(row=31, column=col).value) * contam_N

        OP_in.loc[year, "value"] += val
        OP_in.loc[year, "entries"] += 1
        df_in.loc[year, "in_recycling_95_11"] += val

    # 2012-2023
    workbook = openpyxl.load_workbook("data_files/10513_20260212-104227.xlsx")
    sheet = workbook["10513"]
    for col in range(2, 110, 9):
        year = int(sheet.cell(row=4, column=col).value)
        if year not in YEARS:
            continue
        val = 0.0
        val += float(sheet.cell(row=9, column=col + 1).value) * wood_N
        val += float(sheet.cell(row=11, column=col + 1).value) * paper_N
        val += float(sheet.cell(row=17, column=col + 1).value) * plastic_N
        val += float(sheet.cell(row=18, column=col + 1).value) * rubber_N
        val += float(sheet.cell(row=19, column=col + 1).value) * textiles_N
        val += float(sheet.cell(row=22, column=col + 1).value) * hazardous_N
        val += float(sheet.cell(row=23, column=col + 1).value) * mixed_N
        val += float(sheet.cell(row=24, column=col + 1).value) * other_mat_N
        val += float(sheet.cell(row=25, column=col + 1).value) * contam_N

        OP_in.loc[year, "value"] += val
        OP_in.loc[year, "entries"] += 1
        df_in.loc[year, "in_recycling_12_23"] += val

    # 2d) Data from Report: EF.EC-MP.OP-Fuel used as feedstock & FS.FO-MP.OP-Industrial round wood
    data = pd.read_excel(
        "Report.xlsx",
        sheet_name="2a. Database N flows",
        header=None,
        skiprows=2,
    )
    flow_codes = [
        "EF.EC-MP.OP-Fuel used as feedstock-Nmix",
        "FS.FO-MP.OP-Industrial round wood-Nmix",
    ]
    data[2] = data[2].astype(str).str.strip()
    filtered_data = data[data[2].isin(flow_codes)].copy()
    filtered_data = filtered_data.rename(
        columns={2: "FlowCode", 15: "Year", 13: "Value", 17: "Comment"}
    )
    filtered_data["Comment_norm"] = (
        filtered_data["Comment"].astype(str).str.strip().str.lower()
    )
    ok_flag = (
        filtered_data.groupby("Year")["Comment_norm"]
        .agg(lambda s: (s == "ok").all())
        .rename("all_ok")
        .reset_index()
    )
    grouped_data = (
        filtered_data.groupby(["Year", "FlowCode"])["Value"].sum().reset_index()
    )
    result = grouped_data.pivot(
        index="Year", columns="FlowCode", values="Value"
    ).fillna(0.0)
    result["FinalValue"] = result.sum(axis=1, numeric_only=True)
    result = result.reset_index().merge(ok_flag, on="Year", how="left")
    result["all_ok"] = result["all_ok"].fillna(False)

    for _, row in result.iterrows():
        year = int(row["Year"])
        if year not in YEARS:
            continue
        if row["all_ok"]:
            # add direct fuel/wood to OP_in as separate component
            direct_val = row["FinalValue"]
            OP_in.loc[year, "value"] += direct_val
            OP_in.loc[year, "entries"] += 1
            df_in.loc[year, "in_report_feedstock_wood"] += direct_val

            # ORIGINAL CODE: adds an extra closing term = FinalValue + OP_in - OP_out
            closing_val = row["FinalValue"] + OP_in.loc[year, "value"] - OP_out.loc[year, "value"]
            OP_in.loc[year, "value"] += closing_val
            OP_in.loc[year, "entries"] += 1
            closing_term.loc[year] += closing_val

    # 2e) Other goods import (without NH3)
    types_to_keep = [
        "blomster",
        "kjemikalier",
        "såpe",
        "industrielt protein",
        "skinn",
        "tre",
        "silke",
        "ull",
        "bomull",
        "nylon",
        "tekstil",
    ]
    trade_data = pd.read_csv("data_files/Tab_08801_1988_2024.csv", sep=";", header=None)
    trade_data.columns = [
        "year",
        "impeks",
        "HS_code",
        "country",
        "value_code",
        "amount",
        "value_2",
        "value_3",
    ]
    hs_codes_data = pd.read_excel("data_files/Varekat_web_nor_2025_ny.xlsx")
    relevant_hs_codes = hs_codes_data[hs_codes_data["type"].isin(types_to_keep)].copy()
    relevant_hs_codes = relevant_hs_codes.rename(columns={"Varenr": "HS_code"})
    filtered_trade_data = trade_data[
        (trade_data["HS_code"].isin(relevant_hs_codes["HS_code"]))
        & (trade_data["impeks"].isin([1]))
    ]
    sum_by_year = (
        filtered_trade_data.groupby(["HS_code", "year"])["amount"]
        .sum()
        .reset_index()
    )
    N_content = {row["HS_code"]: row["konv"] for _, row in relevant_hs_codes.iterrows()}
    sum_by_year["N_content"] = sum_by_year["HS_code"].map(N_content)
    sum_by_year["N_amount"] = (
        sum_by_year["amount"] * sum_by_year["N_content"] / 1e6
    )  # kg -> ktN
    aggregated_data = sum_by_year.groupby("year", as_index=False)["N_amount"].sum()
    for _, row in aggregated_data.iterrows():
        year = int(row["year"])
        if year not in YEARS:
            continue
        val = row["N_amount"]
        OP_in.loc[year, "value"] += val
        OP_in.loc[year, "entries"] += 1
        df_in.loc[year, "in_other_goods_import"] += val

    # 2f) Other goods export (without NH3) (OP_out)
    filtered_trade_data = trade_data[
        (trade_data["HS_code"].isin(relevant_hs_codes["HS_code"]))
        & (trade_data["impeks"].isin([2]))
    ]
    sum_by_year = (
        filtered_trade_data.groupby(["HS_code", "year"])["amount"]
        .sum()
        .reset_index()
    )
    sum_by_year["N_content"] = sum_by_year["HS_code"].map(N_content)
    sum_by_year["N_amount"] = (
        sum_by_year["amount"] * sum_by_year["N_content"] / 1e6
    )  # kg -> ktN
    aggregated_data = sum_by_year.groupby("year", as_index=False)["N_amount"].sum()
    for _, row in aggregated_data.iterrows():
        year = int(row["year"])
        if year not in YEARS:
            continue
        val = row["N_amount"]
        OP_out.loc[year, "value"] += val
        OP_out.loc[year, "entries"] += 1
        df_out.loc[year, "out_other_goods_export"] += val

    # ------------------------------------------------------------------
    # Final balance as used in the main script
    # (after closing term has been added)
    # ------------------------------------------------------------------
    final_balance = OP_in["value"] - OP_out["value"]
    final_balance.name = "final_balance"

    return df_in, df_out, closing_term, final_balance


def make_plot(df_in, df_out, closing_term, final_balance):
    """Make stacked area plot of all components, closing term and balance."""
    years = np.array(df_in.index, dtype=int)

    # --- Quick diagnostics in the terminal ---
    print("IN components (min..max):")
    print(df_in.sum(axis=1).describe())
    print("\nOUT components (min..max):")
    print(df_out.sum(axis=1).describe())
    print("\nFinal balance (min..max):")
    print(final_balance.describe())

    # Total magnitudes per year for axis scaling
    total_in = df_in.sum(axis=1)
    total_out = df_out.sum(axis=1)
    ymax = float(
        max(
            total_in.max(),
            total_out.max(),
            abs(final_balance).max(),
            abs(closing_term).max(),
        )
    )
    if ymax == 0:
        ymax = 1.0  # avoid degenerate axis if everything is zero

    fig, ax = plt.subplots(figsize=(12, 8))

    # ------------------------------
    # 1) Stacked IN flows (positive)
    # ------------------------------
    in_labels = df_in.columns.to_list()
    in_arrays = [df_in[col].values for col in in_labels]

    ax.stackplot(
        years,
        in_arrays,
        labels=[lbl + " (in)" for lbl in in_labels],
        alpha=0.6,
    )

    # -------------------------------------------
    # 2) Stacked OUT flows (magnitudes, negative)
    # -------------------------------------------
    out_labels = df_out.columns.to_list()
    out_arrays = [df_out[col].values for col in out_labels]

    # stack the magnitudes, then plot them below zero
    out_stack = ax.stackplot(
        years,
        out_arrays,
        alpha=0.6,
    )
    # shift the out-stack down by plotting a white line at zero afterwards
    # and interpret the stack visually as negative.
    # To make it explicit, we draw outlines at -total_out:
    ax.plot(years, -total_out.values, color="black", linewidth=0.5)

    # Manually add labels for out-flows (legend only)
    for lbl in out_labels:
        ax.plot([], [], color="grey", alpha=0.6, label=lbl + " (out)")

    # ------------------------------
    # 3) Closing term and balance
    # ------------------------------
    ax.plot(
        years,
        closing_term,
        color="black",
        linestyle="--",
        linewidth=1.5,
        label="closing term (added in Report-block)",
    )
    ax.plot(
        years,
        final_balance,
        color="red",
        linewidth=2,
        label="final balance OP_in - OP_out",
    )

    # Axis formatting
    ax.axhline(0, color="grey", linewidth=0.8)
    ax.set_ylim(-1.1 * ymax, 1.1 * ymax)
    ax.set_xlabel("Year")
    ax.set_ylabel("kt N/year")
    ax.set_title("Consumer goods mass balance (MP.OP-HS.HS-Consumer goods-Nmix)")

    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize="small")
    fig.tight_layout()
    fig.savefig("consumer_goods_mass_balance_debug.png", dpi=300)
    plt.show()

    print("\nYears with largest |final_balance|:")
    print(final_balance.abs().sort_values(ascending=False).head(10))


def main():
    df_in, df_out, closing_term, final_balance = build_consumer_goods_mass_balance()
    make_plot(df_in, df_out, closing_term, final_balance)


if __name__ == "__main__":
    main()


def main():
    df_in, df_out, closing_term, final_balance = build_consumer_goods_mass_balance()
    make_plot(df_in, df_out, closing_term, final_balance)


if __name__ == "__main__":
    main()
