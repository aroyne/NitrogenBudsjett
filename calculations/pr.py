#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 13 10:25:28 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl
from calculations.n_params import NParameters
from calculations.shared_flow_calculations import (
    get_waste_frac,
    find_export_for_recycling,
    find_export_for_reuse,
    find_household_waste,
    find_landfill_emissions_to_water,
    find_other_industry_waste,
    find_recycling,
    find_sewage_sludge_biogas,
    find_solid_waste_export)
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
    read_trade_data,
    # find_trade_data
)

expected_years = EXPECTED_YEARS

params = NParameters("data_files/N_parameters.xlsx")
dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
waste_fracs = params.get_table('waste_fractions').set_index('waste_category')
trade_data = read_trade_data('data_files/Tab_08801_1988_2024.csv')


def execute_calculations():
    results = []
    trade_params = params.get_trade_params()      # index: param_id
    trade_mapping = params.get_trade_mapping()    # columns: type, konv, Varenr, ...

    

    years = list(range(1984, 2026))  # 1984..2025 inclusive
    ww_bg = pd.DataFrame({
        'year': years,
        'value': 0.0,               # float zeros; use 0 if you want ints
        'entries': 0    # need to count the right number of entries that year if comment to be 'ok'
    })
    ww_bg.set_index('year', inplace=True)
    
    # _add_sewage_sludge_to_biogas(results, dataset_unc, ww_bg)
    _add_waste_to_energy(results, dataset_unc, ww_bg)
    _add_recycling(results, dataset_unc, trade_mapping, trade_params)
    _add_ag_biologically_treated_organic_waste(results, dataset_unc)
    _add_hs_biologically_treated_organic_waste(results, dataset_unc)
    _add_wastewater_from_landfills(results, dataset_unc, ww_bg)
    _add_biofuels_production_wastewater(results, dataset_unc, ww_bg)
    _add_so_NOx_emissions(results, dataset_unc)
    _add_so_NH3_emissions(results, dataset_unc)
    _add_so_N2O_emissions(results, dataset_unc)
    _add_so_leaching(results, dataset_unc)
    _add_solid_waste_export(results, dataset_unc, trade_data, trade_mapping, trade_params)
    _add_export_for_recycling(results, dataset_unc, trade_data, trade_mapping, trade_params)
    _add_export_for_reuse(results, dataset_unc, trade_data, trade_mapping, trade_params)
    _add_ag_sewage_sludge_fertilizer(results, dataset_unc)
    _add_hs_sewage_sludge_fertilizer(results, dataset_unc)
    _add_sewage_sludge_landfill(results, dataset_unc)
    _add_ww_NH3_emissions(results, dataset_unc)
    _add_ww_NOx_emissions(results, dataset_unc)
    _add_ww_N2O_emissions(results, dataset_unc)
    _add_ww_N2_emissions(results, dataset_unc)
    _add_treated_ww_discharge(results, dataset_unc)


    # paper_N,  u_paper  = get_waste_frac("paper")
    # plastic_N, u_plastic = get_waste_frac("plastic")
    # wood_N,   u_wood   = get_waste_frac("wood")
    # textile_N, u_text  = get_waste_frac("textiles")
    # wet_N,    u_wet    = get_waste_frac("wet_organic")
    # sludge_N, u_sludge = get_waste_frac("sludge")
    # other_N,  u_other  = get_waste_frac("other_materials")
    # haz_N,    u_haz    = get_waste_frac("hazardous")
    # contam_N, u_contam = get_waste_frac("contaminated_masses")
    # mixed_N,  u_mixed  = get_waste_frac("mixed_waste")
    # rubber_N, u_rubber = get_waste_frac("rubber")
    # park_N,   u_park   = get_waste_frac("park_garden")
    


    return results  # Returns a list of flow records

# def _add_sewage_sludge_to_biogas(results, dataset_unc, ww_bg):
#     flow_code = 'PR.WW-PR.SO-Sewage sludge to biogas-Nmix'
#     collected_years = set()
#     # using data from SSB
#     data_sources = 'SSB table 12359'
#     comment = 'ok'
#     year_values, uncertainty = find_sewage_sludge_biogas(dataset_unc)
#     for year, value in year_values.items():
#         collected_years.add(year)
#         ww_bg.loc[year,'value'] += value
#         ww_bg.loc[year,'entries'] += 1
#         results.append({
#             'flow_name': flow_code,
#             'year': year,
#             'value': value, 
#             'comment': comment,
#             'data_sources': data_sources,
#             'uncertainty': uncertainty
#         })
#     missing_years = expected_years - collected_years
#     report_missing_years(flow_code, missing_years, results)



def _add_waste_to_energy(results, dataset_unc, ww_bg):
    flow_code = 'PR.SO-EF.EC-Waste to energy-Nmix'
    collected_years = set()
    # waste to incineration and energy recovery 
    # representative/conservative N-content uncertainty across all waste fractions used
    u_05281 = get_uncertainty(dataset_unc, '05281')
    u_10513 = get_uncertainty(dataset_unc, '10513')    
    paper_N,  u_paper  = get_waste_frac("paper")
    plastic_N, u_plastic = get_waste_frac("plastic")
    wood_N,   u_wood   = get_waste_frac("wood")
    textile_N, u_text  = get_waste_frac("textiles")
    wet_N,    u_wet    = get_waste_frac("wet_organic")
    sludge_N, u_sludge = get_waste_frac("sludge")
    other_N,  u_other  = get_waste_frac("other_materials")
    haz_N,    u_haz    = get_waste_frac("hazardous")
    contam_N, u_contam = get_waste_frac("contaminated_masses")
    mixed_N,  u_mixed  = get_waste_frac("mixed_waste")
    rubber_N, u_rubber = get_waste_frac("rubber")
    park_N,   u_park   = get_waste_frac("park_garden")
    u_waste_max = max(
        u_paper, u_plastic, u_wood, u_text, u_wet, u_sludge,
        u_other, u_haz, u_contam, u_rubber, u_mixed, u_park)    
    uncertainty_05281 = combine_uncertainties_percent(u_05281, u_waste_max)
    uncertainty_10513 = combine_uncertainties_percent(u_10513, u_waste_max)
    data_sources = 'SSB'
    comment = 'ok'
    # 1995-2011: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    workbook = openpyxl.load_workbook('data_files/05281_20260121-140338.xlsx')
    sheet = workbook['Avfall']
    for col in range(4,21):  
        value = 0
        year = int(sheet.cell(row=3, column=col).value)  
        collected_years.add(year)
        # Papir med og uten energiutnyttelse
        value += float(sheet.cell(row=61, column=col).value)*paper_N
        value += float(sheet.cell(row=89, column=col).value)*paper_N
        # Plast
        value += float(sheet.cell(row=63, column=col).value)*plastic_N
        value += float(sheet.cell(row=91, column=col).value)*plastic_N
        # Treavfall
        value += float(sheet.cell(row=66, column=col).value)*wood_N
        value += float(sheet.cell(row=94, column=col).value)*wood_N
        # Tekstiler
        value += float(sheet.cell(row=67, column=col).value)*textile_N
        value += float(sheet.cell(row=95, column=col).value)*textile_N
        # Våtorganisk avfall
        value += float(sheet.cell(row=68, column=col).value)*wet_N
        value += float(sheet.cell(row=96, column=col).value)*wet_N
        # Slam
        value += float(sheet.cell(row=70, column=col).value)*sludge_N
        value += float(sheet.cell(row=98, column=col).value)*sludge_N
        # Andre materialer
        value += float(sheet.cell(row=71, column=col).value)*other_N
        value += float(sheet.cell(row=99, column=col).value)*other_N
        # Farlig avfall
        value += float(sheet.cell(row=72, column=col).value)*haz_N
        value += float(sheet.cell(row=100, column=col).value)*haz_N
        # Forurensede masser
        value += float(sheet.cell(row=73, column=col).value)*contam_N
        value += float(sheet.cell(row=101, column=col).value)*contam_N
        comment = 'ok'
        data_sources = 'SSB'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_05281        
        })
    # 2012-2023: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    workbook = openpyxl.load_workbook('data_files/10513_20260212-104227.xlsx')
    sheet = workbook['10513']
    for col in range(2,110,9):  
        value = 0
        year = int(sheet.cell(row=4, column=col).value)  
        collected_years.add(year)
        # Våtorganisk avfall
        value += float(sheet.cell(row=7, column=col+5).value)*wet_N
        # Park- og hageavfall
        value += float(sheet.cell(row=8, column=col+5).value)*park_N        
        # Treavfall
        value += float(sheet.cell(row=9, column=col+5).value)*wood_N
        # Slam
        value += float(sheet.cell(row=10, column=col+5).value)*sludge_N
        # Papir
        value += float(sheet.cell(row=11, column=col+5).value)*paper_N
        # Plast
        value += float(sheet.cell(row=17, column=col+5).value)*plastic_N
        # Gummi
        value += float(sheet.cell(row=18, column=col+5).value)*rubber_N # Schäppi tabell 23
        # Tekstiler
        value += float(sheet.cell(row=19, column=col+5).value)*textile_N
        # Farlig avfall
        value += float(sheet.cell(row=22, column=col+5).value)*haz_N
        # Blandet avfall
        value += float(sheet.cell(row=23, column=col+5).value)*mixed_N
        # Andre materialer
        value += float(sheet.cell(row=24, column=col+5).value)*other_N
        # Forurensede masser
        value += float(sheet.cell(row=25, column=col+5).value)*contam_N
        comment = 'ok'
        data_sources = 'SSB'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_10513        
        })
    # 1990-1994
    household_waste, unc_waste = find_household_waste(dataset_unc)
    industry_waste, industry_waste_unc = find_other_industry_waste(dataset_unc,params)
    workbook = openpyxl.load_workbook('data_files/kommunalt_avfall_1985_1995.xlsx')
    sheet = workbook['forbrenning og gjenvinning'] # 
    inc_frac_1985 = float(sheet.cell(row=2, column=2).value)/100
    inc_frac_1992 = float(sheet.cell(row=3, column=2).value)/100
    change_per_year = (inc_frac_1992-inc_frac_1985)/7
    r = 3
    for year in range(1990,1995):
        collected_years.add(year)
        waste = household_waste[year] + industry_waste[year]
        if year < 1992:
            inc_frac = inc_frac_1985+change_per_year*(year-1985)
            comment = 'extrapolated'
        else:
            inc_frac = float(sheet.cell(row=r, column=2).value)/100
            comment = 'ok'
            r += 1
        value = waste*inc_frac
        data_sources = 'SSB'
        uncertainty = combine_uncertainties_percent(unc_waste[year],industry_waste_unc[year])
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
    
def _add_recycling(results, dataset_unc, trade_mapping, trade_params):
    flow_code = 'PR.SO-MP.OP-Recycling-Nmix'
    collected_years = set()
    year_values, uncertainty_05281, uncertainty_10513, uncertainty_old = find_recycling(dataset_unc, trade_mapping, trade_params)
    comment = 'ok'
    data_sources = 'SSB'
    # 1995-2011: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    for year, value in year_values.items():
        collected_years.add(year)
        if year < 1995:
            uncertainty = uncertainty_old
        elif year <= 2011:
            uncertainty = uncertainty_05281
        else:
            uncertainty = uncertainty_10513
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty       
        })    
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_ag_biologically_treated_organic_waste(results, dataset_unc):
    flow_code = 'PR.SO-AG.SM-Biologically treated organic waste-Nmix'
    collected_years = set()
    # From 2021: use data from Biogass Norge
    workbook = openpyxl.load_workbook('data_files/biogass_tall.xlsx')
    sheet = workbook['biorest']
    data_sources = 'Biogass Norge'
    comment = 'ok'
    uncertainty   = get_uncertainty(dataset_unc, 'Biogass_Norge')
    for col in range (3,7):
        year = int(sheet.cell(row=7, column=col).value)
        collected_years.add(year)
        value = float(sheet.cell(row=32, column=col).value)/1000
        if year == 2021:
            value_2021 = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    # For 2018-2020: scale with values from SSB table 12818 
    data_sources = 'SSB'
    workbook = openpyxl.load_workbook('data_files/12818_20260526-110921.xlsx')
    sheet = workbook['12818']
    comment = 'ok'
    u_12359   = get_uncertainty(dataset_unc, '12818')
    uncertainty = combine_uncertainties_percent(uncertainty,u_12359)
    tonnes_2021 = float(sheet.cell(row=6, column=5).value)
    for col in range (2,5):
        year = int(sheet.cell(row=4, column=col).value)
        collected_years.add(year)
        value = float(sheet.cell(row=6, column=col).value)*value_2021/tonnes_2021
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    # 2012-2017: find input flows
    # sewage_sludge, unc_sew = find_sewage_sludge_biogas(dataset_unc)
    # manure, unc_man = find_manure_for_biofuel_production(params, dataset_unc)
    # year_values, uncertainty_10514, uncertainty_05282 = find_food_industry_waste(dataset_unc)
    # using numbers from avfallsregnskapet on waste sent to biogas production, with assumptions on loss and agricultural use
    # no data before 2012 (eller kan det være at "slam" og "våtofganisk avfall" fra avfallsregnskapet til energiutnyttelse faktisk er til biogassproduksjon?)
    data_sources = 'SSB, Landbruksdirektoratet, Biogass Norge'
    u_biogass = get_uncertainty(dataset_unc, 'Biogass')
    u_10513   = get_uncertainty(dataset_unc, '10513')
    u_12359   = get_uncertainty(dataset_unc, '12359')   
    paper_N,  u_paper  = get_waste_frac("paper")
    plastic_N, u_plastic = get_waste_frac("plastic")
    wood_N,   u_wood   = get_waste_frac("wood")
    textile_N, u_text  = get_waste_frac("textiles")
    wet_N,    u_wet    = get_waste_frac("wet_organic")
    sludge_N, u_sludge = get_waste_frac("sludge")
    other_N,  u_other  = get_waste_frac("other_materials")
    haz_N,    u_haz    = get_waste_frac("hazardous")
    contam_N, u_contam = get_waste_frac("contaminated_masses")
    mixed_N,  u_mixed  = get_waste_frac("mixed_waste")
    rubber_N, u_rubber = get_waste_frac("rubber")
    park_N,   u_park   = get_waste_frac("park_garden")
    # manure N fraction
    manure_N, u_manureN = params.get_global_param_with_uncertainty("manure_N_frac")
    # fish waste N fraction
    fish_N, u_fishN = params.get_global_param_with_uncertainty("fish_waste_N_frac")
    # digestate process parameters
    loss_factor, u_loss = params.get_global_param_with_uncertainty("digestate_loss_fraction")
    fraction_agr, u_fracA = params.get_global_param_with_uncertainty("digestate_fraction_to_agriculture")    
    # waste mix in biogas input
    u_waste_max = max(
        u_wet, u_park, u_wood, u_sludge, u_paper, u_plastic,
        u_rubber, u_text, u_haz, u_mixed, u_other, u_contam)   
    uncertainty = combine_uncertainties_percent(
        u_biogass, u_10513, u_12359,
        u_waste_max,
        u_manureN, u_fishN, u_loss, u_fracA)
    # manure for biogas production, from Landbruksdirektoratet
    manure_values, manure_unc = find_sewage_sludge_biogas(dataset_unc)
    # # fiskeslam
    # workbook = openpyxl.load_workbook('data_files/12359_20251211-153434.xlsx')
    # sheet = workbook['Mengde'] # unit: 1000 ton waste
    # # N_content: assume typical value of 10 gN/kg (Schäppi2025Ann p 130) gN/kg10 = kgN/t = tN/kt så dele på 1000
    # N_content = fish_N
    # year_values_fish = {}
    # for col in range(4, sheet.max_column + 1):  
    #     year = int(sheet.cell(row=3, column=col).value)  
    #     value = float(sheet.cell(row=29, column=col).value)*N_content # Fiskeavfall og anna maritimt
    #     year_values_fish[year] = value
    # 2012-2023: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    workbook = openpyxl.load_workbook('data_files/10513_20260212-104227.xlsx')
    sheet = workbook['10513']
    for col in range(2,50,9):  
        value = 0
        year = int(sheet.cell(row=4, column=col).value)  
        collected_years.add(year)
        # Våtorganisk avfall
        value += float(sheet.cell(row=7, column=col+2).value)*wet_N
        # Park- og hageavfall
        value += float(sheet.cell(row=8, column=col+2).value)*park_N
        # Treavfall
        value += float(sheet.cell(row=9, column=col+2).value)*wood_N
        # Slam
        value += float(sheet.cell(row=10, column=col+2).value)*sludge_N
        # legg til husdyrmøkk (antar neglisjerbar før 2013)
        if year > 2016:
            value += manure_values[year]  
        comment = 'ok'
        data_sources = 'SSB'
        value *= (1-loss_factor)*fraction_agr
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    for year in range (1984,2012):
        collected_years.add(year)
        value = 0
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_wastewater_from_landfills(results, dataset_unc, ww_bg):
    flow_code = 'PR.SO-PR.WW-Wastewater from landfills-Nmix'
    collected_years = set()
    u_ns = get_uncertainty(dataset_unc, 'norskeutslipp')
    uncertainty = u_ns    
    comment = 'ok'
    emissions_landfills = find_landfill_emissions_to_water()
    mean_connected = emissions_landfills["N_tilkoblet"].mean()
    for year in range (1990,2011):
        collected_years.add(year)
        value = mean_connected/1000 # tN -> ktN
        data_sources = 'extrapolated'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    for year in range (2011,2026):
        collected_years.add(year)
        value = emissions_landfills.loc[emissions_landfills["År"] == year, "N_tilkoblet"].sum()/1000 # tN -> ktN
        data_sources = 'norskeutslipp.no'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_hs_biologically_treated_organic_waste(results, dataset_unc):
    flow_code = 'PR.SO-HS.HS-Biologically treated organic waste-Nmix'
    collected_years = set()
    # 2018-2024:  table 12818 
    data_sources = 'SSB'
    workbook = openpyxl.load_workbook('data_files/12818_20260526-110921.xlsx')
    sheet = workbook['12818']
    comment = 'ok'
    u_12818   = get_uncertainty(dataset_unc, '12818')
    compost_N, u_compost = get_waste_frac("compost_old")
    uncertainty = combine_uncertainties_percent(u_compost,u_12818)
    for col in range (2,9):
        year = int(sheet.cell(row=4, column=col).value)
        collected_years.add(year)
        value = float(sheet.cell(row=7, column=col).value)*compost_N
        value += float(sheet.cell(row=8, column=col).value)*compost_N
        if year == 2018:
            value_12818_2018 = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    # # 2012-2017:SSB table 10513
    u_10513   = get_uncertainty(dataset_unc, '10513')
    wet_N,    u_wet    = get_waste_frac("wet_organic")
    sludge_N, u_sludge = get_waste_frac("sludge")
    park_N,   u_park   = get_waste_frac("park_garden")
    # 2012-2017: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    comment = 'ok'
    data_sources = 'SSB'
    workbook = openpyxl.load_workbook('data_files/10513_20260212-104227.xlsx')
    sheet = workbook['10513']
    N_loss, u_loss_comp = get_waste_frac("compost_N_loss")
    uncertainty = combine_uncertainties_percent(u_wet,u_sludge,u_park,u_loss_comp,u_10513)
    value_10513_2018 = float(sheet.cell(row=7, column=56+2).value)*wet_N+ float(sheet.cell(row=8, column=56+2).value)*park_N+float(sheet.cell(row=10, column=56+2).value)*sludge_N
    for col in range(2,50,9):  
        value = 0
        year = int(sheet.cell(row=4, column=col).value)  
        collected_years.add(year)
        # Våtorganisk avfall
        value += float(sheet.cell(row=7, column=col+2).value)*wet_N
        # Park- og hageavfall
        value += float(sheet.cell(row=8, column=col+2).value)*park_N
        # Slam
        value += float(sheet.cell(row=10, column=col+2).value)*sludge_N
        value = value/value_10513_2018*value_12818_2018
        value *= (1-N_loss)
        if year == 2012:
            value_2012 = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    # 1990-2011: extrapolate 2012 value
    data_sources = 'extrapolated'
    value = value_2012
    for year in range (1984,2012):
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_biofuels_production_wastewater(results, dataset_unc, ww_bg):
    flow_code = 'PR.SO-PR.WW-Biofuels production wastewater-Nmix'
    collected_years = set()
    # using numbers from avfallsregnskapet on waste sent to biogas production, with assumptions on loss and agricultural use
    # no data before 2012 (eller kan det være at "slam" og "våtofganisk avfall" fra avfallsregnskapet til energiutnyttelse faktisk er til biogassproduksjon?)
    data_sources = 'SSB, Landbruksdirektoratet, Biogass Norge'
    comment = 'ok'
    u_biogass = get_uncertainty(dataset_unc, 'Biogass')
    u_10513   = get_uncertainty(dataset_unc, '10513')
    u_12359   = get_uncertainty(dataset_unc, '12359')
    paper_N,  u_paper  = get_waste_frac("paper")
    plastic_N, u_plastic = get_waste_frac("plastic")
    wood_N,   u_wood   = get_waste_frac("wood")
    textile_N, u_text  = get_waste_frac("textiles")
    wet_N,    u_wet    = get_waste_frac("wet_organic")
    sludge_N, u_sludge = get_waste_frac("sludge")
    other_N,  u_other  = get_waste_frac("other_materials")
    haz_N,    u_haz    = get_waste_frac("hazardous")
    contam_N, u_contam = get_waste_frac("contaminated_masses")
    mixed_N,  u_mixed  = get_waste_frac("mixed_waste")
    rubber_N, u_rubber = get_waste_frac("rubber")
    park_N,   u_park   = get_waste_frac("park_garden")
    manure_N, u_manureN = params.get_global_param_with_uncertainty("manure_N_frac")
    fish_N,   u_fishN   = params.get_global_param_with_uncertainty("animal_waste_N_frac")
    loss_factor, u_loss = params.get_global_param_with_uncertainty("digestate_loss_fraction")
    u_waste_max = max(
        u_wet, u_park, u_wood, u_sludge, u_paper, u_plastic,
        u_rubber, u_text, u_haz, u_mixed, u_other, u_contam)
    uncertainty = combine_uncertainties_percent(
        u_biogass, u_10513, u_12359,
        u_waste_max,
        u_manureN, u_fishN, u_loss)
    # manure for biogas production
    workbook = openpyxl.load_workbook('data_files/Biogass.xlsx')
    sheet = workbook['Tabell']
    N_content = manure_N
    year_values_manure = {}
    for r in range(3, 15):  
        year = int(sheet.cell(row=r, column=4).value) 
        value = float(sheet.cell(row=r, column=8).value)/1000*N_content 
        year_values_manure[year] = value
    # fiskeslam
    workbook = openpyxl.load_workbook('data_files/12359_20251211-153434.xlsx')
    sheet = workbook['Mengde'] # unit: 1000 ton waste
    # N_content: assume typical value of 10 gN/kg (Schäppi2025Ann p 130) gN/kg10 = kgN/t = tN/kt så dele på 1000
    N_content = fish_N
    year_values_fish = {}
    for col in range(4, sheet.max_column + 1):  
        year = int(sheet.cell(row=3, column=col).value)  
        value = float(sheet.cell(row=29, column=col).value)*N_content # Fiskeavfall og anna maritimt
        year_values_fish[year] = value
    # 2012-2023: Avfallsregnskap for Norge. Oppgitt i 1000 tonn
    workbook = openpyxl.load_workbook('data_files/10513_20260212-104227.xlsx')
    sheet = workbook['10513']
    for col in range(2,110,9):  
        value = 0
        year = int(sheet.cell(row=4, column=col).value)  
        collected_years.add(year)
        # Våtorganisk avfall
        value += float(sheet.cell(row=7, column=col+2).value)*wet_N
        # Park- og hageavfall
        value += float(sheet.cell(row=8, column=col+2).value)*park_N
        # Treavfall
        value += float(sheet.cell(row=9, column=col+2).value)*wood_N
        # Slam
        value += float(sheet.cell(row=10, column=col+2).value)*sludge_N
        # Papir
        value += float(sheet.cell(row=11, column=col+2).value)*paper_N
        # Plast
        value += float(sheet.cell(row=17, column=col+2).value)*plastic_N
        # Gummi
        value += float(sheet.cell(row=18, column=col+2).value)*rubber_N # Schäppi tabell 23
        # Tekstiler
        value += float(sheet.cell(row=19, column=col+2).value)*textile_N
        # Farlig avfall
        value += float(sheet.cell(row=22, column=col+2).value)*haz_N
        # Blandet avfall
        value += float(sheet.cell(row=23, column=col+2).value)*mixed_N
        # Andre materialer
        value += float(sheet.cell(row=24, column=col+2).value)*other_N
        # Forurensede masser
        value += float(sheet.cell(row=25, column=col+2).value)*contam_N
        # legg til husdyrmøkk (antar neglisjerbar før 2013)
        if year > 2012:
            value += year_values_manure[year]  
        # legg til fiskeavfall (antar neglisjerbar før 2017)
        if year > 2016:
            value += year_values_fish[year]
        if  ww_bg.loc[year,'entries'] == 1:
            value += ww_bg.loc[year,'value'] # legge til avløpsslam
            comment = 'ok'
            data_sources = 'SSB'
            value *= loss_factor
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value,
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty        
            })
    for year in range (1984,2012):
        collected_years.add(year)
        value = 0
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_compost(results, dataset_unc):
    flow_code = 'PR.SO-HS.HS-Compost for private gardens-Nmix'
    ## se nærmere på. Sjekk om husdyrmøkk, fiskeavfall osv er med i begge datasettene. Ser ut som stor forskjell på de to. 
    # compost handled by the renovation sector goes predominantly to private gardens and 
    # green areas because they are willing to pay for the product. Will assign 100 % to this flow. 
    collected_years = set()
    # using data from SSB
    # 2001-2011: 05414 Biologisk behandling av avfall. Fylkesvis. Verdier i 1000 tonn. 
    # Kolonne "ferdig kompost-totalt" (inkluderer "solgt" og "brukt på eget anlegg")
    data_sources = 'SSB'
    comment = 'ok'
    N_content_old, u_comp_old = get_waste_frac("compost_old")
    u_05414 = get_uncertainty(dataset_unc, '05414')
    uncertainty_old = combine_uncertainties_percent(u_05414, u_comp_old)
    workbook = openpyxl.load_workbook('data_files/05414_20260211-130738.xlsx')
    sheet = workbook['05414']
    for col in [29,38,110]:  
        year = int(sheet.cell(row=3, column=col).value) 
        collected_years.add(year)
        # loop through fylker
        value = 0
        for r in range(5,24):
            value += float(sheet.cell(row=r, column=col+4).value)*N_content_old
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_old
        })
    # 2017-2023: 12359 Biologisk behandling av avfall (1000 tonn)
    # ser på alle avfallstyper under ett ("I alt" og antar samme N-innhold, 3%)
    data_sources = 'SSB'
    comment = 'ok'
    N_content_new, u_comp_new = get_waste_frac("compost_new")
    N_loss,        u_loss_comp = get_waste_frac("compost_N_loss")
    u_12359 = get_uncertainty(dataset_unc, '12359')    
    uncertainty_new = combine_uncertainties_percent(u_12359, u_comp_new, u_loss_comp)
    workbook = openpyxl.load_workbook('data_files/12359_20251211-153434.xlsx')
    sheet = workbook['Mengde']
    for col in range(4,11):
        year = int(sheet.cell(row=3, column=col).value)
        collected_years.add(year)
        value = float(sheet.cell(row=6, column=col).value)*N_content_new*(1-N_loss)   # kt to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_new
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_so_NOx_emissions(results, dataset_unc):
    flow_code = 'PR.SO-AT.AT-Emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_crltap, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '1A1a')|(data[2] == '5A')|(data[2] == '5C1a')|(data[2] == '5C1bi')|(
        data[2] == '5C1bii')|(data[2] == '5C1biii')|(data[2] == '5C1biv')|(data[2] == '5C1bv')|(
        data[2] == '5C1bvi')|(data[2] == '5E')|(data[2] == '5B1')|(data[2] == '5B2')]
    data = data[data[3] == 'NOx']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    conv = params.get("NOx_to_N_factor")
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
 

def _add_so_NH3_emissions(results, dataset_unc):
    flow_code = 'PR.SO-AT.AT-Emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")   
    uncertainty = combine_uncertainties_percent(u_crltap, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '1A1a')|(data[2] == '5A')|(data[2] == '5C1a')|(data[2] == '5C1bi')|(
        data[2] == '5C1bii')|(data[2] == '5C1biii')|(data[2] == '5C1biv')|(data[2] == '5C1bv')|(
        data[2] == '5C1bvi')|(data[2] == '5E')|(data[2] == '5B1')|(data[2] == '5B2')]
    data = data[data[3] == 'NH3']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    conv = params.get("NH3_to_N_factor")
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    

def _add_so_N2O_emissions(results, dataset_unc):
    flow_code = 'PR.SO-AT.AT-Emissions-N2O'
    collected_years = set()
    # emissions from waste incineration - also include landfills if relevant
    # norskeutslipp.no does not give N2O emissions from landfills or waste incineration
    # UNFCCC CRT gives N2O emissions from waste incineration
    # read from 2 datafiles: SO and BC (manure used for biofuel production)
    data_sources = 'UNFCCC CRT'
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)
    comment = 'ok'
    data = pd.read_csv('data_files/N2O_SO.csv')
    data = data[['year', 'value']]
    conv = params.get("N2O_to_N_factor")
    for index, row in data.iterrows():
        year = int(row['year'])
        collected_years.add(year)
        value = row['value']*conv # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_so_leaching(results, dataset_unc):
    flow_code = 'PR.SO-HY.SW-Leaching-Nmix'
    collected_years = set()
    u_ns = get_uncertainty(dataset_unc, 'norskeutslipp')
    uncertainty = u_ns    
    comment = 'ok'
    emissions_landfills = find_landfill_emissions_to_water()
    mean_unconnected = emissions_landfills["N_ikke"].mean()
    for year in range (1990,2011):
        collected_years.add(year)
        value = mean_unconnected/1000 # tN -> ktN
        data_sources = 'extrapolated'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    for year in range (2011,2026):
        collected_years.add(year)
        value = emissions_landfills.loc[emissions_landfills["År"] == year, "N_ikke"].sum()/1000 # tN -> ktN
        data_sources = 'norskeutslipp.no'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_solid_waste_export(results, dataset_unc, trade_data, trade_mapping, trade_params):
    flow_code = 'PR.SO-RW.RW-Solid waste export-Nmix'
    collected_years = set()
    # N-content from Table 50, Schâppi2025Ann
    # municipal waste: HS code 382510; 0.9 % (household and similar wastes) 
    # sewage sludge: HS code 382520; 4,40 % (common sludges)
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    year_values, uncertainty = find_solid_waste_export(dataset_unc, trade_data, trade_mapping, trade_params)
    for year,value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    for year in range(1988,2002):
        collected_years.add(year)
        value = 0
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })       
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_export_for_recycling(results, dataset_unc, trade_data, trade_mapping, trade_params):
    flow_code = 'PR.SO-RW.RW-Export-for-recycling-Nmix'
    collected_years = set()
    # N-content from Table 50, Schâppi2025Ann
    # municipal waste: HS code 382510; 0.9 % (household and similar wastes) 
    # sewage sludge: HS code 382520; 4,40 % (common sludges)
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    year_values, uncertainty = find_export_for_recycling(dataset_unc, trade_data, trade_mapping, trade_params)
    for year,value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_export_for_reuse(results, dataset_unc, trade_data, trade_mapping, trade_params):
    flow_code = 'PR.SO-RW.RW-Export-for-reuse-Nmix'
    collected_years = set()
    # N-content from Table 50, Schâppi2025Ann
    # municipal waste: HS code 382510; 0.9 % (household and similar wastes) 
    # sewage sludge: HS code 382520; 4,40 % (common sludges)
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    year_values, uncertainty = find_export_for_reuse(dataset_unc, trade_data, trade_mapping, trade_params)
    for year,value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_ag_sewage_sludge_fertilizer(results, dataset_unc):
    flow_code = 'PR.WW-AG.SM-Sewage sludge fertilizer-Nmix'
    collected_years = set()
    # using data from SSB
    data_sources = 'SSB'
    comment = 'ok'
    u_05279 = get_uncertainty(dataset_unc, '05279')
    sludge_N, u_sludge = get_waste_frac("sludge")
    N_content = sludge_N
    uncertainty = combine_uncertainties_percent(u_05279, u_sludge)
    # 2002-2024
    workbook = openpyxl.load_workbook('data_files/05279_20260121-103739.xlsx')
    sheet = workbook['Slam']
    for col in range(3, sheet.max_column + 1):  
        year = int(sheet.cell(row=3, column=col).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=5, column=col).value)/1000*N_content 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # 1993-2001
    workbook = openpyxl.load_workbook('data_files/slamdisponering.xlsx')
    sheet = workbook['Ark1']
    mean_val = 0
    for r in range(2, 11):  
        year = int(sheet.cell(row=r, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)*float(sheet.cell(row=r, column=4).value)/100*N_content 
        if year < 1996:
            mean_val += value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    mean_val /= 3
    data_sources = 'extrapolated'
    for year in range(1990,1993):  
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': mean_val, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })    
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_hs_sewage_sludge_fertilizer(results, dataset_unc):
    flow_code = 'PR.WW-HS.HS-Sewage sludge fertilizer-Nmix'
    collected_years = set()
    # using data from SSB
    # including sludge for fertilizer also for other green areas and "jordprodusent"
    data_sources = 'SSB table 05279'
    u_05279 = get_uncertainty(dataset_unc, '05279')
    sludge_N, u_sludge = get_waste_frac("sludge")
    N_content = sludge_N
    uncertainty = combine_uncertainties_percent(u_05279, u_sludge)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/05279_20260121-103739.xlsx')
    sheet = workbook['Slam']
    for col in range(3, sheet.max_column + 1):  
        year = int(sheet.cell(row=3, column=col).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=6, column=col).value)/1000*N_content 
        value += float(sheet.cell(row=7, column=col).value)/1000*N_content 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # 1993-2001
    workbook = openpyxl.load_workbook('data_files/slamdisponering.xlsx')
    sheet = workbook['Ark1']
    mean_val = 0
    for r in range(2, 11):  
        year = int(sheet.cell(row=r, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)*float(sheet.cell(row=r, column=3).value)/100*N_content 
        if year < 1996:
            mean_val += value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    mean_val /= 3
    data_sources = 'extrapolated'
    for year in range(1990,1993):  
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': mean_val, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })    
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_sewage_sludge_landfill(results, dataset_unc):
    flow_code = 'PR.WW-PR.SO-Sewage sludge landfill-Nmix'
    # added flow: sewage sludge for landfills
    collected_years = set()
    # using data from SSB
    # including both "deponert" and "dekkmasse avfallsfylling"
    data_sources = 'SSB table 05279'
    u_05279 = get_uncertainty(dataset_unc, '05279')
    sludge_N, u_sludge = get_waste_frac("sludge")
    N_content = sludge_N
    uncertainty = combine_uncertainties_percent(u_05279, u_sludge)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/05279_20260121-103739.xlsx')
    sheet = workbook['Slam']
    for col in range(3, 26):  
        year = int(sheet.cell(row=3, column=col).value) 
        collected_years.add(year)
        # Dekkmasse avfallsfylling
        value = float(sheet.cell(row=8, column=col).value)/1000*N_content   # t to ktN
        # Deponert
        v = sheet.cell(row=9, column=col).value
        if type(v) != str:
            value += float(v)/1000*N_content   # t to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # 1993-2001
    workbook = openpyxl.load_workbook('data_files/slamdisponering.xlsx')
    sheet = workbook['Ark1']
    mean_val = 0
    for r in range(2, 11):  
        year = int(sheet.cell(row=r, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)*float(sheet.cell(row=r, column=5).value)/100*N_content 
        if year < 1996:
            mean_val += value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    mean_val /= 3
    data_sources = 'extrapolated'
    for year in range(1990,1993):  
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': mean_val, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })    
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

# def _add_sewage_sludge_incineration(results, dataset_unc):
#     flow_code = 'PR.WW-PR.SO-Sewage sludge incineration-Nmix'
#     collected_years = set()
#     # using data from SSB
#     # including sludge for fertilizer also for other green areas and "jordprodusent"
#     data_sources = 'SSB table 05279'
#     u_05279 = get_uncertainty(dataset_unc, '05279')
#     sludge_N, u_sludge = get_waste_frac("sludge")
#     N_content = sludge_N
#     uncertainty = combine_uncertainties_percent(u_05279, u_sludge)
#     comment = 'ok'
#     workbook = openpyxl.load_workbook('data_files/05279_20260121-103739.xlsx')
#     sheet = workbook['Slam']
#     for col in range(3, 26):  
#         year = int(sheet.cell(row=3, column=col).value) 
#         # Levert til forbrenning
#         v = sheet.cell(row=10, column=col).value
#         if type(v) != str:
#             collected_years.add(year)
#             value = float(v)/1000*N_content   # t to ktN
#             results.append({
#                 'flow_name': flow_code,
#                 'year': year,
#                 'value': value, 
#                 'comment': comment,
#                 'data_sources': data_sources,
#                 'uncertainty': uncertainty
#             })
#     missing_years = expected_years - collected_years
#     report_missing_years(flow_code, missing_years, results)

def _add_ww_NH3_emissions(results, dataset_unc):
    flow_code = 'PR.WW-AT.AT-Emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_crltap, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '5D1')|(data[2] == '5D2')|(data[2] == '5D3')]
    data = data[data[3] == 'NH3']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_ww_NOx_emissions(results, dataset_unc):
    flow_code = 'PR.WW-AT.AT-Emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_crltap, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '5D1')|(data[2] == '5D2')|(data[2] == '5D3')]
    data = data[data[3] == 'NOx']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_ww_N2O_emissions(results, dataset_unc):
    flow_code = 'PR.WW-AT.AT-Emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 5: 0.16 kt N2O reported from wastewster treatment in 2022. Corresponds to 0,102 kt N
    # not negligible compared to many other streams
    data_sources = 'UNFCCC CRT'
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)
    comment = 'ok'
    data = pd.read_csv('data_files/N2O_SO.csv')
    data = data[['year', 'value']]
    conv = params.get("N2O_to_N_factor")
    for index, row in data.iterrows():
        year = int(row['year'])
        collected_years.add(year)
        value = row['value']*conv # ktN2O to ktN
    # workbook = openpyxl.load_workbook('data_files/N2O_WW.xlsx')
    # sheet = workbook['Ark1']
    # # Industri og bergverk - avfall
    # for row in range(5, 38):  
    #     year = int(sheet.cell(row=row, column=1).value)  
    #     collected_years.add(year)
    #     value = float(sheet.cell(row=row, column=2).value)*conv # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_ww_N2_emissions(results, dataset_unc):
    flow_code = 'PR.WW-AT.AT-Emissions-N2'
    collected_years = set()
    comment = 'ok'
    data_sources = 'treatment plant reports'
    uncertainty = 20
    # denitrification from the WW treatment plants that have N removal
    # read in data from 2002 from norskeutslipp.no and 
    # extrapolate back for those plants that have been in operation longer
    # assume N removal rate of 70 %
    # N_released = N_in*(1-removal)
    # N_in = N_released/(1-removal)
    # N_removed = N_in-N_released
    # N_removed = N_released(1/(1-removal)-1)=N_released(removal/(1-removal))
    # data for single, recent year:
    # 	Rensegrad	N utslipp	Mengde N2	Kilde rensegrad	Kilde utslipp
    # Lillehammer (fra 1995)	0,7	62,594	1,46E+02	antatt (krav)	https://www.norskeutslipp.no/no/Diverse/Virksomhet/?CompanyID=20952
    # Veas (fra 1997)	0,772	688,21	2,33E+03	https://veas.nu/arsrapporter/arsrapport-2023/om-veas/nokkeltall-2023	https://www.norskeutslipp.no/no/Diverse/Virksomhet/?CompanyID=20951
    # Nordre Follo (fra 1997)	0,7	55,294	1,29E+02	antatt (krav)	https://www.norskeutslipp.no/no/Diverse/Virksomhet/?CompanyID=20954
    # Gardermoen (fra 1998)	0,7	44,409	1,04E+02	antatt (krav)	https://www.norskeutslipp.no/no/Diverse/Virksomhet/?CompanyID=20953
    # Bekkelaget (fra 2001)	0,79	317,75	1,20E+03	https://www.statsforvalteren.no/siteassets/fm-oslo-og-viken/miljo-og-klima/egenkotrollrapportering---avlop/oslo-kommune.pdf	https://www.norskeutslipp.no/no/Diverse/Virksomhet/?CompanyID=20950
    # NRVA (fra 2003)
    # Hokksund (fra 2025 - ikke ta med)
    removal_default = 0.7
    N_released = pd.read_excel("data_files/nitrogenrensing_avløp.xlsx", sheet_name="Ark1",nrows = 31) 
    num_cols = N_released.columns[1:]     # alle unntatt "år"
    N_released[num_cols] = N_released[num_cols] / 1000 # tN -> ktN
    mean_Lillehammer = N_released["Lillehammer"].mean() # relativt kontant hele perioden
    mask = (N_released["år"] >= 2002) & (N_released["år"] <= 2003)
    mean_Veas= N_released.loc[mask, "VEAS"].mean() # stiger mye etter to første år
    mean_NordreFollo = N_released["Nordre Follo"].mean() # relativt kontant hele perioden
    mask = (N_released["år"] >= 2002) & (N_released["år"] <= 2009)
    mean_Gardermoen= N_released.loc[mask, "Gardermoen"].mean() # stiger mye etter 2009
    mean_NRVA = N_released["NRVA"].mean() # relativt kontant hele perioden
    for year in expected_years:
        collected_years.add(year)
        if year < 1995:
            value = 0
        elif year < 1997: # kun Lillehammer
            value = mean_Lillehammer*(removal_default/(1-removal_default))
        elif year < 1998: # + VEAS og Nordre Follo
            value = (mean_Lillehammer+mean_Veas+mean_NordreFollo)*(removal_default/(1-removal_default))
        elif year < 2002: # + Gardermoen
            value = (mean_Lillehammer+mean_Veas+mean_NordreFollo+mean_Gardermoen)*(removal_default/(1-removal_default))
        elif year == 2002:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year+1, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) #ekstrapolere fra neste år
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
        elif year == 2003: # + NRVA
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) #ekstrapolere fra neste år
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += mean_NRVA*removal_default
        elif year in [2004,2005]:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += mean_NordreFollo*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year == 2006:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += mean_NordreFollo*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += mean_NRVA*(removal_default/(1-removal_default))
        elif year in [2007,2008]:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year in range(2009,2012):
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year == 2012:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += mean_NordreFollo*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year in range(2013,2016):
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year == 2016:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year in [2017,2018]:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad NRVA"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal/(1-removal))
        elif year in [2019,2020]:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            removal = N_released.loc[N_released["år"] == year, "rensegrad NRVA"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal/(1-removal))
        elif year == 2021:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal_default/(1-removal_default))
        elif year < 2025:
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            removal = N_released.loc[N_released["år"] == year, "rensegrad NRVA"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal/(1-removal))
        else:            
            value = N_released.loc[N_released["år"] == year, "Lillehammer"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad VEAS"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "VEAS"].iloc[0]*(removal_default/(1-removal_default))
            value += N_released.loc[N_released["år"] == year, "Nordre Follo"].iloc[0]*(removal_default/(1-removal_default)) 
            value += N_released.loc[N_released["år"] == year, "Gardermoen"].iloc[0]*(removal_default/(1-removal_default))
            removal = N_released.loc[N_released["år"] == year, "rensegrad Bekkelaget"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "Bekkelaget"].iloc[0]*(removal/(1-removal))
            removal = N_released.loc[N_released["år"] == year, "rensegrad NRVA"].iloc[0]
            value += N_released.loc[N_released["år"] == year, "NRVA"].iloc[0]*(removal/(1-removal))
            value += N_released.loc[N_released["år"] == year, "Hokksund"].iloc[0]*(removal_default/(1-removal_default))
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,  
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_treated_ww_discharge(results, dataset_unc):
    flow_code = 'PR.WW-HY.CW-Treated wastewater discharge-Nmix'
    collected_years = set()
    # assuming all water discharges from WW to CW, not SW (not separated in statistics)
    # using data from SSB
    data_sources = 'SSB table 05280'
    u_05280 = get_uncertainty(dataset_unc, '05280')
    uncertainty = u_05280    
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/05280_20251113-113329.xlsx')
    sheet = workbook['Nitrogen']
    for col in range(4, 27):  
        year = int(sheet.cell(row=3, column=col).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=4, column=col).value)/1000     # t to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # 1997-2001
    workbook = openpyxl.load_workbook('data_files/utslipp_avløp.xlsx')
    sheet = workbook['Ark1']
    for r in range(2, 7):  
        year = int(sheet.cell(row=r, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)
        if year == 1997:
            value_1997 = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    data_sources = 'extrapolated'
    for year in range(1990,1997):  
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value_1997, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })    
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)



if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)
