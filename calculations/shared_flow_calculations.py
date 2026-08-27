#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Flow calculations shared across two or more pools (e.g. aquaculture
production feeds both HY and MP/RW; household/industry waste composition
feeds both HS and MP/PR), kept here so each pool module doesn't reimplement
the same source-data parsing.
"""
import pandas as pd
import numpy as np
import openpyxl

from calculations.utils import (
    EXPECTED_YEARS,
    process_generic_trade_flow
)


def find_aquaculture_production(df_aqua_modern, df_aqua_old, current_params, dataset_noise):
    """
    Computes N in aquaculture production (salmon/trout) from Fiskeridirektoratet
    sales data, used across HY, MP and RW for flows built on aquaculture output
    (e.g. HY.CW-HY.MM-Aquaculture production-Nmix). df_aqua_modern (1994
    onward, one column per year) and df_aqua_old (1984-1993) are both already
    cleaned to numeric dtypes at load time in data_loader.py.
    """
    aquaculture_production = {}

    fish_N_frac = float(current_params.get('fish_N_frac'))
    key_fisk = 'Fiskeridirektoratet'
    noise_aqua = dataset_noise[key_fisk]

    # --- Modern data (1994 onward) ---
    for col in df_aqua_modern.columns:
        year = int(col)
        col_data = pd.to_numeric(df_aqua_modern[col], errors='coerce').fillna(0)
        value_tonn = col_data.sum()

        # tonnes / 1000 -> kt round weight, times N fraction and activity noise
        val_kt_N = (value_tonn / 1000) * fish_N_frac * noise_aqua

        aquaculture_production[year] = val_kt_N

    # --- Historical data (1984-1993) ---
    for _, row in df_aqua_old.iterrows():
        year = int(row.iloc[0])
        value_base = float(row.iloc[1])

        # kt round weight * N fraction * activity noise
        val_kt_N = value_base * fish_N_frac * noise_aqua

        aquaculture_production[year] = val_kt_N

    return aquaculture_production


def find_export_for_recycling(results, preloaded_data, current_params, current_trade_factors, dataset_noise):
    year_values = process_generic_trade_flow(
        preloaded_data=preloaded_data, 
        current_params=current_params,
        current_trade_factors=current_trade_factors, 
        flow_code='PR.SO-RW.RW-Export for recycling-Nmix',
        target_types=['plastavfall', 'papiravfall', 'tekstilavfall'],
        is_import=False,
        dataset_noise=dataset_noise,
        results=results, 
    )
    
    return year_values


def find_export_for_reuse(results, preloaded_data, current_params, current_trade_factors, dataset_noise):
    year_values = process_generic_trade_flow(
        preloaded_data=preloaded_data, 
        current_params=current_params,
        current_trade_factors=current_trade_factors, 
        flow_code='PR.SO-RW.RW-Export for reuse-Nmix',
        target_types=['tekstil_brukt'],
        is_import=False,
        dataset_noise=dataset_noise,
        results=results, 
    )
    
    return year_values

def find_feedstock_fuel(preloaded_data, current_params, dataset_noise):
    """
    Computes N in fossil fuel used as chemical feedstock rather than burned
    for energy (used for EF.EC-MP.OP-Fuel used as feedstock-Nmix in ef_mc.py,
    and folded into the consumer-goods mass balance in mp_mc.py). Reads the
    "11 Netto innenlands forbruk som råstoff" (net domestic consumption as
    feedstock) section of SSB table 11561, which breaks the feedstock total
    down by energy product - this row selection deliberately targets only
    that section, not the table's overall consumption figures.
    """
    year_values = {}

    noise_energy = float(dataset_noise['11561'])

    GWh_to_TJ_factor = float(current_params.get('GWh_to_TJ_factor'))
    coal_NCV         = float(current_params.get('coal_feedstock_NCV'))
    oil_NCV          = float(current_params.get('oil_feedstock_NCV'))
    coal_N_frac      = float(current_params.get('coal_feedstock_N_frac'))
    oil_N_frac       = float(current_params.get('oil_feedstock_N_frac'))

    df_energy = preloaded_data.get('ssb_energy_balance_11561')

    # --- Coal and coal products (rows 38-72, 1990-2024) ---
    for row_idx in range(38, 73):
        if row_idx >= len(df_energy):
            break
        row_data = df_energy.iloc[row_idx]
        year_val = row_data.iloc[2]
        value_val = row_data.iloc[3]

        if pd.notna(year_val) and pd.notna(value_val) and value_val != '-':
            year = int(year_val)
            value = float(value_val) / (GWh_to_TJ_factor * coal_NCV) * coal_N_frac
            year_values[year] = year_values.get(year, 0.0) + (value * noise_energy)

    # --- Oil and oil products, excl. bio (rows 108-142, 1990-2024) ---
    for row_idx in range(108, 143):
        if row_idx >= len(df_energy):
            break
        row_data = df_energy.iloc[row_idx]
        year_val = row_data.iloc[2]
        value_val = row_data.iloc[3]

        if pd.notna(year_val) and pd.notna(value_val) and value_val != '-':
            year = int(year_val)
            value = float(value_val) / (GWh_to_TJ_factor * oil_NCV) * oil_N_frac
            year_values[year] = year_values.get(year, 0.0) + (value * noise_energy)

    return year_values

def find_food_industry_waste(df_05282, df_10514, current_params, dataset_noise):
    """
    Computes N in food industry waste (used by mp_mc.py for
    MP.FP-PR.SO-Food industry waste-Nmix) from the "wet organic" waste
    category (row 14 in 05282 / row 7 in 10514) for the mining, manufacturing
    and other/unspecified-industry sectors - the same sector selection as
    find_other_industry_waste, since 'wet organic' is assumed to be entirely
    food-industry-related.

    Table 05282 (1995-2011) does not report an equivalent 2012 baseline, so
    the older values are scaled by the ratio of the 2012 value (from 10514)
    to the function's own 2011 value (from 05282), rather than combined
    directly - this keeps the two tables' differing category definitions
    from creating a level jump at the boundary the way the unscaled
    'mixed waste' transition does elsewhere in this file.
    """
    year_values = {}
    wet_org_N = float(current_params.get('wet_organic'))

    noise_05282 = float(dataset_noise['05282'])
    noise_10514 = float(dataset_noise['10514'])
    noise_trend = float(dataset_noise['trend interpolation'])

    value_2012_base = 0.0

    # --- PART 1: years 2012-2023 (table 10514) ---
    for col in range(2, 115, 10):
        p_col = col - 1  # convert to pandas' 0-based column index
        year_val = df_10514.iloc[3, p_col]
        if pd.isna(year_val):
            continue
        year = int(float(year_val))

        # Wet organic waste (row 7 -> index 6), Bergverk/Industri/Annen-uspesifisert sectors
        v_base = 0.0
        v_base += float(df_10514.iloc[6, p_col+1]) * wet_org_N
        v_base += float(df_10514.iloc[6, p_col+3]) * wet_org_N
        v_base += float(df_10514.iloc[6, p_col+8]) * wet_org_N

        if year == 2012:
            value_2012_base = v_base

        year_values[year] = {
            'value': max(0.0, v_base * noise_10514),
            'comment': 'ok',
            'data_sources': 'SSB'
        }

    if value_2012_base == 0.0:
        raise ValueError("[KRITISK] Fant ikke basisverdi for år 2012 i Tabell 10514. Skalering umulig!")

    # --- PART 2: years 1995-2011 (table 05282), scaled to match 2012's level ---
    p_col_2011 = 162 - 1
    value_2011_base = float(df_05282.iloc[13, p_col_2011+1]) * wet_org_N
    value_2011_base += float(df_05282.iloc[13, p_col_2011+3]) * wet_org_N
    value_2011_base += float(df_05282.iloc[13, p_col_2011+8]) * wet_org_N

    mean_val_accumulator = 0.0
    mean_year_count = 0

    for col in range(2, 170, 10):
        p_col = col - 1
        year_val = df_05282.iloc[3, p_col]
        if pd.isna(year_val):
            continue
        year = int(float(year_val))

        # Wet organic waste (row 14 -> index 13), same three sectors as above
        v_base = 0.0
        v_base += float(df_05282.iloc[13, p_col+1]) * wet_org_N
        v_base += float(df_05282.iloc[13, p_col+3]) * wet_org_N
        v_base += float(df_05282.iloc[13, p_col+8]) * wet_org_N

        v_scaled = v_base * (value_2012_base / value_2011_base)

        if 1995 <= year < 2000:
            mean_val_accumulator += v_scaled
            mean_year_count += 1

        year_values[year] = {
            'value': max(0.0, v_scaled * noise_05282),
            'comment': 'ok',
            'data_sources': 'SSB'
        }

    # --- PART 3: 1990-1994, extrapolated as the mean of the scaled 1995-1999 values ---
    final_mean = (mean_val_accumulator / mean_year_count) if mean_year_count > 0 else 0.0
    for year in range(1990, 1995):
        year_values[year] = {
            'value': max(0.0, final_mean * noise_05282 * noise_trend),
            'comment': 'ok',
            'data_sources': 'extrapolated'
        }

    return year_values


def find_household_waste(preloaded_data, current_params, dataset_noise):
    household_waste = {y: 0.0 for y in range(1990, 2024)}
    
    noise_05282 = float(dataset_noise['05282'])
    noise_10514 = float(dataset_noise['10514'])
    noise_interp = float(dataset_noise['trend interpolation'])

    paper_N   = float(current_params.waste_N_frac('paper'))
    plastic_N = float(current_params.waste_N_frac('plastic'))
    wood_N    = float(current_params.waste_N_frac('wood'))
    textile_N = float(current_params.waste_N_frac('textiles'))
    wet_N     = float(current_params.waste_N_frac('wet_organic'))
    other_N   = float(current_params.waste_N_frac('other_materials'))
    haz_N     = float(current_params.waste_N_frac('hazardous'))
    contam_N  = float(current_params.waste_N_frac('contaminated_masses'))
    park_N    = float(current_params.waste_N_frac('park_garden'))
    mixed_N   = float(current_params.waste_N_frac('mixed_waste'))

    # =========================================================================
    # TABLE 05281 / 05282 (1995-2011)
    # =========================================================================
    # Sectors summed here: Bygge- og anleggsvirksomhet (construction),
    # Tjenesteytende næringer (services), Private husholdninger (households).
    # Unlike the 2012+ block below, this period does NOT include the
    # power/water supply or water/sewage/waste-management sectors - 05282
    # doesn't group them the same way 10514 does, and no attempt is made to
    # reconstruct an equivalent. This is a real (small) source of
    # under-coverage for 1995-2011 relative to 2012+.
    #
    # 'Blandet avfall' (mixed waste) has no equivalent row in this table at
    # all. Verified this isn't hidden elsewhere: 05282's own row totals for
    # these sectors are fully reconciled by metal + glass + concrete + sludge
    # (the rows deliberately excluded here) - there's no unaccounted tonnage.
    # 'Blandet avfall' is a new SSB reporting category that starts with table
    # 10514 in 2012, persistently large (~2300-2400 kt/year in these sectors,
    # not a one-off), not a reclassification of tonnage that existed before
    # under another name. This means 1995-2011 household waste is likely a
    # real undercount of the true total relative to 2012 onward, not a
    # double-counting or calibration artifact in this function - visible as a
    # ~40% jump in the computed flow between 2011 and 2012.
    #
    # Separately, 2009 shows a real dip (should not be corrected): this
    # table's own footnote states industries are classified under SN2007 from
    # 2008 (SN2002 before), and the dip is entirely in the construction and
    # services sectors specifically - private households alone shows no dip
    # and grows smoothly throughout. Consistent with the 2008-2009 financial
    # crisis, the classification change, or both.
    df_05282 = preloaded_data['ssb_05282']
    value_1995 = 0.0
    width_05282 = df_05282.shape[1]

    col_to_year = {}
    for col_idx in range(1, width_05282):
        val = str(df_05282.iloc[3, col_idx]).strip()
        if val.replace('.0', '').isdigit():
            y = int(float(val))
            if 1995 <= y <= 2011:
                col_to_year[col_idx] = y

    for col_idx, year in col_to_year.items():
        val_year = 0.0
        
        # Paper (row 7 -> index 6)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[6, col_idx + c]) * paper_N
        # Plastic (row 9 -> index 8)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[8, col_idx + c]) * plastic_N
        # Wood waste (row 12 -> index 11)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[11, col_idx + c]) * wood_N
        # Textiles (row 13 -> index 12)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[12, col_idx + c]) * textile_N
        # Wet organic (row 14 -> index 13)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[13, col_idx + c]) * wet_N
        # Other materials (row 17 -> index 16)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[16, col_idx + c]) * other_N
        # Hazardous waste (row 18 -> index 17)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[17, col_idx + c]) * haz_N
        # Contaminated masses (row 19 -> index 18)
        for c in [5, 6, 9]:
            if col_idx + c < width_05282: val_year += float(df_05282.iloc[18, col_idx + c]) * contam_N

        household_waste[year] = val_year * noise_05282
        if year == 1995:
            value_1995 = household_waste[year]

    # =========================================================================
    # TABLE 10513 / 10514 (2012-2023)
    # =========================================================================
    df_10514 = preloaded_data['ssb_10514']
    width_10514 = df_10514.shape[1]
    
    col_to_year_10514 = {}
    for col_idx in range(1, width_10514):
        val = str(df_10514.iloc[3, col_idx]).strip()
        if val.replace('.0', '').isdigit():
            y = int(float(val))
            if 2012 <= y <= 2023:
                col_to_year_10514[col_idx] = y

    for col_idx, year in col_to_year_10514.items():
        val_year = 0.0
        
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[6, col_idx + c]) * wet_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[7, col_idx + c]) * park_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[8, col_idx + c]) * wood_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[10, col_idx + c]) * paper_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[16, col_idx + c]) * plastic_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[18, col_idx + c]) * textile_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[21, col_idx + c]) * haz_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[22, col_idx + c]) * mixed_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[23, col_idx + c]) * other_N
        for c in [4, 5, 6, 7, 9]:
            if col_idx + c < width_10514: val_year += float(df_10514.iloc[24, col_idx + c]) * contam_N

        household_waste[year] = val_year * noise_10514

    # =========================================================================
    # EXTRAPOLATION, 1990-1994
    # =========================================================================
    inhabitants_1990 = 4233116
    inhabitants_1995 = 4348410
    waste_kg_person_1990 = 200
    waste_kg_person_1995 = 289
    
    waste_kt_1990 = waste_kg_person_1990 * inhabitants_1990 * 1e-6
    waste_kt_1995 = waste_kg_person_1995 * inhabitants_1995 * 1e-6
    
    N_frac = value_1995 / waste_kt_1995
    value_1990 = waste_kt_1990 * N_frac
    change_per_year = (value_1995 - value_1990) / 5.0
    
    for idx, year in enumerate(range(1990, 1995)):
        household_waste[year] = (value_1990 + change_per_year * idx) * noise_interp

    return household_waste


def find_other_industry_waste(df_05282, df_10514, df_hist_waste, current_params, dataset_noise):
    """
    Computes N in other industry waste (used by mp_mc.py for
    MP.OP-PR.SO-Other industry waste-Nmix), combining SSB tables 05282
    (1995-2011) and 10514 (2012-2023) for the mining, manufacturing and
    other/unspecified-industry sectors, plus a pre-1995 linear extrapolation
    calibrated against df_hist_waste's 1992/1995 totals.

    Table 10514 introduces a 'Blandet avfall' (mixed waste) category with no
    equivalent in 05282 (the same table-transition issue documented in
    find_household_waste above), but here the deliberately lower N-content
    parameter for 'other_materials' relative to 'mixed_waste' keeps the
    1995-2011-to-2012-2023 transition continuous for these sectors - unlike
    household waste, where the same transition is not compensated.
    """
    industry_waste = {}
    
    paper_N     = float(current_params.waste_N_frac('paper'))
    plastic_N   = float(current_params.waste_N_frac('plastic'))
    wood_N      = float(current_params.waste_N_frac('wood'))
    textiles_N  = float(current_params.waste_N_frac('textiles'))
    wet_org_N   = float(current_params.waste_N_frac('wet_organic'))
    other_mat_N = float(current_params.waste_N_frac('other_materials'))
    hazardous_N = float(current_params.waste_N_frac('hazardous'))
    mixed_N     = float(current_params.waste_N_frac('mixed_waste'))
    
    noise_05282 = float(dataset_noise['05282'])
    noise_10514 = float(dataset_noise['10514'])
    noise_trend = float(dataset_noise['trend interpolation'])        

    arr_05282 = df_05282.values
    arr_10514 = df_10514.values
    
    base_value_1995 = 0.0

    # --- PART 1: YEARS 1995-2011 (table 05282) ---
    # c offsets 2, 3, 8 select the Bergverk (mining), Industri (manufacturing) and
    # Annen eller uspesifisert næring (other/unspecified industry) sector columns.
    for col in range(1, 169, 10):
        year = int(arr_05282[3, col])

        value_base = 0.0
        # Paper (row 7 -> index 6)
        for c in [2, 3, 8]: value_base += float(arr_05282[6, col + c]) * paper_N
        # Plastic (row 9 -> index 8)
        for c in [2, 3, 8]: value_base += float(arr_05282[8, col + c]) * plastic_N
        # Wood waste (row 12 -> index 11)
        for c in [2, 3, 8]: value_base += float(arr_05282[11, col + c]) * wood_N
        # Textiles (row 13 -> index 12)
        for c in [2, 3, 8]: value_base += float(arr_05282[12, col + c]) * textiles_N
        # Wet organic (row 14 -> index 13) - mining sector only (c=2)
        for c in [2]:       value_base += float(arr_05282[13, col + c]) * wet_org_N
        # Other materials (row 17 -> index 16)
        for c in [2, 3, 8]: value_base += float(arr_05282[16, col + c]) * other_mat_N
        # Hazardous waste (row 18 -> index 17)
        for c in [2, 3, 8]: value_base += float(arr_05282[17, col + c]) * hazardous_N

        if year == 1995:
            base_value_1995 = value_base

        industry_waste[year] = value_base * noise_05282

    # --- PART 2: YEARS 2012-2023 (table 10514) ---
    # Same c offsets 2, 3, 8 select the same three sectors in this table, even
    # though it splits power/water supply into two extra columns not used here.
    for col in range(1, 114, 10):
        year = int(arr_10514[3, col])

        value_base = 0.0
        # Wet organic (row 7 -> index 6)
        for c in [2]:       value_base += float(arr_10514[6, col + c]) * wet_org_N
        # Wood waste (row 9 -> index 8)
        for c in [2, 3, 8]: value_base += float(arr_10514[8, col + c]) * wood_N
        # Paper (row 11 -> index 10)
        for c in [2, 3, 8]: value_base += float(arr_10514[10, col + c]) * paper_N
        # Plastic (row 17 -> index 16)
        for c in [2, 3, 8]: value_base += float(arr_10514[16, col + c]) * plastic_N
        # Textiles (row 19 -> index 18)
        for c in [2, 3, 8]: value_base += float(arr_10514[18, col + c]) * textiles_N
        # Other materials (row 24 -> index 23)
        for c in [2, 3, 8]: value_base += float(arr_10514[23, col + c]) * other_mat_N
        # Hazardous waste (row 22 -> index 21)
        for c in [2, 3, 8]: value_base += float(arr_10514[21, col + c]) * hazardous_N
        # Mixed waste (row 23 -> index 22)
        for c in [2, 3, 8]: value_base += float(arr_10514[22, col + c]) * mixed_N

        industry_waste[year] = value_base * noise_10514

    # --- PART 3: LINEAR EXTRAPOLATION BACK TO 1990 ---
    waste_kt_1992 = float(df_hist_waste.iloc[1, 2])
    waste_kt_1995 = float(df_hist_waste.iloc[2, 2])

    # Calibrated on raw (noise-free) 1995 data so noise isn't compounded twice.
    N_frac = base_value_1995 / waste_kt_1995
    base_value_1992 = waste_kt_1992 * N_frac
    change_per_year = (base_value_1995 - base_value_1992) / 3

    step = 0
    for year in range(1990, 1995):
        base_value_extrapolated = base_value_1992 + (change_per_year * step)
        step += 1
        # Noise applied once, at the end, not to the raw calibration inputs above.
        industry_waste[year] = base_value_extrapolated * noise_05282 * noise_trend

    return industry_waste


def find_industrial_crop_products(df_gnb_sheet30, dataset_noise):
    """
    Computes N in crop products for industrial use (used by mp_mc.py for
    AG.SM-MP.OP-Crop products for industrial use-Nmix) from the Eurostat
    Gross Nutrient Balance sheet's year/value row pair.
    """
    year_values = {}

    key_gnb = 'Gross nutrient balance'
    noise_gnb_val = dataset_noise[key_gnb]

    key_interp = 'trend interpolation'
    noise_interp_val = dataset_noise[key_interp]

    year_row = df_gnb_sheet30.iloc[8]
    value_row = df_gnb_sheet30.iloc[10]

    for col_idx in range(1, len(df_gnb_sheet30.columns)):
        year_val = year_row.iloc[col_idx]
        val_val = value_row.iloc[col_idx]

        if pd.notna(year_val) and pd.notna(val_val) and val_val != '-':
            # Eurostat marks 2017-2019 as missing with ':' rather than leaving
            # the cell blank, which float() can't parse - those years are the
            # ones filled in by the interpolation block below.
            try:
                year = int(year_val)
                base_value = float(val_val) * 1.0e-3  # kg -> kt
                value = base_value * noise_gnb_val

                year_values[year] = value
            except ValueError:
                continue

    # Eurostat reports no data at all for 2017-2019; fill the gap with the
    # mean of all other years.
    if year_values:
        mean_value = float(np.mean(list(year_values.values())))
        
        for year in range(2017, 2020):
            value_interp = mean_value * noise_interp_val
                
            year_values[year] = value_interp
            
    return year_values

def find_industrial_round_wood(preloaded_data, current_params, dataset_noise):
    """
    Computes N in industrial round wood (used by fs_mc.py for
    FS.FO-MP.OP-Industrial round wood-Nmix) from FAOSTAT forestry production
    volumes, converted to mass via wood density and split into conifer/
    non-conifer N content.
    """
    year_values = {}

    noise_faostat = dataset_noise['Forestry production and trade']
    wood_density  = float(current_params.get('wood_density'))
    conifer_N     = float(current_params.get('conifer_N_frac'))
    nonconifer_N   = float(current_params.get('nonconifer_N_frac'))

    # 'faostat_forestry' <- FAOSTAT_data_en_2-20-2026.csv (data_loader.py DATA_MAP):
    # FAOSTAT forestry production and trade.
    data = preloaded_data.get('faostat_forestry')

    filtered_data = data[(data['Element'] == 'Production') & (data['Value'] != 0)].copy()
    
    items_conifer = ['Industrial roundwood, coniferous']
    items_nonconifer = ['Industrial roundwood, non-coniferous']
    
    final_data = filtered_data[filtered_data['Item'].isin(items_conifer + items_nonconifer)].copy()
    
    final_data['tonnes'] = final_data['Value'] * wood_density
    
    mask_conifer = final_data['Item'].isin(items_conifer)
    mask_nonconifer = final_data['Item'].isin(items_nonconifer)
    
    final_data['N_kg_per_kg'] = 0.0
    final_data.loc[mask_conifer, 'N_kg_per_kg'] = conifer_N
    final_data.loc[mask_nonconifer, 'N_kg_per_kg'] = nonconifer_N
    
    # tonnes * kg N/tonne / 1e3 -> kt N
    final_data['N_amount'] = final_data['tonnes'] * final_data['N_kg_per_kg'] / 1e3

    total_N_per_year = final_data.groupby('Year')['N_amount'].sum().to_dict()

    for year in EXPECTED_YEARS:
        value = total_N_per_year.get(year, 0.0)
        if value > 0:
            year_values[year] = value * noise_faostat

    return year_values

def find_industrial_waste_fuels(df_bio_08205, df_bio_hist, current_params, dataset_noise):
    """
    Computes N in industrial waste fuels (used by mp_mc.py for
    MP.OP-EF.IC-Industrial waste fuels-Nmix) from SSB table 08205 bioenergy
    volumes (2003-2024) plus an older historical series (1998-2002), with
    1990-1997 extrapolated from the average of all pre-2008 years.
    """
    year_values = {}

    noise_08205 = float(dataset_noise['08205'])
    noise_trend = float(dataset_noise['trend interpolation'])

    NCV              = float(current_params.get('firewood_NCV'))
    N_content        = float(current_params.get('firewood_N_frac'))
    GWh_to_TJ_factor = float(current_params.get('GWh_to_TJ_factor'))

    arr_08205 = df_bio_08205.values
    arr_hist = df_bio_hist.values

    raw_sum_pre_2008 = 0.0

    # --- PART 1: SSB table 08205 (2003-2024) ---
    for col in range(3, 25):
        year_val = arr_08205[2, col]
        value_val = arr_08205[9, col]

        year = int(year_val)
        # GWh -> TJ, divide by NCV for kt of fuel, multiply by N_content for kt N
        value_raw = float(value_val) / GWh_to_TJ_factor / NCV * N_content

        year_values[year] = value_raw * noise_08205
        if year < 2008:
            raw_sum_pre_2008 += value_raw

    # --- PART 2: historical series, 1998-2002 (df_bio_hist) ---
    # df_bio_hist predates table 08205 and has no uncertainty entry of its own;
    # it is reused here since both are SSB bioenergy accounting series.
    for r in range(1, 6):
        year_val = arr_hist[r, 0]
        val_col2 = arr_hist[r, 1]
        val_col3 = arr_hist[r, 2]

        year = int(year_val)
        value_raw = (float(val_col2) + float(val_col3)) / GWh_to_TJ_factor / NCV * N_content

        year_values[year] = value_raw * noise_08205

        if year < 2008:
            raw_sum_pre_2008 += value_raw

    # --- PART 3: 1990-1997, extrapolated as the mean of all pre-2008 years ---
    mean_value_raw = raw_sum_pre_2008 / 10.0

    for year in range(1990, 1998):
        year_values[year] = mean_value_raw * noise_trend

    return year_values


def find_non_edible_animal_products(df_hides_clean, df_wool, df_sheep, current_params, dataset_noise):
    """
    Computes N in non-edible animal products (hides and wool) for
    AG.MM-MP.OP-Non-edible animal products-Nmix. Hides come from FAOSTAT
    throughout; wool switches from directly reported Landbruksdirektoratet
    data (from 2005) to an SSB sheep-count-based extrapolation for earlier
    years, with 2001 interpolated from 2000 and 2002 due to a gap in
    ssb_sheep_numbers.

    This branching is duplicated in ag_mc.py's
    _add_non_edible_animal_products_flow_mc, which must build a matching
    data_sources string for each year - keep the two in sync.
    """
    year_values = {}

    noise_faostat = dataset_noise['Crops and livestock products']
    noise_ssb = dataset_noise['03710']
    noise_wool = dataset_noise['Landbruksdirektoratet_wool']
    noise_trend = dataset_noise['trend interpolation']

    N_content_hides = current_params.get('prod_Raw hides and skins')
    wool_pr_sheep = current_params.get('wool_per_sheep')
    N_content_wool = current_params.get('wool_N_frac')

    df_hides = df_hides_clean.copy()
    df_hides['N_amount'] = df_hides['Value'] * float(N_content_hides) * 1e-5 * float(noise_faostat)
    total_N_per_year = df_hides.groupby('Year')['N_amount'].sum().to_dict()

    for year in range(1990, 2024):
        value = total_N_per_year.get(year, 0.0)

        if year > 2004:
            # Reported wool deliveries (Landbruksdirektoratet)
            wool_row = df_wool[df_wool['år'] == year]
            if not wool_row.empty:
                value += float(wool_row['ull'].iloc[0]) * float(N_content_wool) * float(noise_wool)

        elif year != 2001:
            # Wool extrapolated from SSB sheep counts
            sheep_row = df_sheep[df_sheep['År'] == year]
            if not sheep_row.empty:
                value += float(sheep_row['Husdyr (sau)'].iloc[0]) * float(wool_pr_sheep) * float(N_content_wool) * 1e-6 * float(noise_ssb)

        else:
            # 2001 interpolated from the surrounding years' sheep counts
            sheep_prev = df_sheep[df_sheep['År'] == year-1]
            sheep_next = df_sheep[df_sheep['År'] == year+1]
            if not sheep_prev.empty and not sheep_next.empty:
                avg_sheep = 0.5 * (
                    float(sheep_prev['Husdyr (sau)'].iloc[0]) +
                    float(sheep_next['Husdyr (sau)'].iloc[0])
                )
                value += (avg_sheep * float(wool_pr_sheep) * float(N_content_wool) * 1e-6 * float(noise_ssb)) * float(noise_trend)
                
        year_values[year] = value
        
    return year_values

    
def find_recycling(preloaded_data, current_params, current_trade_factors, dataset_noise, 
                    prepared_trade_recycling, prepared_trade_reuse, trade_params):
    year_values = {y: 0.0 for y in range(1990, 2024)}
    
    noise_05281 = float(dataset_noise['05281'])
    noise_10513 = float(dataset_noise['10513'])
    noise_old = float(dataset_noise['historical_waste'])

    paper_N   = float(current_params.waste_N_frac('paper'))
    plastic_N = float(current_params.waste_N_frac('plastic'))
    wood_N    = float(current_params.waste_N_frac('wood'))
    textile_N = float(current_params.waste_N_frac('textiles'))
    other_N   = float(current_params.waste_N_frac('other_materials'))
    haz_N     = float(current_params.waste_N_frac('hazardous'))
    mixed_N   = float(current_params.waste_N_frac('mixed_waste'))
    rubber_N  = float(current_params.waste_N_frac('rubber'))
    contam_N  = float(current_params.waste_N_frac('contaminated_masses'))

    # 'Blandet avfall' (mixed waste) has no matching row text in table 05281 below -
    # it is a new SSB reporting category introduced with table 10513 in 2012, not a
    # reclassification of tonnage recorded under another name before (see the fuller
    # reconciliation in find_household_waste's comment above, which covers the same
    # table transition). Recycling captures two of the largest 'Blandet avfall'
    # contributions, so this flow's 2011-to-2012 jump is even larger than household
    # waste's: roughly 3.45 kt N in 2011 to 5.6 kt N in 2012, about +62%.
    df_05281 = preloaded_data.get('ssb_waste_05281')
    value_1995 = 0.0

    col_to_year_05281 = {}
    for col_idx in range(3,20):
        val = str(df_05281.iloc[2, col_idx]).strip()
        col_to_year_05281[col_idx] = int(val)

    for idx in range(17, 30):
        row_text = str(df_05281.iloc[idx, 2]).strip()
        n_frac = 0.0
        if 'Papir' in row_text: n_frac = paper_N
        elif 'Plast' in row_text: n_frac = plastic_N
        elif 'Treavfall' in row_text: n_frac = wood_N
        elif 'Tekstiler' in row_text: n_frac = textile_N
        elif 'Andre materialer' in row_text: n_frac = other_N
        elif 'Farlig avfall' in row_text: n_frac = haz_N
        elif 'Forurensede masser' in row_text: n_frac = contam_N

        if n_frac > 0:
            for col_idx, year in col_to_year_05281.items():
                val_kt = float(df_05281.iloc[idx, col_idx])
                year_values[year] += val_kt * n_frac * noise_05281

    value_1995 = year_values.get(1995, 0.0)

    df_10513 = preloaded_data.get('ssb_waste_10513')
        
    col_to_year_10513 = {}
    for col in range(1, df_10513.shape[1], 9):  
        cell_year = str(df_10513.iloc[3, col]).strip()
        if cell_year.replace('.0', '').isdigit():
            current_year = int(float(cell_year))
            target_col = col + 1  # Kolonneindeksen for materialgjenvinning
            if target_col < df_10513.shape[1]:
                col_to_year_10513[target_col] = current_year

    for idx in range(5, 25):
        row_text = str(df_10513.iloc[idx, 0]).strip()
        n_frac = 0.0
        
        if 'Papir' in row_text: n_frac = paper_N
        elif 'Plast' in row_text: n_frac = plastic_N
        elif 'Treavfall' in row_text: n_frac = wood_N
        elif 'Gummi' in row_text: n_frac = rubber_N
        elif 'Tekstiler' in row_text: n_frac = textile_N
        elif 'Farlig avfall' in row_text: n_frac = haz_N
        elif 'Blandet avfall' in row_text: n_frac = mixed_N
        elif 'Andre materialer' in row_text: n_frac = other_N
        elif 'Lett forurensede masser' in row_text.lower(): n_frac = contam_N

        if n_frac > 0:
            for col_idx, year in col_to_year_10513.items():
                val_kt = float(df_10513.iloc[idx, col_idx])
                year_values[year] += val_kt * n_frac * noise_10513
                
    household_waste = find_household_waste(preloaded_data, current_params, dataset_noise)
    
    df_05282_ind = preloaded_data['ssb_05282']
    df_10514_ind = preloaded_data['ssb_10514']
    df_hist_ind  = preloaded_data['ssb_hist_industry_waste']
    
    industry_waste = find_other_industry_waste(
        df_05282_ind, 
        df_10514_ind, 
        df_hist_ind, 
        current_params, 
        dataset_noise
    )

    workbook = openpyxl.load_workbook('data_files/kommunalt_avfall_1985_1995.xlsx')
    sheet = workbook['forbrenning og gjenvinning']
    
    rec_frac_1985 = float(sheet.cell(row=2, column=2).value) / 100
    rec_frac_1992 = float(sheet.cell(row=3, column=2).value) / 100
    
    change_per_year = (rec_frac_1992 - rec_frac_1985) / 7
    rec_frac_1995 = rec_frac_1985 + change_per_year * (1995 - 1985)
    
    N_frac = value_1995 / ((household_waste[1995] + industry_waste[1995]) * rec_frac_1995)
    
    r = 3
    for year in range(1990, 1995):
        waste = household_waste[year] + industry_waste[year]
        
        if year < 1992:
            rec_frac = rec_frac_1985 + change_per_year * (year - 1985)
        else:
            rec_frac = float(sheet.cell(row=r, column=2).value) / 100
            r += 1
            
        value = waste * N_frac * rec_frac * noise_old
        year_values[year] = value        
        
    export_resirk = find_export_for_recycling([], preloaded_data, current_params, current_trade_factors, dataset_noise)
    for year, val in export_resirk.items():
        if year in year_values: year_values[year] -= val

    export_reuse = find_export_for_reuse([], preloaded_data, current_params, current_trade_factors, dataset_noise)
    for year, val in export_reuse.items():
        if year in year_values: year_values[year] -= val

    return year_values


def find_treated_wastewater_discharge(df_05280, df_utslipp, dataset_noise):
    """
    Computes N discharged from treated municipal wastewater (used by
    hy_mc.py's _add_inflow_to_coastal_waters), combining SSB table 05280
    (2002-2023) with an older historical series (1997-2001), with 1990-1996
    held flat at the 1997 value.
    """
    ww_discharge = {}

    key_ssb = '05280'
    noise_ww = dataset_noise[key_ssb]

    value_1997 = 0.0

    # --- PART 1: SSB table 05280 (2002-2023) ---
    years_row = df_05280.iloc[2]
    values_row = df_05280.iloc[3]

    for col_idx in range(3, min(26, df_05280.shape[1])):
        year = int(years_row.iloc[col_idx])
        val_t = float(values_row.iloc[col_idx])
        val_kt_N = (val_t / 1000.0) * noise_ww
        ww_discharge[year] = max(0.0, val_kt_N)

    # --- PART 2: historical series, 1997-2001 (utslipp_avløp.xlsx) ---
    for r_idx in range(1, min(6, len(df_utslipp))):
        year = int(df_utslipp.iloc[r_idx, 0])
        val_kt_N = float(df_utslipp.iloc[r_idx, 1]) * noise_ww
        ww_discharge[year] = max(0.0, val_kt_N)

        if year == 1997:
            value_1997 = val_kt_N

    # --- PART 3: 1990-1996, held flat at the 1997 value ---
    for year in range(1990, 1997):
        ww_discharge[year] = value_1997

    return ww_discharge

