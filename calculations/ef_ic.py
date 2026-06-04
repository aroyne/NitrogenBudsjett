#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 28 15:26:53 2025

@author: anja
"""

import pandas as pd  # Ensure you have pandas installed
import openpyxl

def read_data(file_path):
    """Read dataset from a CSV file."""
    return pd.read_csv(file_path)

def execute_calculations():
    data_sources = 'text' # placeholder for when nothing else is given
    results = []
    
    # Assuming the data file is named 'dataset.csv' located in a 'data_files' folder
#    data_file_path = os.path.join("data_files", "dataset.csv")
#    data = read_data(data_file_path)

    flow_code = 'EF.EC-EF.IC-Fuel for industry-Nmix'
    year = 2018
    value = 3
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-EF.TR-Fuel for transport-Nmix'
    year = 2022
    value = 3
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-EF.OE-Fuel for heating-Nmix'
    year = 2018
    value = 3
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-MP.OP-Fuel used as feedstock-Nmix' 
    year = 2018
    value = 1
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-AT.AT-Emissions-NH3'
    year = 2018
    value = 1
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-AT.AT-Emissions-NOx'
    year = 2018
    value = 1
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-AT.AT-Emissions-N2O'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })
    

    flow_code = 'EF.EC-AT.AT-Emissions-N2'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })

    flow_code = 'EF.EC-RW.RW-Fuel export-Nmix'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
        'flow_name': flow_code,
        'year': year,
        'value': value,
        'comment': comment,
        'data_sources': data_sources
    })

    flow_code = 'EF.IC-AT.AT-Emissions-NH3'
    # use data from SSB
    data_sources = 'SSB table 08941'
    comment = 'ok'
    results = []
    workbook = openpyxl.load_workbook('data_files/08941_20251106-130347.xlsx')
    sheet = workbook['Utslipp']
    year_values = {}
    # Oppvarming i andre næringer og husholdninger
    for row in range(1929, 1964):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*14/17  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*14/17  # Initialize the total for the year
    for year, total_value in year_values.items():
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': total_value,
            'comment': comment,
            'data_sources': data_sources        
        })

    flow_code = 'EF.IC-AT.AT-Emissions-NOx'
    # use data from SSB
    data_sources = 'SSB table 08941'
    comment = 'ok'
    results = []
    workbook = openpyxl.load_workbook('data_files/08941_20251106-130347.xlsx')
    sheet = workbook['Utslipp']
    year_values = {}
    # Olje- og gassutvinning
    for row in range(4, 39):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Energiforsyning, kull etc. 
    for row in range(1334, 1369):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Energiforsyning, ved etc. 
    for row in range(1404, 1439):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Energiforsyning, gass 
    for row in range(1474, 1509):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Energiforsyning, diesel osv. 
    for row in range(1614, 1649):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Energiforsyning, tungolje osv 
    for row in range(1684, 1719):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # Oppvarming i andre næringer og husholdninger
    for row in range(1894, 1928):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    for year, total_value in year_values.items():
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': total_value,
            'comment': comment,
            'data_sources': data_sources        
        })

    flow_code = 'EF.IC-AT.AT-Emissions-N2O'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.IC-AT.AT-Emissions-N2'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.TR-AT.AT-Emissions-NH3'
    # use data from SSB
    data_sources = 'SSB table 08941'
    comment = 'ok'
    results = []
    workbook = openpyxl.load_workbook('data_files/08941_20251106-130347.xlsx')
    sheet = workbook['Utslipp']
    year_values = {}
    # 5 Veitrafikk
    for row in range(2559, 2594):  # 232 is exclusive, so it actually reads up to row 231
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*14/17  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*14/17  # Initialize the total for the year
    for year, total_value in year_values.items():
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': total_value,
            'comment': comment,
            'data_sources': data_sources        
        })

    flow_code = 'EF.TR-AT.AT-Emissions-NOx'
    comment = 'in progress'
    # problem: utslipp fra fiske osv. skal egentlig til subpool OE, men er ikke delt opp i statistikken
    # 5 Veitrafikk
    for row in range(2524, 2559):  
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    # 6 Luftfart, sjøfart, fiske mm. 
    for row in range(3154, 3189):  
        year = sheet.cell(row=row, column=4).value  # Column D is the 4th column
        value = sheet.cell(row=row, column=5).value  # Column E is the 5th column
        # Check if year and value are not None before processing
        if year is not None and value is not None:
            if year in year_values:
                year_values[year] += value/1000*0.3043  # Add the value to the existing total for the year
            else:
                year_values[year] = value/1000*0.3043  # Initialize the total for the year
    for year, total_value in year_values.items():
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': total_value,
            'comment': comment,
            'data_sources': data_sources
        })
   
    flow_code = 'EF.TR-AT.AT-Emissions-N2O'
    # use data from SSB
    data_sources = 'SSB table 13932'
    comment = 'ok'
    results = []
    workbook = openpyxl.load_workbook('data_files/13932_20251106-153015.xlsx')
    sheet = workbook['Utslipp']
    for col in range(3, sheet.max_column + 1):  
        year = sheet.cell(row=4, column=col).value  
        value1 = sheet.cell(row=5, column=col).value    
        value2 = sheet.cell(row=6, column=col).value    
        value3 = sheet.cell(row=7, column=col).value 
        value = value1 + value2 + value3
        value = float(value)/1000*0.6364 #converting from t to kt
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': value, #converting from t to kt
            'comment': comment,
            'data_sources': data_sources
        })


    flow_code = 'EF.TR-AT.AT-Emissions-N2'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.TR-RW.RW-Export of transport fuels-Nmix'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.OE-AT.AT-Emissions-NH3'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.OE-AT.AT-Emissions-NOx'
    year = 2018
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.OE-AT.AT-Emissions-N2O'
    year = 2023
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    flow_code = 'EF.OE-AT.AT-Emissions-N2'
    year = 2023
    value = 2
    comment = 'not done'
    results.append({
         'flow_name': flow_code,
         'year': year,
         'value': value,
         'comment': comment,
         'data_sources': data_sources
    })

    return results  # Returns a dictionary of flow code to (value, comment) pairs
