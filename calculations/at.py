#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov  6 11:34:21 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl

from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
)
from calculations.shared_flow_calculations import (
    find_ammonia_import
    )

expected_years = EXPECTED_YEARS

def execute_calculations(mc_params=None):
    results = []
    params = NParameters("data_files/N_parameters.xlsx")
    
    # HVIS vi kjører Monte Carlo, overstyrer vi de statiske verdiene
    if mc_params is not None:
        params.override_global_params(mc_params)
        
    dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
    trade_params = params.get_trade_params()      # index: param_id
    trade_mapping = params.get_trade_mapping()    # columns: type, konv, Varenr, ...
    
    # # Combustion N2 fixation (mass balances)
    # _add_OE_N2_fixation(results, dataset_unc)
    # _add_TR_N2_fixation(results, dataset_unc)
    # # consider moving uncertainty for mass balanced N fixation to data file
    
    # Industrial ammonia synthesis
    _add_OP_N2_fixation(results, dataset_unc, trade_mapping, trade_params)

    # Biological N2 fixation
    _add_AG_N2_fixation(results, params, dataset_unc)
    _add_FO_N2_fixation(results, params, dataset_unc)
    _add_OL_N2_fixation(results, params, dataset_unc)
    _add_SW_N2_fixation(results, params, dataset_unc)
    # note on N fixation: check out recent data here https://aslopubs.onlinelibrary.wiley.com/doi/full/10.1002/lol2.10459

    # Atmospheric outflows
    _add_atmospheric_outflow_oxn(results, dataset_unc)
    _add_atmospheric_outflow_rdn(results, dataset_unc)

    # Deposition flows (per land class)
    _deposition_flow(results,
                     flow_code='AT.AT-AG.SM-Deposition-OXN',
                     class4='jordbruk',
                     poll='NOx',
                     dataset_unc=dataset_unc)    
    _deposition_flow(results,
                 flow_code='AT.AT-AG.SM-Deposition-RDN',
                 class4='jordbruk',
                 poll='Nred',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-FS.FO-Deposition-OXN',
                 class4='skog',
                 poll='NOx',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-FS.FO-Deposition-RDN',
                 class4='skog',
                 poll='Nred',
                 dataset_unc=dataset_unc)  
    _deposition_flow(results,
                 flow_code='AT.AT-FS.OL-Deposition-OXN',
                 class4='annet',
                 poll='NOx',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-FS.OL-Deposition-RDN',
                 class4='annet',
                 poll='Nred',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-HS.HS-Deposition-OXN',
                 class4='bebyggelse',
                 poll='NOx',
                 dataset_unc=dataset_unc) 
    _deposition_flow(results,
                 flow_code='AT.AT-HS.HS-Deposition-RDN',
                 class4='bebyggelse',
                 poll='Nred',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-HY.SW-Deposition-OXN',
                 class4='overflatevann',
                 poll='NOx',
                 dataset_unc=dataset_unc)
    _deposition_flow(results,
                 flow_code='AT.AT-HY.SW-Deposition-RDN',
                 class4='overflatevann',
                 poll='Nred',
                 dataset_unc=dataset_unc)

    return results  # Returns a list of flow records

def _deposition_flow(results, flow_code, class4, poll, dataset_unc, data_file='data_files/N_per_class_period_distributed_unallocated_long.csv'):
    data = pd.read_csv(data_file)
    u_dep = get_uncertainty(dataset_unc, 'Deposition')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_base = u_dep
    uncertainty_extrap = combine_uncertainties_percent(u_dep, u_interp)

    def period_for_year(y):
        if y < 1988:
            return "1983-1987"
        elif y < 1992:
            return "1988-1992"
        elif y < 1997:
            return "1992-1996"
        elif y < 2002:
            return "1997-2001"
        elif y < 2007:
            return "2002-2006"
        elif y < 2012:
            return "2007-2011"
        else:
            return "2012-2016"

    for year in sorted(expected_years):
        if year < 2017:
            period = period_for_year(year)
            mask = (
                (data["period"] == period) &
                (data["pollutant"] == poll) &
                (data["class4"] == class4)
            )
            subset = data.loc[mask, "N_tonn"]
            if subset.empty:
                raise ValueError(
                    f"No deposition data for flow={flow_code}, class4={class4}, "
                    f"poll={poll}, year={year}"
                )
            value = subset.iloc[0] / 1000
            data_sources = 'NILU and geodata.no'
            uncertainty = uncertainty_base
            if year == 2016:
                value_2016 = value
        elif year < 2022: #bruker NILU data, men siste periode ligger ikke i Excel-arket jeg bruker
            if poll == 'NOx':
                value = value_2016*61440/68166
            else:
                value = value_2016*61175/73494
            value_last = value
        else:
            # mask = (
            #     (data["period"] == "2012-2016") &
            #     (data["pollutant"] == poll) &
            #     (data["class4"] == class4)
            # )
            # values = data.loc[mask, "N_tonn"].values
            # if values.size == 0:
            #     raise ValueError(
            #         f"No deposition data for extrapolation: flow={flow_code}, "
            #         f"class4={class4}, poll={poll}"
            #     )
            value = value_last
            data_sources = 'extrapolated'
            uncertainty = uncertainty_extrap

        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': 'ok',
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

def _add_OP_N2_fixation(results, dataset_unc, trade_mapping, trade_params):
    flow_code = 'AT.AT-MP.OP-Ammonia synthesis N2 fixation-N2'
    collected_years = set()
    # estimated from domestic fertilizer production, use FAOSTAT data and subtracting ammonia import
    comment = 'ok'
    u_fert = get_uncertainty(dataset_unc, 'Fertilizer by nutrient')
    data_sources = 'FAOSTAT Fertilizer by nutrient + SSB'
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-25-2025.csv')
    data = data[['Year', 'Value']]
    # trekke fra import av NH3
    year_values, unc_amm = find_ammonia_import(dataset_unc, trade_mapping, trade_params)
    # Combined, assuming independence
    uncertainty = combine_uncertainties_percent(u_fert, unc_amm)
    for _, row in data.iterrows():
        year = int(row['Year'])
        if year in year_values: # trade data from 1988
            collected_years.add(year)
            value = row['Value']/1000-year_values[year]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    

def _add_AG_N2_fixation(results, params, dataset_unc):
    flow_code = 'AT.AT-AG.SM-Biological N2 fixation-N2'
    # i mangel av noe bedre: bruker konstant verdi tilsvarende Bleken&Bakken
    comment = 'ok'
    data_sources = 'Bleken & Bakken'
    value, u_fix = params.get_global_param_with_uncertainty("AG_biological_fixation_N2")
    uncertainty = u_fix if u_fix is not None else 50.0
    for year in expected_years:
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })

def _add_FO_N2_fixation(results, params, dataset_unc):
    flow_code = 'AT.AT-FS.FO-N2 fixation-N2'
    # following Swedish NBB (Moldan2025Whe): using 1.5 kgN/ha/yr and 12.0 mill ha 
    # (SSB table 14368) -> 18 ktN each year. 
    comment = 'ok'
    data_sources = 'Moldan (2025) and SSB'
    value, u_fix = params.get_global_param_with_uncertainty("FO_biological_fixation_N2")
    uncertainty = u_fix if u_fix is not None else 50.0
    for year in expected_years:
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    
def _add_OL_N2_fixation(results, params, dataset_unc):
    flow_code = 'AT.AT-FS.OL-N2 fixation-N2'
    # Schäffi2025Ann: use biological N2 fixation rates in wetlands, Table 62, with 
    # landcover data from CORINE land cover inventory
    # assume no changes during time period
    # see file Land_cover_CORINE.ipynb for calculation
    comment = 'ok'
    data_sources = 'CORINE land cover inventory and REddy & DeLaune (2008)'
    value, u_fix = params.get_global_param_with_uncertainty("OL_biological_fixation_N2")
    uncertainty = u_fix if u_fix is not None else 50.0
    for year in expected_years:
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    
def _add_SW_N2_fixation(results, params, dataset_unc):
    flow_code = 'AT.AT-HY.SW-N2 fixation-N2'
    # Schäppi2025Ann: The uptake of N2 by nitrogen-fixing microbes is widely 
    # recognized as a significant source of bioavailable nitrogen in marine 
    # environments (Gruber & Galloway, 2008). Nitrogen fixation is mediated by 
    # a variety of autotrophic and heterotrophic bacteria. Cyanobacteria appear 
    # responsible for most planktonic fixation in aquatic ecosystems, and rates 
    # of fixation are high only when these organisms make up a major percentage of the planktonic biomass.
    # Volumes of surface water bodies and coastal waters need to be determined from national statistics.
    # Nitrogen fixation rates for lakes can be found in Howarth et al. 1988: 
    # “Planktonic nitrogen  fixation tends to be low in oligotrophic and 
    # mesotrophic lakes ≪ 0.1 g N m−2 yr−1) but is often high in eutrophic lakes (0.2–9.2 g N m−2 yr−1).”
    # -> assume 1 gN/m2 = 1e6 gN/km2 = 1 ktN/km2
    # according to NIBIO, the surface water area is 20 457 km2 https://arealbarometer.nibio.no/nb/norge/
    # this gives (1 km2 = 100 ha -> 1 t/km2) 2 ktN/y
    comment = 'ok'
    data_sources = 'NIBIO and Reddy & DeLaune (2008)'
    value, u_fix = params.get_global_param_with_uncertainty("SW_biological_fixation_N2")
    uncertainty = u_fix if u_fix is not None else 50.0
    for year in expected_years:
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    
def _add_atmospheric_outflow_oxn(results, dataset_unc):
    flow_code = 'AT.AT-RW.RW-Atmospheric outflow-OXN'
    collected_years = set()
    # using source-receptor data from EMEP, according to Schäppi2025Ann
    comment = 'ok'
    u_sr = get_uncertainty(dataset_unc, 'Source-receptor')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_base = u_sr
    uncertainty_interp = combine_uncertainties_percent(u_sr, u_interp)
    workbook = openpyxl.load_workbook('data_files/atm_in_out.xlsx')
    sheet = workbook['Ark1']    
    for r in range(6,46):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=3).value)/10 # from 100 tN to ktN  
        if sheet.cell(row=r, column=6).value == 'interpolated':
            data_sources = 'interpolated'
            uncertainty = uncertainty_interp
        else:
            data_sources = 'EMEP SR tables'
            uncertainty = uncertainty_base
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_atmospheric_outflow_rdn(results, dataset_unc):
    flow_code = 'AT.AT-RW.RW-Atmospheric outflow-RDN'
    collected_years = set()
    # using source-receptor data from EMEP, according to Schäppi2025Ann
    u_sr = get_uncertainty(dataset_unc, 'Source-receptor')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_base = u_sr
    uncertainty_interp = combine_uncertainties_percent(u_sr, u_interp)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/atm_in_out.xlsx')
    sheet = workbook['Ark1']    
    for r in range(6,46):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=5).value)/10 # from 100 tN to ktN  
        if sheet.cell(row=r, column=6).value == 'interpolated':
            data_sources = 'interpolated'
            uncertainty = uncertainty_interp
        else:
            data_sources = 'EMEP SR tables'
            uncertainty = uncertainty_base
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)




# Example usage
if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)

