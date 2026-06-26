#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 17 11:49:52 2026

@author: anja
"""
import openpyxl
import pandas as pd  

class NParameters:
    def __init__(self, filename="N_parameters.xlsx"):
        # Viktig: lagre filnavnet for senere bruk i get_table
        self.filename = filename

        wb = openpyxl.load_workbook(filename, data_only=True)

        # global_parameters sheet
        self.global_params = {}
        df = pd.read_excel(filename, sheet_name='global_parameters')
        for _, row in df.iterrows():
            param_id = row['parameter_id']
            value = row['value']
            self.global_params[param_id] = value

        # waste_fractions sheet
        self.waste_fractions = {}
        if "waste_fractions" in wb.sheetnames:
            df = pd.read_excel(filename, sheet_name='waste_fractions')
            for _, row in df.iterrows():
                category = row['waste_category']
                n_frac = row['N_frac']
                self.waste_fractions[category] = n_frac

        # animal_weights sheet
        self.animal_weights = {}
        if "animal_weights" in wb.sheetnames:
            df = pd.read_excel(filename, sheet_name='animal_weights')
            for _, row in df.iterrows():
                item_name = row['item_name']
                avg_w = row['avg_weight_kg']
                self.animal_weights[item_name] = avg_w

    def get(self, param_id):
        """Søker etter global parameter. Krasjer hvis den ikke finnes."""
        if hasattr(self, param_id):
            return getattr(self, param_id)
        if param_id in self.global_params:
            return self.global_params[param_id]
        raise KeyError(f"[STOPP] Global parameter '{param_id}' mangler helt i systemet!")
    
    def waste_N_frac(self, category):
        """Søker etter avfallsfraksjon. Krasjer hvis den ikke finnes."""
        if hasattr(self, category):
            return getattr(self, category)
        if category in self.waste_fractions:
            return self.waste_fractions[category]
        raise KeyError(f"[STOPP] Avfallsfraksjon '{category}' mangler helt i waste_fractions!")

#     def animal_weight(self, item_name):
#         """Søker etter dyrevekt. Krasjer hvis den ikke finnes."""
#         if hasattr(self, item_name):
#             return getattr(self, item_name)
#         if item_name in self.animal_weights:
#             return self.animal_weights[item_name]
#         raise KeyError(f"[STOPP] Dyrevekt for '{item_name}' mangler helt i animal_weights!")
        
    def get_table(self, sheet_name):
        """
        Returnerer et pandas DataFrame for et gitt ark i N_parameters.xlsx.
        Antar at første rad er header.
        """
        return pd.read_excel(self.filename, sheet_name=sheet_name)
    
    def get_trade_params(self):
        df = pd.read_excel(self.filename, sheet_name='trade_parameters')
        df['value'] = pd.to_numeric(df['value'], errors='coerce')
        
        # SJEKK FOR NY VS GAMMEL STRUKTUR:
        # Hvis den gamle usikkerhetskolonnen finnes, konverter den.
        if 'uncertainty' in df.columns:
            df['uncertainty'] = pd.to_numeric(df['uncertainty'], errors='coerce')
        
        # Hvis de nye kolonnene finnes, konverter dem også så de er klare for MC.
        if 'lower_bound' in df.columns:
            df['lower_bound'] = pd.to_numeric(df['lower_bound'], errors='coerce')
        if 'upper_bound' in df.columns:
            df['upper_bound'] = pd.to_numeric(df['upper_bound'], errors='coerce')
            
        df = df.set_index('param_id')
        return df

    def get_trade_mapping(self):
        """
        Mapping from SSB trade codes (Varenr) to N-content parameters.
        Sheet: 'trade_mapping'
        Columns (at least): type, konv, Varenr, År fra, År til, Varebetegnelse
        'konv' holds the param_id used in trade_N_parameters.
        """
        df = self.get_table('trade_mapping')
        return df

    def override_global_params(self, custom_dict):
        """
        Gjør det mulig for Monte Carlo-motoren å dytte inn simulerte 
        verdier før beregningene kjøres. Håndterer både eksisterende globale 
        parametere og nye, flate MC-nøkler (f.eks. prod_ og weight_).
        """
        for param_id, new_value in custom_dict.items():
            if param_id in self.global_params:
                # Hvis det er en eksisterende global parameter, oppdater den i dict-en
                self.global_params[param_id] = new_value
            else:
                # Hvis det er en ny, flat parameter (f.eks. prod_ eller weight_),
                # legger vi den direkte på objektet slik at .get() finner den!
                setattr(self, param_id, new_value)