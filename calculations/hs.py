#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 13 12:33:13 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl

from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
)
from calculations.shared_flow_calculations import (
    find_household_waste)

expected_years = EXPECTED_YEARS

params = NParameters("data_files/N_parameters.xlsx")
dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
waste_fracs = params.get_table('waste_fractions').set_index('waste_category')

def get_waste_frac(cat):
    row = waste_fracs.loc[cat]
    return float(row['N_frac']), float(row['uncertainty'])

paper_N,   u_paper   = get_waste_frac("paper")
plastic_N, u_plastic = get_waste_frac("plastic")
metal_N,   u_metal   = get_waste_frac("metal")
ewaste_N,  u_ewaste  = get_waste_frac("E_waste")
wood_N,    u_wood    = get_waste_frac("wood")
textile_N, u_textile = get_waste_frac("textiles")
haz_N,     u_haz     = get_waste_frac("hazardous")
mixed_N,   u_mixed   = get_waste_frac("mixed_waste")
other_N,   u_other   = get_waste_frac("other_materials")
constr_N,  u_constr  = get_waste_frac("construction_waste")
wet_N,     u_wet     = get_waste_frac("wet_organic")
park_N,    u_park    = get_waste_frac("park_garden")
contam_N,  u_contam  = get_waste_frac("contaminated_masses")

def execute_calculations():
    results = []
    
    # recyc = _add_recycling(results, dataset_unc)
    _add_mixed_household_waste(results, dataset_unc)
    _add_municipal_wastewater(results, dataset_unc)
    _add_NH3_emissions(results, dataset_unc)
    _add_LUC_NH3_emissions(results, dataset_unc)
    _add_LUC_NOx_emissions(results, dataset_unc)
    _add_LUC_N2O_emissions(results, dataset_unc)
    _add_overland_flow(results, dataset_unc)
    
    return results  # Returns a list of flow records


def _add_mixed_household_waste(results, dataset_unc):
    flow_code = 'HS.HS-PR.SO-Household waste-Nmix'
    collected_years = set()
    household_waste, unc_waste = find_household_waste(dataset_unc)
    data_sources = 'SSB'
    for year, value in household_waste.items():
        collected_years.add(year)   
        if year < 1995:
            comment = 'extrapolated'
        else:
            comment = 'ok'
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': unc_waste[year]        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_municipal_wastewater(results, dataset_unc):
    flow_code = 'HS.HS-PR.WW-Municipal wastewater-Nmix'
    collected_years = set()
    # using population data from SSB
    u_06913 = get_uncertainty(dataset_unc, '06913')
    N_amount, u_percap = params.get_global_param_with_uncertainty("per_capita_WW_N_load_kg")  
    uncertainty = combine_uncertainties_percent(u_06913, u_percap)
    data_sources = 'SSB table 06913'
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/06913_20251113-124117.xlsx')
    sheet = workbook['Folkemengde']
    # Industri og bergverk - avfall
    for row in range(37, 79):  
        year = int(sheet.cell(row=row, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=2).value)*N_amount*1e-6
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_NH3_emissions(results, dataset_unc):
    flow_code = 'HS.HS-AT.AT-Emissions-NH3'
    collected_years = set()
    # emissions from human body
    data_sources = 'SSB table 07459'
    comment = 'ok'
    # Schappi2025Ann, Eq. 46 p. 214. Equation from Sutton (2000)
    # FHS−AT (tN) = 1. 7 ∙ 10−5Ptotal + 1. 17 ∙ 10−5P<1y + 1. 46 ∙ 10−5P1−3y + 3. 4 ∙  10−9n
    # Ptotal: total population
    # P1y: population of children aged < 1 year
    # P13y: population of children aged 1-3 year2
    data = pd.read_excel('data_files/07459_20251119-151434.xlsx', sheet_name='Personer1', header=None, skiprows=3)
    data.columns = ['Gender', 'AgeGroup', 'Year', 'Value']
    data['Gender'] = data['Gender'].fillna(method='ffill')  
    data['AgeGroup'] = data['AgeGroup'].fillna(method='ffill')  
    age_0 = data[data['AgeGroup'] == '0 år']  # For age 0
    total_age_0 = age_0.groupby('Year')['Value'].sum().reset_index()
    age_1_3 = data[data['AgeGroup'].isin(['1 år', '2 år', '3 år'])]
    total_age_1_3 = age_1_3.groupby('Year')['Value'].sum().reset_index()
    total_population = data.groupby('Year')['Value'].sum().reset_index()
    population = total_age_0.merge(total_age_1_3, on='Year', suffixes=('_0', '_1_3'))
    population = population.merge(total_population, on='Year')
    population.columns = ['Year','Age0','Age1-3','Total']
    # n: average number of cigarettes smoked per year
    # cigarettes: use table 05307 that gives percentage of daily and occational smokers per year
    smoking = pd.read_excel('data_files/05307_20251119-152214.xlsx', sheet_name='Dagroyk', header=None, skiprows=3)
    smoking.columns = ['A','B','Year', 'Daily', 'Occ']  
    smoking = smoking[['Year', 'Daily', 'Occ']]  
    # assume daily smokers smoke 750 cigarettes per year, occasional smoke 2 per week (100 per year)
    u_pop = get_uncertainty(dataset_unc, '07459')
    u_smk = get_uncertainty(dataset_unc, '05307')    
    c_total,   u_c_total  = params.get_global_param_with_uncertainty("NH3_emission_factor_total_pop")
    c_age0,    u_c_age0   = params.get_global_param_with_uncertainty("NH3_emission_factor_age0")
    c_age1_3,  u_c_age13  = params.get_global_param_with_uncertainty("NH3_emission_factor_age1_3")
    c_smoke,   u_c_smoke  = params.get_global_param_with_uncertainty("NH3_emission_factor_cigarettes")
    cig_daily, u_cig_d    = params.get_global_param_with_uncertainty("daily_smoker_cigs_per_year")
    cig_occ,   u_cig_o    = params.get_global_param_with_uncertainty("occasional_smoker_cigs_per_year")  
    uncertainty = combine_uncertainties_percent(
        u_pop, u_smk,
        u_c_total, u_c_age0, u_c_age13, u_c_smoke,
        u_cig_d, u_cig_o)
    merged_data = population.merge(smoking, on='Year', how='inner')
    merged_data['Total_Smoked'] = (merged_data['Daily']*cig_daily + merged_data['Occ']*cig_occ)/100 * merged_data['Total']
    merged_data['Emissions'] = (c_total * merged_data['Total'] + c_age0 * merged_data['Age0'] + c_age1_3 * merged_data['Age1-3'] + c_smoke * merged_data['Total_Smoked'])    
    merged_data['Emissions'] = merged_data['Emissions']/1000 # tN to ktN
    for index, row in merged_data.iterrows():
        year = row['Year']                         # Extract the Year
        collected_years.add(year)
        value = row['Emissions']               # Extract Emissions    
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_LUC_NH3_emissions(results, dataset_unc):
    flow_code = 'HS.HS-AT.AT-LUC emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")  
    uncertainty = combine_uncertainties_percent(u_crltap, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '6')|(data[2] == '4E1')|(data[2] == '4E2')]
    data = data[data[3] == 'NH3']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_LUC_NOx_emissions(results, dataset_unc):
    flow_code = 'HS.HS-AT.AT-LUC emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '6')|(data[2] == '4E1')|(data[2] == '4E2')]
    data = data[data[3] == 'NOx']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

    
def _add_LUC_N2O_emissions(results, dataset_unc):
    flow_code = 'HS.HS-AT.AT-LUC emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 4
    # NOx given as NA
    data_sources = 'UNFCCC CRT'
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/N2O_NOx_HS_FS.xlsx')
    sheet = workbook['Ark1']
    for row in range(6, 39):  
        year = int(sheet.cell(row=row, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=2).value)*conv # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
    
def _add_overland_flow(results, dataset_unc):
    flow_code = 'HS.HS-HY.SW-Overland flow-Nmix'
    collected_years = set()
    # adding this flow for runoff from urban areas, given in TEOTIL
    data_sources = 'TEOTIL'
    comment = 'ok'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    ret, u_ret = params.get_global_param_with_uncertainty("HS_urban_retention_fraction")    
    # old TEOTIL (background * (1-ret)): combine TEOTIL + retention + interpolation
    uncertainty_early = combine_uncertainties_percent(u_teotil, u_ret, u_interp)
    # TEOTIL3 (direct source, no retention factor in code)
    uncertainty_recent = u_teotil    # prior to 2013: use TEOTIL data published on    https://www.miljodirektoratet.no/ansvarsomrader/overvaking-arealplanlegging/miljoovervaking/overvakingsprogrammer/forurensning-og-klimagasser/teotil/
    # use retention rate of 5% as found from recent teotil3 data
    workbook = openpyxl.load_workbook('data_files/Tilførsel av nitrogen til kystområdene fordelt på kilder.xlsx')
    sheet = workbook['Data fra Miljødirektoratet']
    comment = 'ok'
    for r in range(2, 25):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=5).value)/1000*(1-ret)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_early
        })
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet = workbook['totn_by_source']
    comment = 'ok'
    for r in range(2, 13):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=10).value)/1000 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_recent
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)




if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)
