#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov  6 11:34:21 2025

@author: anja
"""

import pandas as pd  
import openpyxl
import numpy as np

from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    fill_missing_with_mean,
    read_year_value_row,
    get_uncertainty,
    load_crltap_emissions_to_N,
)
from calculations.shared_flow_calculations import (
    find_industrial_crop_products,
    # find_manure_for_biofuel_production,
    find_non_edible_animal_products
    )

expected_years = EXPECTED_YEARS

CRLTAP_FILE = 'data_files/webdabData1863365.txt'

AG_SM_CRLTAP_SECTORS = [
    '3Da1','3Da2a','3Da2b','3Da2c','3Da3','3Da4',
    '3Db','3Dc','3De','3Df','4B1','4B2','4C1','4C2',
]

AG_MM_CRLTAP_SECTORS = [
    '3B1a','3B1b','3B2','3B3',
    '3B4a','3B4d','3B4e','3B4f',
    '3B4gi','3B4gii','3B4giii','3B4giv','3B4h',
]

GNB_FILE = 'data_files/aei_pr_gnb__custom_18744910_spreadsheet.xlsx'


def execute_calculations():
    """
    Compute all AG pool nitrogen flows.

    Returns
    -------
    list of dict
        Each dict has keys:
        'flow_name', 'year', 'value', 'comment', 'data_sources', 'uncertainty'.
    """

    results = []
    params = NParameters("data_files/N_parameters.xlsx")
    dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
        
    # NB! The order of execution below is important. 
    _add_food_crop_products_flow(results, params, dataset_unc)
    _add_industrial_crop_products_flow(results, params, dataset_unc)
    _add_fodder_crops_flow(results, params, dataset_unc)
    _add_NH3_emissions_soil_management(results, params, dataset_unc)
    _add_N2O_emissions_soil_management(results, params, dataset_unc)
    _add_NOx_emissions_soil_management(results, params, dataset_unc)
    _add_leaching_soil_management(results, params, dataset_unc)
    _add_leaching_manure_management(results, params, dataset_unc)
    _add_animal_products(results, params, dataset_unc)
    _add_non_edible_animal_products(results, params, dataset_unc)
    _add_manure_application(results, params, dataset_unc)
    # _add_manure_for_biofuel_production(results, params, dataset_unc)
    _add_NH3_emissions_manure_management(results, params, dataset_unc)
    _add_N2O_emissions_manure_management(results, params, dataset_unc)
    _add_NOx_emissions_manure_management(results, params, dataset_unc)
    _add_live_animal_export(results, params, dataset_unc)
    _add_N2_emissions_soil_management(results, params, dataset_unc)

    return results  # Returns a list of flow records

def _add_food_crop_products_flow(results, params, dataset_unc):
    flow_code = 'AG.SM-MP.FP-Food crop products-Nmix'
    collected_years = set()

    # use data from Gross nutrient balance, Eurostat as advised in Annexes: 
    # "Nutrient removal by harvest of crops" minus "Industrial crops"
    workbook = openpyxl.load_workbook(GNB_FILE)
    data_sources = 'Eurostat Gross nutrient balance'
    comment = 'ok'

    u_dataset = get_uncertainty(dataset_unc, 'Gross nutrient balance')
    uncertainty = u_dataset
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_interp = combine_uncertainties_percent(u_dataset, u_interp)

    # Sheet 26: total crops
    sheet = workbook['Sheet 26']  # nutrient removal by harvest of crops
    year_values = read_year_value_row(
        sheet,
        year_values=None,
        year_row=9,
        value_row=11,
        first_col=2,
        unit_factor=1.0e-3,
        op='+',
    )

    # Sheet 30: industrial crops (subtract)
    sheet = workbook['Sheet 30']  # nutrient removal by harvest of industrial crops
    year_values = read_year_value_row(
        sheet,
        year_values=year_values,
        year_row=9,
        value_row=11,
        first_col=2,
        unit_factor=-1.0e-3,  # subtract industrial crops
        op='+',
    )

    # Write results 
    for year, total_value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': total_value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

    # Fill missing years with mean, then explicit zeros
    mean_value = np.mean(list(year_values.values()))
    fill_missing_with_mean(
        flow_code, year_values, collected_years, results,
        mean=mean_value,
        comment=comment,
        data_sources='interpolated',
        uncertainty=uncertainty_interp,
    )
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_industrial_crop_products_flow(results, params, dataset_unc):
    flow_code = 'AG.SM-MP.OP-Crop products for industrial use-Nmix'
    collected_years = set()
    data_sources = 'Eurostat Gross nutrient balance, Nutrient removal by harvest of industrial crops'
    year_values, uncertainty = find_industrial_crop_products(dataset_unc,GNB_FILE)
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    # Write results 
    for year, value in year_values.items():
        collected_years.add(year)
        if year in range(2017,2020):
            comment = 'interpolated'
            unc = u_interp
        else:
            comment = 'ok'
            unc = uncertainty
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': unc,
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_fodder_crops_flow(results, params, dataset_unc):
    flow_code = 'AG.SM-AG.MM-Fodder crops-Nmix'
    collected_years = set()

    # This N flow covers N removal with grass and fodder production for livestock feeding. 
    # It includes grass from temporary and permanent pasture, as well as fodder plants harvested 
    # green (forage, e.g. green maize). National statistics should be used.

    # protein content
    # https://cdnmedia.eurofins.com/european-east/media/2709204/naeringsinnhold-i-grovfôr-til-droevtyggere.pdf, 
    # N per protein from https://www.fao.org/4/y5022e/y5022e03.htm
    fodder_prot, u_fodder = params.get_global_param_with_uncertainty("fodder_protein_frac")
    Jones, u_Jones = params.get_global_param_with_uncertainty("Jones_factor")
    N_content = fodder_prot / Jones

    u_13648 = get_uncertainty(dataset_unc, '13648')
    u_05772 = get_uncertainty(dataset_unc, '05772')
    uncertainty_13648 = combine_uncertainties_percent(u_fodder, u_Jones, u_13648)
    uncertainty_05772 = combine_uncertainties_percent(u_fodder, u_Jones, u_05772)

    comment = 'ok'

    # SSB table 13648
    workbook = openpyxl.load_workbook('data_files/13648_20251117-154625.xlsx')
    sheet = workbook['Avling']  # unit: 1000 ton crops
    data_sources = 'SSB table 13648'
    for col in range(2, 6):
        year = int(sheet.cell(row=4, column=col).value)
        collected_years.add(year)
        value1 = float(sheet.cell(row=5, column=col).value) * N_content  # Eng til slått
        value2 = float(sheet.cell(row=6, column=col).value) * N_content  # Grøntfôr- og silovekstar
        value = value1 + value2
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_13648,
        })
    # SSB table 05772
    workbook = openpyxl.load_workbook('data_files/05772_20251210-142618.xlsx')
    sheet = workbook['Gronfor']  # unit: 1000 ton crops
    data_sources = 'SSB table 05772'
    for col in range(2, 23):
        year = int(sheet.cell(row=3, column=col).value)
        collected_years.add(year)
        value1 = float(sheet.cell(row=4, column=col).value) * N_content  # Grøntfôr- og silovekstar
        value2 = float(sheet.cell(row=5, column=col).value) * N_content  # Høy
        value = value1 + value2
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_05772,
        })
    # Before 2000: data from SSB Jordbruksstatistikk
    workbook = openpyxl.load_workbook('data_files/grovfor_før_2000.xlsx')
    sheet = workbook['Ark1']  # unit: 1000 ton crops
    data_sources = 'SSB Jordbruksstatistikk'
    for r in range(3, 19):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value1 = float(sheet.cell(row=r, column=2).value) * N_content  # Grøntfôr- og silovekstar
        value2 = float(sheet.cell(row=r, column=3).value) * N_content  # Høy
        value = value1 + value2
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_05772,
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
def _add_NH3_emissions_soil_management(results, params, dataset_unc):    
    flow_code = 'AG.SM-AT.AT-Emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_dataset_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")    
    uncertainty = combine_uncertainties_percent(u_dataset_crltap, u_conv)
    sums = load_crltap_emissions_to_N(
        filename=CRLTAP_FILE,
        categories=AG_SM_CRLTAP_SECTORS,
        pollutant='NH3',
        conv_to_N=conv,
    )
    for year, val in sums.items():
        year = int(year)
        value = float(val)
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_N2O_emissions_soil_management(results, params, dataset_unc):    
    flow_code = 'AG.SM-AT.AT-Emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 3
    data_sources = 'UNFCCC CRT'
    u_dataset_n2o = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv_N2O, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_dataset_n2o, u_conv)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/N2O_NOx_AG.xlsx')
    sheet = workbook['Ark1']
    for row in range(5, 38):  
        year = int(sheet.cell(row=row, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=3).value)*conv_N2O # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty      
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_NOx_emissions_soil_management(results, params, dataset_unc):    
    flow_code = 'AG.SM-AT.AT-Emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_dataset_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_dataset_crltap, u_conv)
    sums = load_crltap_emissions_to_N(
        filename=CRLTAP_FILE,
        categories=AG_SM_CRLTAP_SECTORS,
        pollutant='NOx',
        conv_to_N=conv,
    )
    for year, val in sums.items():
        year = int(year)
        value = float(val)
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_leaching_soil_management(results, params, dataset_unc):    
    flow_code = 'AG.SM-HY.SW-Leaching-Nmix'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 3
    u_dataset_leach = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    uncertainty = u_dataset_leach
    data_sources = 'UNFCCC CRT'
    comment = 'ok'
    data = pd.read_csv('data_files/Nr_AG--HY.csv')
    data = data[['year', 'Nr_SM']]
    for index, row in data.iterrows():
        year = int(row['year'])
        collected_years.add(year)
        value = row['Nr_SM']
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_leaching_manure_management(results, params, dataset_unc):    
    flow_code = 'AG.MM-HY.SW-Leaching-Nmix'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 3
    u_dataset_leach = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    uncertainty = u_dataset_leach
    data_sources = 'UNFCCC CRT'
    comment = 'ok'
    data = pd.read_csv('data_files/Nr_AG--HY.csv')
    data = data[['year', 'Nr_MM']]
    for index, row in data.iterrows():
        year = int(row['year'])
        collected_years.add(year)
        value = row['Nr_MM']
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
def _add_animal_products(results, params, dataset_unc):
    flow_code = 'AG.MM-MP.FP-Animal products-Nmix'
    collected_years = set()
    # using data from FAO
    # all units: t product
    u_dataset_fao = get_uncertainty(dataset_unc, 'Crops and livestock products')
    comment = 'ok'
    data_sources = 'FAOSTAT Crops and livestock products'
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-18-2025.csv')
    filtered_data = data[
        (data['Element'] == 'Production') & 
        (data['Value'] != 0) & 
        ~data['Item'].str.contains('hides', case=False, na=False)]
    final_data = filtered_data[['Item', 'Year', 'Value']].copy()
    N_table = params.get_table('animal_products')
    # forventer kolonner: item, N_content_kg_per_t, (ev. source)
    N_table = N_table.set_index('item')
    N_table['uncertainty'] = pd.to_numeric(N_table['uncertainty'], errors='coerce')   # slå inn N_content på final_data via Item
    final_data = final_data.join(N_table['N_content_percent'],on='Item')
    final_data['N_content_kg_per_t'] = final_data['N_content_percent'].fillna(0)  
    # beregn N-mengde i kt N per år:
    # Value = tonn produkt
    final_data['N_amount_kt'] = final_data['Value'] * final_data['N_content_kg_per_t'] / 1e5
    used_items = final_data['Item'].unique()    
    # Parameter uncertainty: take max of uncertainties of used items (conservative)
    u_params = N_table.loc[used_items, 'uncertainty'].max()
    # Combined relative uncertainty for this flow
    uncertainty = combine_uncertainties_percent(u_dataset_fao, u_params)
    # summer per år
    total_N_per_year = (
        final_data
        .groupby('Year', as_index=False)['N_amount_kt']
        .sum())
    # fill only for years where FAO has data
    for year in expected_years:          # covers 1984–2025
        row = total_N_per_year[total_N_per_year['Year'] == year]
        if row.empty:
            # treat as missing; do not add to collected_years here
            continue
        collected_years.add(year)
        value = float(row['N_amount_kt'].iloc[0])
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_non_edible_animal_products(results, params, dataset_unc):
    flow_code = 'AG.MM-MP.OP-Non-edible animal products-Nmix'
    collected_years = set()
    # Schäppi advices using FAOSTAT Commodity Balances (non-food) but for Norway it only contains wool for 4 (random?) years. 
    # NB: approx 2500 t wool for those years, N-content 15 % (Schâppi2025Ann Table 24) gives 375 tN, comparable with other products, should be added
    # use raw hides and skins from FAOSTAT Crops and livestock products
    # all units: t product
    year_values = find_non_edible_animal_products(params)
    u_fao   = get_uncertainty(dataset_unc, 'Crops and livestock products')
    u_03710 = get_uncertainty(dataset_unc, '03710')
    u_LD_wool = get_uncertainty(dataset_unc, 'Landbruksdirektoratet_wool')   
    N_table = params.get_table('animal_products')
    N_table = N_table.set_index('item')
    N_table['uncertainty'] = pd.to_numeric(N_table['uncertainty'], errors='coerce')
    u_hides = get_uncertainty(N_table, 'Raw hides and skins')
    _, u_wool_per = params.get_global_param_with_uncertainty("wool_per_sheep")
    _, u_wool_N  = params.get_global_param_with_uncertainty("wool_N_frac")
    unc_fao_only = combine_uncertainties_percent(u_fao, u_hides)
    unc_with_wool = combine_uncertainties_percent(
        u_fao, u_hides,
        u_03710, u_LD_wool,
        u_wool_per, u_wool_N)
    for year, value in year_values.items():
        collected_years.add(year)
        comment = 'ok'
        if year > 2004 and year != 2001:
            data_sources = 'FAOSTAT Crops and livestock products + Landbruksdirektoratet'
            uncertainty = unc_with_wool
        elif year != 2001: # 2001 lacking data for sheep
            data_sources = 'FAOSTAT Crops and livestock products + Landbruksdirektoratet + SSB, extrapolated'
            uncertainty = unc_fao_only
        else: # interpolating for 2001
            data_sources = 'FAOSTAT Crops and livestock products'
            uncertainty = unc_fao_only
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_manure_application(results, params, dataset_unc):
    flow_code = 'AG.MM-AG.SM-Manure application-Nmix'
    collected_years = set()
    # use data from Gross nutrient balance, Eurostat as advised in Annexes
    u_dataset = get_uncertainty(dataset_unc, 'Gross nutrient balance')
    uncertainty = u_dataset
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_interp = combine_uncertainties_percent(u_dataset, u_interp)
    workbook = openpyxl.load_workbook(GNB_FILE)
    sheet = workbook['Sheet 12']
    comment = 'ok'
    year_values = read_year_value_row(sheet, year_values = None, year_row=9, value_row=11, first_col=2, unit_factor=1.0e-3, op='+')
    data_sources = 'Eurostat Gross nutrient balance, Manure input'
    for year, value in year_values.items():
        collected_years.add(year)
        if year == 2016:
            value_2016 = value
        elif year == 2020:
            value_2020 = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # 2017-2019: interpolate
    for year in range(2017,2020):
        collected_years.add(year)
        value = value_2016 + (value_2020-value_2016)/4*(year-2016)
        data_sources = 'interpolated'
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_interp
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


# def _add_manure_for_biofuel_production(results, params, dataset_unc):
#     flow_code = 'AG.MM-PR.SO-Manure for biofuel production-Nmix'
#     collected_years = set()
#     data_sources = 'Landbruksdirektoratet'
#     comment = 'ok'
#     year_values, uncertainty = find_manure_for_biofuel_production(params, dataset_unc)
#     for year, value in year_values.items():
#         collected_years.add(year)
#         results.append({
#             'flow_name': flow_code,
#             'year': year,
#             'value': value,
#             'comment': comment,
#             'data_sources': data_sources,
#             'uncertainty': uncertainty
#         })
#     missing_years = expected_years - collected_years
#     report_missing_years(flow_code, missing_years, results)

def _add_NH3_emissions_manure_management(results, params, dataset_unc):
    flow_code = 'AG.MM-AT.AT-Emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_dataset_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_dataset_crltap, u_conv)
    sums = load_crltap_emissions_to_N(
        filename=CRLTAP_FILE,
        categories=AG_MM_CRLTAP_SECTORS,
        pollutant='NH3',
        conv_to_N=conv,
    )
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_N2O_emissions_manure_management(results, params, dataset_unc):
    flow_code = 'AG.MM-AT.AT-Emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 3
    data_sources = 'UNFCCC CRT'
    u_dataset_n2o = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    conv_N2O, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_dataset_n2o, u_conv)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/N2O_NOx_AG.xlsx')
    sheet = workbook['Ark1']
    for row in range(5, 38):  
        year = int(sheet.cell(row=row, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=2).value)*conv_N2O # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_NOx_emissions_manure_management(results, params, dataset_unc):
    flow_code = 'AG.MM-AT.AT-Emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    u_dataset_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_dataset_crltap, u_conv)
    sums = load_crltap_emissions_to_N(
        filename=CRLTAP_FILE,
        categories=AG_MM_CRLTAP_SECTORS,
        pollutant='NOx',
        conv_to_N=conv,
    )
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
def _add_live_animal_export(results, params, dataset_unc):
    flow_code = 'AG.MM-RW.RW-Live animal export-Nmix'
    collected_years = set()
    # using data from FAO
    comment = 'ok'
    data_sources = 'FAOSTAT Crops and livestock products'
    u_dataset_fao = get_uncertainty(dataset_unc, 'Crops and livestock products')
    prot_frac, u_prot = params.get_global_param_with_uncertainty("live_animal_protein_frac")
    Jones, u_Jones    = params.get_global_param_with_uncertainty("Jones_factor")
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-12-2025.csv')
    filtered_data = data[(data['Element'] == 'Export quantity') & (data['Value'] != 0)]
    final_data = filtered_data[['Item', 'Year', 'Unit', 'Value']].copy()
    # Estimating typical weight (kg) of imported animals based on google search. Assume young animals so on the light side
    # Look up average live weight per animal from N_parameters.xlsx (sheet 'animal_weights')
    weights_table = params.get_table('animal_weights')
    # expect columns: item, avg_weight_kg, uncertainty
    weights_table = weights_table.set_index('item_name')
    weights_table['uncertainty'] = pd.to_numeric(weights_table['uncertainty'], errors='coerce')
    # params.animal_weight(item_name) returns avg_weight_kg
    final_data = final_data.join(weights_table[['avg_weight_kg', 'uncertainty']], on='Item')
    final_data.rename(columns={'uncertainty': 'weight_uncertainty'}, inplace=True)
    # assuming average 13 % protein in whole animal based on https://www.fao.org/4/X5557E/x5557e0a.htm
    # and Jones factor 6.25 for nitrogen to protein (standard) 
    final_data.loc[:, 'N_amount'] = (
        final_data['avg_weight_kg']
        * final_data['Value']   # number of animals
        * prot_frac
        * 1e-6                  # kg -> kt
        / Jones)    # final_data['N_amount'] = final_data['Item'].map(weight) * final_data['Value'] * .013 * 1e-6 / 6.25
    total_N_per_year = final_data.groupby('Year', as_index=False)['N_amount'].sum()
    u_weight = final_data['weight_uncertainty'].max()
    uncertainty = combine_uncertainties_percent(u_dataset_fao, u_prot, u_Jones, u_weight)
    # Fill only for years where FAO has data; let report_missing_years create zeros
    for year in expected_years:
        row = total_N_per_year[total_N_per_year['Year'] == year]
        if row.empty:
            # treat as missing; do not add to collected_years here
            continue
        collected_years.add(year)
        value = float(row['N_amount'].iloc[0])
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_N2_emissions_soil_management(results, params, dataset_unc):
    flow_code = 'AG.SM-AT.AT-Emissions-N2'
    collected_years = set()
    # Schäppi2025Ann recommends using a value of 14 kgN/ha for denitrification 
    # (estimated for agricultural soils in Germany) if no other data is available. 
    # Together with a total agricultural area of 1 132 693 ha, this gives around
    # 16 kgN/year for agricultural soils. 
    denit_N, u_denit = params.get_global_param_with_uncertainty("denitrification_AG_N2")
    value = denit_N
    uncertainty = u_denit    
    comment = 'ok'
    data_sources = 'Schäppi2025Ann + NIBIO'
    for year in expected_years:
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,  
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)




# Example usage
if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)

