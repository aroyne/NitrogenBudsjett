#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 22 15:57:29 2026

@author: anja
"""

# calculations/utils.py

import numpy as np
import pandas as pd
import openpyxl


start_year = 1984
end_year = 2025
EXPECTED_YEARS = set(range(start_year, end_year + 1))



def report_missing_years(flow_code, missing_years, results,
                         comment='not done', data_sources='no data',
                         default_value=np.nan):
    """
    Append zero flows for missing_years to results.

    Parameters
    ----------
    flow_code : str
    missing_years : iterable of int
    results : list of dict  (modified in-place)
    comment : str
    data_sources : str
    default_value : float
    """
    for year in missing_years:
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': default_value,
            'comment': comment,
            'data_sources': data_sources,
        })


def combine_uncertainties_percent(*percs):
    """
    Combine multiple independent relative uncertainties (in %) in quadrature.
    None values are ignored. Returns total in % (float).
    """
    vals = [p for p in percs if p is not None]
    if not vals:
        return None
    return (sum((p / 100.0) ** 2 for p in vals) ** 0.5) * 100.0


def fill_missing_with_mean(flow_code, year_values, collected_years, results,
                           mean=None, comment='ok', data_sources='interpolated',
                           uncertainty=None, expected_years=EXPECTED_YEARS):
    """
    Fill missing years (expected_years - collected_years) with a constant mean.

    year_values : dict {year: value} for years with data.
    collected_years : set, modified in-place.
    """
    if not year_values:
        return
    if mean is None:
        mean = float(np.mean(list(year_values.values())))
    missing_years = expected_years - collected_years
    for year in sorted(missing_years):
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': mean,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

def read_year_value_row(sheet,
                        year_values=None,
                        year_row=9,
                        value_row=11,
                        first_col=2,
                        unit_factor=1.0,
                        op='+'):
    """
    Read year/value pairs from one row of an Excel sheet and accumulate
    them into a dict.

    Parameters
    ----------
    sheet : openpyxl worksheet
    year_values : dict or None
        If None, a new dict is created. If dict, updated in place.
        Keys: int year, values: float.
    year_row : int
        Row index containing the years.
    value_row : int
        Row index containing the values corresponding to the years.
    first_col : int
        First column index to read (default 2 = column B).
    unit_factor : float
        Multiplier applied to each numeric value (e.g. 1e-3 to convert t → kt).
    op : str
        How to combine with existing values:
        '+' (default) → add to existing;
        'replace' → overwrite existing.

    Returns
    -------
    dict
        The updated dict mapping year → value.
    """
    if year_values is None:
        year_values = {}

    for col in range(first_col, sheet.max_column + 1):
        year = sheet.cell(row=year_row, column=col).value
        value = sheet.cell(row=value_row, column=col).value
        if year is None or value in (None, ':'):
            continue

        year = int(year)
        val = float(value) * unit_factor

        if op == '+':
            year_values[year] = year_values.get(year, 0.0) + val
        elif op == 'replace':
            year_values[year] = val
        else:
            raise ValueError("op must be '+' or 'replace'")

    return year_values

def read_crltap_sum(file_path, sectors, pollutant):
    """
    Read a CRLTAP webdab file and return annual summed emissions
    for a given pollutant and a set of sectors.

    Parameters
    ----------
    file_path : str
        Path to the CRLTAP txt file.
    sectors : iterable of str
        Sector codes (column 2) to include, e.g. ['3Da1','3Da2a',...].
    pollutant : str
        Pollutant code in column 3, e.g. 'NH3' or 'NOx'.

    Returns
    -------
    pandas.Series
        Index: year (int, from column 1).
        Values: summed emissions (float, from column 5).
    """
    data = pd.read_csv(file_path, sep=';', header=None, skiprows=4)
    data = data[data[2].isin(sectors)]
    data = data[data[3] == pollutant]
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    return data.groupby(1)[5].sum()

def get_uncertainty(dataset_unc, name, default=None):
    """
    Safely retrieve an uncertainty (as float) from the dataset_unc table.

    Parameters
    ----------
    dataset_unc : pandas.DataFrame
        DataFrame indexed by dataset_name, with a column 'uncertainty'.
    name : str
        Dataset name (index) to look up.
    default : float or None, optional
        If provided, return this value when 'name' is missing.
        If None, re-raise KeyError.

    Returns
    -------
    float
        The uncertainty value.
    """
    try:
        return float(dataset_unc.loc[name, 'uncertainty'])
    except KeyError:
        if default is None:
            raise
        return float(default)
    

import pandas as pd

def load_crltap_emissions_to_N(raw_lines, categories, pollutant, conv_to_N, dataset_noise, noise_key='CRLTAP', skiprows=4, sep=';'):
    """
    MC-VERSJON: Laster CRLTAP/WebDab utslippsdata fra ferdiginnlastede tekstlinjer i RAM.
    Speiler nøyaktig logikken til den opprinnelige read_csv-baserte funksjonen.
    Påfører parameterstøy (conv_to_N) og asymmetrisk datasettstøy robust.
    """
    # 1. Klargjør asymmetrisk datasettstøy
    has_noise = dataset_noise and noise_key in dataset_noise
    noise_val = dataset_noise[noise_key]
    
    # 2. Prosesser tekstlinjene fra RAM (hopp over skiprows)
    valid_lines = raw_lines[skiprows:] if len(raw_lines) > skiprows else []
    
    # Vi akkumulerer råverdier pr år i en ordbok
    yearly_raw_sums = {}
    
    for line in valid_lines:
        if not line.strip():
            continue
            
        parts = line.split(sep)
        if len(parts) < 6:  # Vi må ha minst opp til indeks 5
            continue
            
        try:
            cat_val = parts[2].strip()
            poll_val = parts[3].strip()
            
            # Sjekk om raden matcher kategori og stoff (tilsvarer .isin() og ==)
            if cat_val in categories and poll_val == pollutant:
                year = int(parts[1].strip())
                
                # Konverter verdi (tilsvarer to_numeric med coerce og fillna(0))
                val_str = parts[5].strip()
                raw_val = float(val_str) if (val_str and val_str != '-') else 0.0
                
                yearly_raw_sums[year] = yearly_raw_sums.get(year, 0.0) + raw_val
        except (ValueError, IndexError):
            continue

    # 3. Konverter til N og påfør støy matematisk korrekt for hvert år
    final_sums = {}
    for year, raw_sum in yearly_raw_sums.items():
        # Grunnverdi etter konverteringsfaktor (f.eks. NH3 -> N)
        base_value = raw_sum * conv_to_N
        value = base_value * noise_val
            
        if value < 0:
            value = 0.0
            
        final_sums[year] = value
        
    return final_sums

def read_trade_data(trade_file):
    trade_data = pd.read_csv(trade_file, sep=';', header=None)
    trade_columns = ['year', 'impeks', 'HS_code', 'country', 'value_code', 'amount', 'value_2', 'value_3']
    trade_data.columns = trade_columns
    return trade_data
    
def find_trade_data(trade_data, hs_N_content, impeks):
    # Only keep entries with N-content that is not zero
    N_dict = {
        row['HS-code']: row['N-content']
        for index, row in hs_N_content.iterrows()
        if row['N-content'] != 0
    }
    hs_codes = list(N_dict.keys())
    filtered_trade_data = trade_data[
                                  (trade_data['HS_code'].isin(hs_codes))& 
                                  (trade_data['impeks'].isin([impeks]))]
    sum_by_year = filtered_trade_data.groupby(['HS_code', 'year'])['amount'].sum().reset_index()
    sum_by_year.loc[:,'N_content'] = sum_by_year['HS_code'].map(N_dict)
    sum_by_year.loc[:,'N_amount'] = (sum_by_year['amount'] * sum_by_year['N_content']) / 1e6 # to go from kg to ktN
    aggregated_data = sum_by_year.groupby('year', as_index=False)['N_amount'].sum()
    return aggregated_data

# def find_trade_flow(trade_data, trade_mapping, trade_params, types_to_keep, impeks, dataset_unc, wide = False):
#     # Mapping: filter types we want
#     mapping_subset = trade_mapping[trade_mapping['type'].isin(types_to_keep)].copy()    
#     data_impeks = trade_data[trade_data['impeks'] == impeks].copy()    
#     # Join trade_data with mapping: HS_code <-> Varenr
#     data_impeks = data_impeks.merge(
#         mapping_subset[['Varenr', 'konv']],   # 'konv' holds param_id
#         left_on='HS_code',
#         right_on='Varenr',
#         how='inner'
#     )    
#     # Join with trade_N_parameters to get N-content (kg N/kg) and param uncertainty
#     data_impeks = data_impeks.merge(
#         trade_params[['value', 'uncertainty']],
#         left_on='konv',         # konv is param_id
#         right_index=True,
#         how='left'
#     )   
#     # Compute N_amount in kt N
#     data_impeks['N_amount'] = data_impeks['amount'] * data_impeks['value'] / 1e6  # kg -> kt    
#     # Aggregate per year
#     aggregated_data = data_impeks.groupby('year', as_index=False)['N_amount'].sum()
#     # Dataset uncertainty (trade statistics)
#     u_dataset_trade = get_uncertainty(dataset_unc, '08801')    
#     # Parameter uncertainty: max of the parameters used in this flow (conservative)
#     used_param_ids = data_impeks['konv'].unique()
#     u_params = trade_params.loc[used_param_ids, 'uncertainty'].max()    
#     # Combined relative uncertainty for this flow
#     uncertainty = combine_uncertainties_percent(u_dataset_trade, u_params)
    # return aggregated_data, uncertainty
    
def find_trade_flow(
    data_impeks,            # <--- Tar nå imot data som allerede er filtrert på impeks og merget med mapping!
    trade_params,
    dataset_unc,
    wide=False,
):
    # Join med trade_N_parameters for å få N-innhold (kg N/kg) og usikkerhet
    # (Dette kjører superraskt fordi data_impeks nå kun har de få radene som gjelder denne strømmen)
    cols_to_include = ['value', 'uncertainty'] if 'uncertainty' in trade_params.columns else ['value']
    
    data_impeks = data_impeks.merge(
        trade_params[cols_to_include],
        left_on='konv',        # konv is param_id
        right_index=True,
        how='left'
    )

    # Beregn N_amount i kt N
    data_impeks['N_amount'] = data_impeks['amount'] * data_impeks['value'] / 1e6  # kg -> kt

    # Aggreger per år og type
    aggregated_data_type = (
        data_impeks
        .groupby(['year', 'type'], as_index=False)['N_amount']
        .sum()
    )

    if wide:
        aggregated_wide = (
            aggregated_data_type
            .pivot(index='year', columns='type', values='N_amount')
            .reset_index()
        )
        aggregated_wide.columns.name = None
        aggregated_data = aggregated_wide
    else:
        aggregated_data = (
            aggregated_data_type
            .groupby(['year'], as_index=False)['N_amount']
            .sum()
        )

    # MC-TILPASNING FOR USIKKERHET
    if dataset_unc is None:
        uncertainty = 0.0
    else:
        try:
            u_dataset_trade = get_uncertainty(dataset_unc, '08801')
            unc_col = 'uncertainty' if 'uncertainty' in trade_params.columns else 'upper_bound'
            used_param_ids = data_impeks['konv'].unique()
            u_params = trade_params.loc[used_param_ids, unc_col].max()
            uncertainty = combine_uncertainties_percent(u_dataset_trade, u_params)
        except Exception:
            uncertainty = 0.0

    return aggregated_data, uncertainty


def process_generic_trade_flow(preloaded_data, current_params, current_trade_factors, 
                               target_types, is_import=True, dataset_noise=None, flow_code=None, results=None, data_sources='SSB tab 08801'):
    """
    MC-KLAR GENERISK KJERNEFUNKSJON: Prosesserer handelsstrømmer fra SSB 08801.
    Nå fullstendig koblet mot det sentrale usikkerhetssystemet (dataset_noise).
    """
    df_vol = preloaded_data.get('compressed_trade_volume')
    if df_vol is None or current_trade_factors is None:
        if flow_code:
            print(f"[ADVARSEL] Mangler handelsdata eller faktorer for {flow_code}.")
        return {}

    # --- 1. KORREKT HÅNDTERING AV DATASETT-STØY FOR SSB 08801 ---
    dataset_key = '08801' 
    noise_val = dataset_noise[dataset_key]
    
    # 2. Sørg for at target_types er et sett med små bokstaver
    if isinstance(target_types, (str, int, float)):
        types_to_match = {str(target_types).lower().strip()}
    else:
        types_to_match = {str(t).lower().strip() for t in target_types}

    # 3. Filtrer ut gitte typer og riktig handelsretning (1 = Import, 2 = Eksport)
    impeks_target = ['1', '1.0'] if is_import else ['2', '2.0']
    is_correct_type = df_vol['type'].astype(str).str.lower().str.strip().isin(types_to_match)
    is_correct_direction = df_vol['impeks'].astype(str).str.strip().isin(impeks_target)
    
    df_filtered = df_vol[is_correct_type & is_correct_direction].copy()
    
    flow_dict = {}
    
    if not df_filtered.empty:
        # 4. Map inn unike, perturberte N-faktorer (støy på produkt-attributt)
        v_factors = df_filtered['konv'].astype(str).str.strip().map(current_trade_factors).fillna(0.0)
        
        # 5. Påfør datasett-støy vektorisert (støy på mengde/aktivitetsnivå)
        df_filtered['perturbed_amount'] = df_filtered['amount'] * noise_val
            
        # Sluttberegning: [Mengde med støy] * [N-faktor med støy] / 1e6 (kg -> kt N)
        df_filtered['N_amount'] = df_filtered['perturbed_amount'] * v_factors / 1e6 
        
        # Fysisk sperre mot negative mengder etter støy
        df_filtered.loc[df_filtered['N_amount'] < 0, 'N_amount'] = 0.0

        # 6. Aggreger nitrogenmengde per år dynamisk
        for yr_raw, amt in zip(df_filtered['year'].values, df_filtered['N_amount'].values):
            try:
                yr = int(float(yr_raw))
                flow_dict[yr] = flow_dict.get(yr, 0.0) + amt
            except (ValueError, TypeError):
                continue

    # 7. Pakk inn i det offisielle formatet for tidsserier hvis results er sendt med
    if results is not None and flow_code is not None:
        collected_years = set()
        for year in EXPECTED_YEARS:
            if year in flow_dict:
                collected_years.add(year)
                results.append({
                    'flow_name': flow_code,
                    'year': year,
                    'value': float(flow_dict[year]), 
                    'comment': 'ok (MC-støy påført mengde og N-fraksjon)',
                    'data_sources': data_sources
                })
        report_missing_years(flow_code, EXPECTED_YEARS - collected_years, results)

    return flow_dict