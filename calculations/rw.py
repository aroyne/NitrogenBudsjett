#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 11 16:16:58 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl
from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
    read_trade_data,
    find_trade_data,
    find_trade_flow
)
from calculations.shared_flow_calculations import (
    find_ammonia_import,
    find_aquaculture_production,
    find_other_goods_import
    )

expected_years = EXPECTED_YEARS

params = NParameters("data_files/N_parameters.xlsx")
dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
waste_fracs = params.get_table('waste_fractions').set_index('waste_category')
trade_data = read_trade_data('data_files/Tab_08801_1988_2024.csv')



def get_waste_frac(cat):
    row = waste_fracs.loc[cat]
    # N_frac is in kg N/kg, uncertainty is in %
    return float(row['N_frac']), float(row['uncertainty'])

def execute_calculations():
    results = []
    trade_params = params.get_trade_params()      # index: param_id
    trade_mapping = params.get_trade_mapping()    # columns: type, konv, Varenr, ...
    
    _add_fuel_import(results, dataset_unc, trade_data)
    _add_transport_fuel_import(results, dataset_unc, trade_data)
    _add_solid_waste_import(results, dataset_unc, trade_data, trade_mapping,trade_params)
    _add_atmospheric_inflow_OXN(results, dataset_unc)
    _add_atmospheric_inflow_RDN(results, dataset_unc)
    _add_food_import(results, dataset_unc, trade_data,trade_mapping, trade_params)
    _add_other_goods_import(results, dataset_unc, trade_mapping, trade_params, trade_data)
    _add_ammonia_import(results, dataset_unc, trade_mapping, trade_params)
    _add_animal_feed_import(results, dataset_unc)
    _add_aquaculture_feed_import(results, dataset_unc)
    _add_live_animal_import(results, dataset_unc)
    _add_mineral_fertilizer_import(results, dataset_unc)

    return results  # Returns a list of flow records


def _add_fuel_import(results, dataset_unc, trade_data):
    flow_code = 'RW.RW-EF.EC-Fuel import-Nmix'
    collected_years = set()
    # using trade data from SSB
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    u_08801 = get_uncertainty(dataset_unc, '08801')
    uncertainty = u_08801  
    # HS-koder for energivarer starter på 27. 
    # Importing N-contents for HS codes
    hs_N_content = pd.read_excel('data_files/N_content_fuels.xlsx')
    # only the ones not labeled T under "transport?"
    hs_N_content = hs_N_content[hs_N_content['transport?'].isna()]
    hs_N_content['N-content'] *= 1e-2 # from weight % to frac
    impeks = 1 # 1 for import, 2 for export
    aggregated_data = find_trade_data(trade_data, hs_N_content, impeks)
    for year in expected_years:
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_transport_fuel_import(results, dataset_unc, trade_data):
    flow_code = 'RW.RW-EF.TR-Import of transport fuel-Nmix'
    collected_years = set()
    # using trade data from SSB
    u_08801 = get_uncertainty(dataset_unc, '08801')
    uncertainty = u_08801
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    # HS-koder for energivarer starter på 27. 
    # Importing N-contents for HS codes
    hs_N_content = pd.read_excel('data_files/N_content_fuels.xlsx')
    # only the ones not labeled T under "transport?"
    hs_N_content = hs_N_content[hs_N_content['transport?'] == 'T']
    hs_N_content['N-content'] *= 1e-2 # from weight % to frac
    impeks = 1 # 1 for import, 2 for export
    aggregated_data = find_trade_data(trade_data, hs_N_content, impeks)
    for year in expected_years:
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_solid_waste_import(results, dataset_unc, trade_data, trade_mapping,trade_params):
    flow_code = 'RW.RW-PR.SO-Solid waste import-Nmix'
    collected_years = set()
    # N-content from Table 50, Schâppi2025Ann
    # municipal waste: HS code 382510; 0.9 % (household and similar wastes) 
    # sewage sludge: HS code 382520; 4,40 % (common sludges)
    comment = 'ok'
    data_sources = 'SSB tab 08801'
    types_to_keep = ['kommunalt_avfall','annet_avfall','slam','farlig_avfall','tekstilavfall','plastavfall','papiravfall',]
    impeks = 1
    aggregated_data, uncertainty = find_trade_flow(trade_data, trade_mapping, trade_params, types_to_keep, impeks, dataset_unc, wide = False)
    # oversikt over detaljerte varekategorier lastet ned herfra: https://www.ssb.no/utenriksokonomi/utenrikshandel/artikler/statistisk-varefortegnelse-for-utenrikshandelen
    # hs_codes = ['38251000_2002','38252000_2002']
    # mixed_N,  u_mixed_export  = get_waste_frac("mixed_waste")
    # common_N, u_common_export = get_waste_frac("common_sludge")
    # N_values = [mixed_N, common_N]
    # hs_N_content = pd.DataFrame({"HS-code": hs_codes,"N-content": N_values})    
    # u_waste_export = max(u_mixed_export, u_common_export)
    # u_08801 = get_uncertainty(dataset_unc, '08801')
    # uncertainty = combine_uncertainties_percent(u_08801, u_waste_export)
    # impeks = 1 # 1 for import, 2 for export
    # aggregated_data = find_trade_data(trade_data, hs_N_content, impeks)
    for year in range (1988,2025):
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_atmospheric_inflow_OXN(results, dataset_unc):
    flow_code = 'RW.RW-AT.AT-Atmospheric inflow-OXN'
    collected_years = set()
    # using source-receptor data from EMEP, according to Schäppi2025Ann
    u_sr = get_uncertainty(dataset_unc, 'Source-receptor')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')    
    uncertainty_base = u_sr
    uncertainty_interp = combine_uncertainties_percent(u_sr, u_interp)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/atm_in_out.xlsx')
    sheet = workbook['Ark1']    
    for r in range(6,46):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)/10 # from 100 tN to ktN  
        if sheet.cell(row=r, column=6).value == 'interpolated':
            data_sources = 'interpolated'
            uncertainty = uncertainty_interp
        else:
            data_sources = 'EMEP SR tables'
            uncertainty = uncertainty_base
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)



def _add_atmospheric_inflow_RDN(results, dataset_unc):
    flow_code = 'RW.RW-AT.AT-Atmospheric inflow-RDN'
    collected_years = set()
    # using source-receptor data from EMEP, according to Schäppi2025Ann
    u_sr = get_uncertainty(dataset_unc, 'Source-receptor')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')   
    uncertainty_base = u_sr
    uncertainty_interp = combine_uncertainties_percent(u_sr, u_interp)
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/atm_in_out.xlsx')
    sheet = workbook['Ark1']    
    for r in range(6,46):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=4).value)/10 # from 100 tN to ktN  
        if sheet.cell(row=r, column=6).value == 'interpolated':
            data_sources = 'interpolated'
            uncertainty = uncertainty_interp
        else:
            data_sources = 'EMEP SR tables'
            uncertainty = uncertainty_base
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)



def _add_food_import(results, dataset_unc, trade_data,trade_mapping, trade_params):
    flow_code = 'RW.RW-MP.FP-Food import-Nmix'
    collected_years = set()
    # using trade data from SSB
    u_08801 = get_uncertainty(dataset_unc, '08801')
    uncertainty = u_08801
    data_sources = 'SSB'
    comment = 'ok'
    types_to_keep = ['korn/planter', 'kjøtt/fisk/meieri/egg', 'mat']
    impeks = 1
    aggregated_data, uncertainty = find_trade_flow(trade_data, trade_mapping, trade_params, types_to_keep, impeks, dataset_unc, wide = False)
    for year in range (1988,2025):
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    

def _add_other_goods_import(results, dataset_unc, trade_mapping, trade_params, trade_data):
    flow_code = 'RW.RW-MP.OP-Other goods import -Nmix'
    collected_years = set()
    data_sources = 'SSB'
    comment = 'ok'
    year_values, uncertainty = find_other_goods_import(dataset_unc, trade_mapping, trade_params, trade_data)
    for year, value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_ammonia_import(results, dataset_unc, trade_mapping, trade_params):
    flow_code = 'RW.RW-MP.OP-Ammonia import -Nmix'
    collected_years = set()
    data_sources = 'SSB'
    comment = 'ok'
    year_values, uncertainty = find_ammonia_import(dataset_unc, trade_mapping, trade_params)
    for year, value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
  
def _add_animal_feed_import(results, dataset_unc):
    flow_code = 'RW.RW-AG.MM-Animal feed import-Nmix'
    collected_years = set()
    N_content_carb, u_carb = params.get_global_param_with_uncertainty("feed_carb_N_frac")
    N_content_prot, u_prot = params.get_global_param_with_uncertainty("feed_prot_N_frac")
    u_kraft = get_uncertainty(dataset_unc, 'Kraftforstatistikk')
    uncertainty = combine_uncertainties_percent(u_kraft, u_carb, u_prot)
    comment = 'ok'
    data_sources = 'Landbruksdirektoratets kraftfôrstatistikk - årlig råvareforbruk'
    workbook = openpyxl.load_workbook('data_files/Årlig råvareforbruk.xlsx')
    sheet = workbook['Varegrupper']    
    N_cont = 0
    i = 1
    for r in range(5,30):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value_carb = sheet.cell(row=r, column=3).value  # importert karbohydratråvare  
        value_prot = sheet.cell(row=r, column=9).value  # importert proteinråvare  
        imported_feed_N = (value_carb*N_content_carb + value_prot*N_content_prot)/1000 # t to kt   
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': imported_feed_N,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
        N_cont += imported_feed_N/(value_carb+value_prot)
        i += 1
    N_cont_before_2000 = N_cont/i*1e3
    # years before 2000
    comment = 'ok'
    data_sources = 'NIBIO Totalkalkylen'
    workbook = openpyxl.load_workbook('data_files/NibioStatistics-4.xlsx')
    uncertainty = get_uncertainty(dataset_unc, 'Totalkalkylen') # såkorn (cereal seed)
    # uncertainty = combine_uncertainties_percent(u_seed_cereal_N, u_seed_oil_N, u_seed_pea_N,u_seed_grass_N, u_seed_rootveg_N,u_dataset)
    sheet = workbook['Sum innkjøpt kraftfôr ukorr.']
    for r in range(28, 43):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value) # tonn kraftfôr
        if year < 1995:            
            dom_frac = float(sheet.cell(row=r, column=5).value)
            comment = 'ok'
        else:
            dom_frac = 0.694
            comment = 'interpolated'            
        value *= 1e-3 * N_cont_before_2000  * (1-dom_frac)     # tons → kt, then * N fraction
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, #converting from t to kt
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_aquaculture_feed_import(results, dataset_unc):
    flow_code = 'RW.RW-HY.AC-Aquaculture feed import-Nmix'
    collected_years = set()
    # Feed for aquaculture
    import_fraction, u_import = params.get_global_param_with_uncertainty("aquafeed_import_fraction")
    fish_N_frac, u_fish_N = params.get_global_param_with_uncertainty("fish_N_frac")
    prot_ret, u_ret = params.get_global_param_with_uncertainty("aquafeed_N_retention")
    feed_waste, u_waste = params.get_global_param_with_uncertainty("aquafeed_waste_fraction")    
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    u_fisk = get_uncertainty(dataset_unc, 'Fiskeridirektoratet')   
    aquaculture_production, _ = find_aquaculture_production(dataset_unc)
    uncertainty = combine_uncertainties_percent(u_fisk, u_import, u_fish_N, u_ret, u_waste)
    for year, value in aquaculture_production.items():
        collected_years.add(year)
        # according to Schäppl2025Ann, p254: N content in fish and shellfish: 2.8% according to UNECE Guidance, Annex 6 Table 12
        eaten_feed_N = value / prot_ret        # kt N in eaten feed
        total_feed_N = eaten_feed_N / (1 - feed_waste)
        imported_feed_N = total_feed_N * import_fraction
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': imported_feed_N,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_live_animal_import(results, dataset_unc):
    flow_code = 'RW.RW-AG.MM-Live animal import-Nmix'
    collected_years = set()
    # using data from FAO
    comment = 'ok'
    data_sources = 'FAOSTAT Crops and livestock products'
    u_fao = get_uncertainty(dataset_unc, 'Crops and livestock products')
    prot_frac, u_prot = params.get_global_param_with_uncertainty("live_animal_protein_frac")
    prot_to_N, u_Jones = params.get_global_param_with_uncertainty("Jones_factor")    
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-12-2025.csv')
    filtered_data = data[(data['Element'] == 'Import quantity') & (data['Value'] != 0)]
    final_data = filtered_data[['Item', 'Year', 'Unit', 'Value']].copy()
    u_fao = get_uncertainty(dataset_unc, 'Crops and livestock products')
    prot_frac, u_prot = params.get_global_param_with_uncertainty("live_animal_protein_frac")
    prot_to_N, u_Jones = params.get_global_param_with_uncertainty("Jones_factor")    # Estimating typical weight (kg) of imported animals based on google search. Assume young animals so on the light side
    weights_table = params.get_table('animal_weights')
    # expected columns: item, avg_weight_kg, uncertainty
    weights_table = weights_table.set_index('item_name')
    weights_table['uncertainty'] = pd.to_numeric(weights_table['uncertainty'], errors='coerce')   
    final_data = final_data.join(weights_table[['avg_weight_kg', 'uncertainty']], on='Item')
    final_data.rename(columns={'uncertainty': 'weight_uncertainty'}, inplace=True)
    final_data.loc[:, 'N_amount'] = (
        final_data['avg_weight_kg']
        * final_data['Value']   # number of animals
        * prot_frac
        * 1e-6                  # kg -> kt
        / prot_to_N
    )
    u_weight = final_data['weight_uncertainty'].max()
    uncertainty = combine_uncertainties_percent(u_fao, u_prot, u_Jones, u_weight)
    # final_data['N_amount'] = final_data['Item'].map(weight) * final_data['Value'] * .013 * 1e-6 / 6.25
    total_N_per_year = final_data.groupby('Year', as_index=False)['N_amount'].sum()
    for year in range (1984,2025):
        if total_N_per_year['Year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = total_N_per_year[total_N_per_year['Year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_mineral_fertilizer_import(results, dataset_unc):
    flow_code = 'RW.RW-AG.SM-Mineral fertilizer import-Nmix'
    collected_years = set()
    # using data from FAO
    u_fert = get_uncertainty(dataset_unc, 'Fertilizer by nutrient')
    uncertainty = u_fert    
    comment = 'ok'
    data_sources = 'FAOSTAT Fertilizer by Nutrient'
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-12-2025-2.csv')
    filtered_data = data[(data['Element'] == 'Import quantity') & (data['Value'] != 0)]
    final_data = filtered_data[['Year', 'Value']]
    for year in range (1984,2024):
        if final_data['Year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = final_data[final_data['Year'] == year]
            value = n_amount_row['Value'].values[0]/1000 # from t to kt
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)




if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)
