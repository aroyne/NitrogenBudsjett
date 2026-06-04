#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Nov  4 14:18:59 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
    read_trade_data,
    find_trade_flow,
)
from calculations.shared_flow_calculations import (
    find_aquaculture_production,
    find_feedstock_fuel,
    find_food_industry_waste,
    find_industrial_round_wood,
    find_industrial_waste_fuels,
    find_other_goods_export,
    find_other_goods_import,
    find_other_industry_waste,
    find_other_industry_wastewater,
    find_op_untreatewd_wastewater,
    find_recycling,
    find_industrial_crop_products,
    find_non_edible_animal_products
    )

expected_years = EXPECTED_YEARS

params = NParameters("data_files/N_parameters.xlsx")
dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
trade_data = read_trade_data('data_files/Tab_08801_1988_2024.csv')



def protein_per_group(params, mapping_sheet, group_index):
    """
    Gir proteinfraksjon (kg protein / kg mat) for hver kode i group_index.

    params        : NParameters-objekt
    mapping_sheet : 'protein_map_new' eller 'protein_map_old'
    group_index   : mengde.index (SSB-koder)
    """

    # 1) Les master-tabellen med matvaretyper (protein_food_items)
    items = params.get_table('protein_food_items')
    # forventer kolonner: food_group, protein_content, source
    items = items.set_index('food_group')

    # 2) Les mapping (kode -> food_group)
    mapping = params.get_table(mapping_sheet)
    # forventer kolonner: code, food_group
    mapping = mapping.set_index('code')

    # 3) Tilpass mapping til de kodene vi faktisk har i mengde
    mapping = mapping.reindex(group_index)

    # 4) Join for å få protein_content per kode
    joined = mapping.join(items[['protein_content']], on='food_group')

    # 5) NaN (manglende) -> 0, og prosent -> fraksjon
    protein_pct = joined['protein_content'].fillna(0)
    protein_frac = protein_pct / 100.0

    return protein_frac


def execute_calculations():
    results = []
    trade_params = params.get_trade_params()      # index: param_id
    trade_mapping = params.get_trade_mapping()    # columns: type, konv, Varenr, ...
    
    # initializing values for mass balance
    years = list(range(1984, 2026))  # 1984..2025 inclusive
    OP_in = pd.DataFrame({
        'year': years,
        'value': 0.0,               # float zeros; use 0 if you want ints
        'entries': 0    # need to count the right number of entries that year if comment to be 'ok'
    })
    OP_in.set_index('year', inplace=True)
    OP_out = pd.DataFrame({
        'year': years,
        'value': 0.0,               # float zeros; use 0 if you want ints
        'entries': 0    # need to count the right number of entries that year if comment to be 'ok'
    })
    OP_out.set_index('year', inplace=True)
    
    _add_seeds_and_planting_material(results, dataset_unc)
    _add_farm_animal_feed(results, dataset_unc)
    _add_food_industry_waste(results, dataset_unc)
    _add_food_industry_wastewater(results, dataset_unc)
    _add_food_products(results, dataset_unc)
    _add_fp_untreated_wastewater(results, dataset_unc)
    _add_aquaculture_feed(results, dataset_unc, trade_mapping, trade_params)
    _add_food_export(results, dataset_unc, trade_mapping, trade_params, trade_data)
    _add_feed_export(results, dataset_unc, trade_mapping, trade_params, trade_data)
    _add_ag_mineral_fertilizer(results, dataset_unc)
    _add_industrial_waste_fuels(results, dataset_unc)
    _add_other_industry_waste(results, dataset_unc, OP_out)
    _add_other_industry_wastewater(results, dataset_unc, OP_out)
    _add_hs_mineral_fertilizer(results, dataset_unc)
    _add_fo_mineral_fertilizer(results, dataset_unc)
    _add_op_NH3_emissions(results, dataset_unc)
    _add_op_NOx_emissions(results, dataset_unc)
    _add_op_N2O_emissions(results, dataset_unc)
    _add_op_untreated_wastewater(results, dataset_unc, OP_out)
    _add_mineral_fertilizer_export(results, dataset_unc)
    _add_other_goods_export(results, dataset_unc, OP_out, trade_mapping, trade_params, trade_data)
    _add_consumer_goods(results, dataset_unc, params, OP_out, OP_in, trade_mapping, trade_params)



    return results  # Returns a list of flow records

def _add_seeds_and_planting_material(results, dataset_unc):
    flow_code = 'MP.FP-AG.SM-Seeds and planting material -Nmix'
    collected_years = set()
    # use data from Totalkalkylen, NIBIO
    comment = 'ok'
    data_sources = 'NIBIO Totalkalkylen'
    year_values = {}
    # Seed protein fractions from parameter file
    seed_cereal_prot, u_seed_cereal_prot  = params.get_global_param_with_uncertainty("seed_cereal_protein_frac")
    seed_cereal_fac,  u_seed_cereal_fac   = params.get_global_param_with_uncertainty("seed_cereal_protein_to_N")
    seed_oil_prot,    u_seed_oil_prot     = params.get_global_param_with_uncertainty("seed_oilseed_protein_frac")
    seed_pea_prot,    u_seed_pea_prot     = params.get_global_param_with_uncertainty("seed_pea_protein_frac")
    seed_grass_prot,  u_seed_grass_prot   = params.get_global_param_with_uncertainty("seed_grass_protein_frac")
    seed_rootveg_prot,u_seed_rootveg_prot = params.get_global_param_with_uncertainty("seed_rootveg_protein_frac")
    Jones,            u_Jones             = params.get_global_param_with_uncertainty("Jones_factor")
    # Compute N fractions: protein_frac / protein_to_N_factor
    seed_cereal_N = seed_cereal_prot / seed_cereal_fac       # 0.12 / 5.83
    seed_oil_N = seed_oil_prot / Jones                       # 0.196 / 6.25
    seed_pea_N = seed_pea_prot / Jones                       # 0.225 / 6.25
    seed_grass_N = seed_grass_prot / Jones                   # 0.12 / 6.25
    seed_rootveg_N = seed_rootveg_prot / Jones               # 0.12 / 6.25
    # --- Combine uncertainties for each N-fraction (ratio product rule) ---
    u_seed_cereal_N = combine_uncertainties_percent(u_seed_cereal_prot, u_seed_cereal_fac)
    u_seed_oil_N    = combine_uncertainties_percent(u_seed_oil_prot,    u_Jones)
    u_seed_pea_N    = combine_uncertainties_percent(u_seed_pea_prot,    u_Jones)
    u_seed_grass_N  = combine_uncertainties_percent(u_seed_grass_prot,  u_Jones)
    u_seed_rootveg_N= combine_uncertainties_percent(u_seed_rootveg_prot,u_Jones)
    workbook = openpyxl.load_workbook('data_files/NibioStatistics-5.xlsx')
    u_dataset = get_uncertainty(dataset_unc, 'Totalkalkylen') # såkorn (cereal seed)
    uncertainty = combine_uncertainties_percent(u_seed_cereal_N, u_seed_oil_N, u_seed_pea_N,u_seed_grass_N, u_seed_rootveg_N,u_dataset)
    sheet = workbook['Sum innkjøpt såkorn']
    for r in range(27, 69):
        year = int(sheet.cell(row=r, column=1).value)
        value = float(sheet.cell(row=r, column=2).value)  # tonn korn
        value *= 1e-3 * seed_cereal_N                     # tons → kt, then * N fraction
        year_values[year] = year_values.get(year, 0.0) + value
    # oljefrø til modning (oilseed)
    sheet = workbook['Oljefrø til modning']
    for r in range(27, 69):
        year = int(sheet.cell(row=r, column=1).value)
        value = float(sheet.cell(row=r, column=2).value) * 1e-3 * seed_oil_N
        year_values[year] = year_values.get(year, 0.0) + value
    # erter (peas)
    sheet = workbook['Erter']
    for r in range(27, 69):
        year = int(sheet.cell(row=r, column=1).value)
        value = float(sheet.cell(row=r, column=2).value) * 1e-3 * seed_pea_N
        year_values[year] = year_values.get(year, 0.0) + value
    # engfrø (grass seed)
    sheet = workbook['Sum engfrø']
    for r in range(27, 69):
        year = int(sheet.cell(row=r, column=1).value)
        value = float(sheet.cell(row=r, column=2).value) * 1e-3 * seed_grass_N
        year_values[year] = year_values.get(year, 0.0) + value
    # rotvekst- og grønnsakfrø (root & vegetable seed)
    sheet = workbook['Sum rotvekst- og grønnsakfrø']
    for r in range(27, 69):
        year = int(sheet.cell(row=r, column=1).value)
        value = float(sheet.cell(row=r, column=2).value) * 1e-3 * seed_rootveg_N
        year_values[year] = year_values.get(year, 0.0) + value
    for year, value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_farm_animal_feed(results, dataset_unc):
    flow_code = 'MP.FP-AG.MM-Farm animal feed-Nmix'
    collected_years = set()
    comment = 'ok'
    data_sources = 'Landbruksdirektoratets kraftfôrstatistikk - årlig råvareforbruk'
    N_content_carb, u_carb = params.get_global_param_with_uncertainty("feed_carb_N_frac")
    N_content_prot, u_prot = params.get_global_param_with_uncertainty("feed_prot_N_frac")
    u_dataset = get_uncertainty(dataset_unc, 'Kraftforstatistikk')
    uncertainty = combine_uncertainties_percent(u_carb, u_prot, u_dataset)
    workbook = openpyxl.load_workbook('data_files/Årlig råvareforbruk.xlsx')
    sheet = workbook['Varegrupper']    
    N_cont = 0
    i = 1
    for r in range(5,30):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value_carb = sheet.cell(row=r, column=2).value  # norsk karbohydratråvare  
        value_prot = sheet.cell(row=r, column=8).value  # norsk proteinråvare  
        value = (value_carb*N_content_carb + value_prot*N_content_prot)/1000 # t to kt   
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, #converting from t to kt
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
        N_cont += value/(value_carb+value_prot)
        i += 1
    N_cont_before_2000 = N_cont/i*1e3
    # years before 2000
    comment = 'ok'
    data_sources = 'NIBIO Totalkalkylen'
    workbook = openpyxl.load_workbook('data_files/NibioStatistics-4.xlsx')
    uncertainty = get_uncertainty(dataset_unc, 'Totalkalkylen') # såkorn (cereal seed)
    # uncertainty = combine_uncertainties_percent(u_seed_cereal_N, u_seed_oil_N, u_seed_pea_N,u_seed_grass_N, u_seed_rootveg_N,u_dataset)
    sheet = workbook['Sum innkjøpt kraftfôr ukorr.']
    for r in range(28, 43):
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value) # tonn kraftfôr
        if year < 1995:            
            dom_frac = float(sheet.cell(row=r, column=5).value)
            comment = 'ok'
        else:
            dom_frac = 0.694
            comment = 'interpolated'            
        value *= 1e-3 * N_cont_before_2000  * dom_frac     # tons → kt, then * N fraction
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, #converting from t to kt
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_food_industry_waste(results, dataset_unc):
    flow_code = 'MP.FP-PR.SO-Food industry waste-Nmix'
    collected_years = set()
    year_values, uncertainty_10514, uncertainty_05282 = find_food_industry_waste(dataset_unc)
    comment = 'ok'
    for year,value in year_values.items():
        collected_years.add(year)
        if year > 2011:
            uncertainty = uncertainty_10514
        else:
            uncertainty = uncertainty_05282
        if year < 1995:
            data_sources = 'extrapolated'
        else:
            data_sources = 'SSB'            
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': value,
            'comment': comment,
            'data_sources': data_sources,        
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_food_industry_wastewater(results, dataset_unc):
    flow_code = 'MP.FP-PR.WW-Food industry wastewater-Nmix'
    collected_years = set()
    # bruker data fra Miljødirektoratet, utslipp per virksomhet, manuelt kategorisert
    data_sources = 'Miljødirektoratet'
    uncertainty = get_uncertainty(dataset_unc, 'norskeutslipp')
    comment = 'ok'
    emissions = pd.read_excel('data_files/Årlig utslipp til vann - Landbasert 02-02-2026.xlsx',header = 0)
    emissions = emissions[emissions['Komponent'] == 'nitrogen, totalt']
    categories = pd.read_excel('data_files/industry_categories.xlsx')
    categories_keep = categories[
        (categories['kategori'] == 'FP') 
        & (categories['kommunalt nett?'].isin(['ja']))]
    emissions_FP = emissions[emissions['AnleggNavn'].isin(categories_keep['Virksomhet'])]
    sum_by_year = emissions_FP.groupby(['År'])['Mengde'].sum().reset_index()
    for index, row in sum_by_year.iterrows():
        year = int(row['År'])
        value = (row['Mengde'])/1000
        if year in range(1989,2024):
            collected_years.add(year)
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value,
                'comment': comment,
                'data_sources': data_sources,        
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_food_products(results, dataset_unc):
    flow_code = 'MP.FP-HS.HS-Food products-Nmix'
    collected_years = set()
    year_values = {}
    # # pet food:can safely assume dogs and cats account for > 90 % of N in pet food intake
    # # use various statistics from Norway and extract trendlines for years between 1995 and 2025 (see file 'hunder_katter_norge.xlsx')
    # # nitrogen intake per animal per year taken from Schäppi2025Ann Table 19: dogs 4.37 kg/year, cats 3.04 kg/year
    # # number of dogs: 7123.8*year-13837700
    # # number of cats: 6650*year-12697800
    # pet_N = (7123.8*year-13837700)*4.37e-6 + (6650*year-12697800)*3.04e-6
    # Pet food parameters
    dog_N_per_year, u_dog_N = params.get_global_param_with_uncertainty("dog_feed_N_per_year")
    cat_N_per_year, u_cat_N = params.get_global_param_with_uncertainty("cat_feed_N_per_year")
    dog_slope,      u_dog_slope = params.get_global_param_with_uncertainty("dog_number_trend_slope")
    dog_intercept,  u_dog_int   = params.get_global_param_with_uncertainty("dog_number_trend_intercept")
    cat_slope,      u_cat_slope = params.get_global_param_with_uncertainty("cat_number_trend_slope")
    cat_intercept,  u_cat_int   = params.get_global_param_with_uncertainty("cat_number_trend_intercept")
    Jones,          u_Jones     = params.get_global_param_with_uncertainty("Jones_factor")
    u_pet = combine_uncertainties_percent(
        u_dog_N, u_cat_N,
        u_dog_slope, u_dog_int,
        u_cat_slope, u_cat_int)
    def pet_N_year(year):
        """Return pet food N (kt N/year) for given year."""
        n_dogs = dog_slope * year + dog_intercept
        n_cats = cat_slope * year + cat_intercept
        # convert kg N/year to kt N/year: *1e-6
        return (n_dogs * dog_N_per_year + n_cats * cat_N_per_year) * 1e-6
    # human food:
    # use SSB data
    # data for 2013-2017 missing, unclear why
    # dataset uncertainties for SSB tables (names as in dataset_uncertainties)
    u_13695  = get_uncertainty(dataset_unc, '13695')  # g protein/cap/day
    u_10249  = get_uncertainty(dataset_unc, '10249')   # kg food/cap/year
    u_06376  = get_uncertainty(dataset_unc, '06376')   # food intervals
    u_pop    = get_uncertainty(dataset_unc, '06913')
    # For each period, combine relevant uncertainties:
    u_period_13695 = combine_uncertainties_percent(u_13695, u_pop, u_Jones, u_pet)
    u_period_10249 = combine_uncertainties_percent(u_10249, u_pop, u_Jones, u_pet)
    u_period_06376 = combine_uncertainties_percent(u_06376, u_pop, u_Jones, u_pet)
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_interp = combine_uncertainties_percent(u_period_10249, u_interp)
    # 2018-2023: total amount of protein in sold amounts of food and drink per person and day
    workbook = openpyxl.load_workbook('data_files/13695_20260129-155515.xlsx')
    sheet = workbook['13695']
    Jones = params.get("Jones_factor")  # 6.25
    comment = 'ok'
    data_sources = 'SSB'
    population = pd.read_excel('data_files/06913_20251113-124117.xlsx',skiprows = 2,skipfooter = 42)
    population = population.set_index('Unnamed: 0')
    for col in range(2, sheet.max_column + 1):  
        year = int(sheet.cell(row=4, column=col).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=7, column=col).value)*1e-9*365 #protein, g/dag/pers -> kt/år/pers
        value *= population.loc[year, 'Befolkning 1. januar'] #protein, kt/år/pers -> kt/år
        value /= Jones # kt protein -> ktN
        value += pet_N_year(year)
        year_values[year] = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': u_period_13695
        })
    # 1999-2012: total amount of food groups given, in kg/person/year
    mengde = pd.read_excel('data_files/10249_20260129-155747.xlsx', sheet_name='10249', index_col=0,skipfooter=17,header=2).iloc[:,0::2]
    mengde = mengde.astype(str).applymap(lambda s: s.replace(',','.') if pd.notna(s) else s)
    mengde = mengde.apply(lambda col: pd.to_numeric(col, errors='coerce'))
    mengde = mengde.dropna(how='all')    
    protein = protein_per_group(params, 'protein_map_new', mengde.index)    
    # kg protein/pers/år per kode -> summer over koder
    total_protein_per_pers_år = mengde.mul(protein, axis=0).sum(axis=0)
    # kg N/pers/år -> kt N/pers/år
    total_N_per_pers_år = total_protein_per_pers_år / Jones * 1e-6
    for year, value in total_N_per_pers_år.items():
        year = int(year)
        collected_years.add(year)
        value *= population.loc[year, 'Befolkning 1. januar'] # ktN/år/pers -> ktN/år
        value += pet_N_year(year)
        year_values[year] = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': u_period_10249
        })
    # 2010-2011 2013-2017: interpolate using trendline for 1999-2023 data
    y = np.array(sorted(year_values.keys()))
    v = np.array([year_values[k] for k in y])
    m, b = np.polyfit(y, v, 1)
    data_sources = 'interpolated'
    comment = 'ok'
    for year in list(range(2010, 2012)) + list(range(2013, 2018)):
        collected_years.add(year)
        value = m*year + b        
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_interp
        })
    # 1984-1998: gitt av SSB for årsintervaller
    mengde = pd.read_excel('data_files/06376_20260129-155937.xlsx', sheet_name='06376', index_col=0,skipfooter=18,header=3).iloc[:,0::2]
    mengde = mengde.astype(str).applymap(lambda s: s.replace(',','.') if pd.notna(s) else s)
    mengde = mengde.apply(lambda col: pd.to_numeric(col, errors='coerce'))
    protein = protein_per_group(params, 'protein_map_old', mengde.index)
    total_protein_per_pers_år = mengde.mul(protein, axis=0).sum(axis=0)
    total_N_per_pers_år = total_protein_per_pers_år / Jones * 1e-6  # kgN -> ktN    
    for year in range(1984,1999):
        collected_years.add(year)
        pop = population.loc[year, 'Befolkning 1. januar'] # ktN/år/pers -> ktN/år
        comment= 'ok'
        if year < 1986:
            value = total_N_per_pers_år['1983-1985']*pop
            data_sources = 'SSB'
        elif year < 1989:
            value = (total_N_per_pers_år['1983-1985']+total_N_per_pers_år['1989-1991'])/2*pop
            data_sources = 'interpolated'
        elif year < 1992:
            value = total_N_per_pers_år['1989-1991']*pop
            data_sources = 'SSB'
        elif year < 1996:
            value = (total_N_per_pers_år['1989-1991']+total_N_per_pers_år['1996-1998'])/2*pop
            data_sources = 'interpolated'
        elif year < 1999:
            value = total_N_per_pers_år['1996-1998']*pop
            data_sources = 'SSB'
        value += pet_N_year(year)           
        year_values[year] = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': u_period_06376
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    # # use FAO Availability (protein)
    # comment = 'ok'
    # data_sources = 'FAOSTAT Availability'
    # data = pd.read_csv('data_files/FAOSTAT_data_en_11-13-2025.csv')
    # data = data[['Year', 'Value']]
    # # Value is now protein supply for all food groups combined, given as g/cap/d
    # # need to multiply by 365 (for year) and population, divide by 1e9 (g -> kt) and divide by 6.25 (Jones factor for protein)
    # # using population data from SSB
    # workbook = openpyxl.load_workbook('data_files/06913_20251113-124117.xlsx')
    # sheet = workbook['Folkemengde']
    # population = {}
    # for row in range(37, 79):  
    #     year = int(sheet.cell(row=row, column=1).value)
    #     pop = float(sheet.cell(row=row, column=2).value)
    #     population[year] = pop
    # # pet food:can safely assume dogs and cats account for > 90 % of N in pet food intake
    # # use various statistics from Norway and extract trendlines for years between 1995 and 2025 (see file 'hunder_katter_norge.xlsx')
    # # nitrogen intake per animal per year taken from Schäppi2025Ann Table 19: dogs 4.37 kg/year, cats 3.04 kg/year
    # # number of dogs: 7123.8*year-13837700
    # # number of cats: 6650*year-12697800
    # for index, row in data.iterrows():
    #     year = int(row['Year'])  
    #     collected_years.add(year)
    #     pet_N = (7123.8*year-13837700)*4.37e-6 + (6650*year-12697800)*3.04e-6
    #     value = row['Value']*population[year]*365*1e-9/6.25 + pet_N
    #     results.append({
    #         'flow_name': flow_code,
    #         'year': year,
    #         'value': value, 
    #         'comment': comment,
    #         'data_sources': data_sources
    #     })
    # missing_years = expected_years - collected_years
    # report_missing_years(flow_code, missing_years, results)

    

def _add_fp_untreated_wastewater(results, dataset_unc):
    flow_code = 'MP.FP-HY.SW-Untreated wastewater-Nmix'
    collected_years = set()
    # bruker data fra Miljødirektoratet, utslipp per virksomhet, manuelt kategorisert
    data_sources = 'Miljødirektoratet'
    uncertainty = get_uncertainty(dataset_unc, 'norskeutslipp')
    comment = 'ok'
    emissions = pd.read_excel('data_files/Årlig utslipp til vann - Landbasert 02-02-2026.xlsx',header = 0)
    emissions = emissions[emissions['Komponent'] == 'nitrogen, totalt']
    categories = pd.read_excel('data_files/industry_categories.xlsx')
    categories_keep = categories[
        (categories['kategori'] == 'FP') 
        & (categories['kommunalt nett?'].isin(['nei', 'ukjent']))]
    emissions_FP = emissions[emissions['AnleggNavn'].isin(categories_keep['Virksomhet'])]
    sum_by_year = emissions_FP.groupby(['År'])['Mengde'].sum().reset_index()
    mean_value = 0
    for index, row in sum_by_year.iterrows():
        year = int(row['År'])
        value = (row['Mengde'])/1000
        if year in range(1994,1999):
            mean_value += value
        if year in range(1994,2024):
            collected_years.add(year)
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value,
                'comment': comment,
                'data_sources': data_sources,        
                'uncertainty': uncertainty
            })
    mean_value /= 5
    for year in range (1990,1994):
        value = mean_value
        data_sources = 'extrapolated'
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,        
            'uncertainty': uncertainty
        })       
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_aquaculture_feed(results, dataset_unc, trade_mapping, trade_params):
    flow_code = 'MP.FP-HY.AC-Feed to coastal aquaculture-Nmix'
    # domestically produced aquaculture feed
    collected_years = set()
    # Feed for aquaculture
    import_fraction, u_import = params.get_global_param_with_uncertainty("aquafeed_import_fraction")
    fish_N_frac, u_fish_N = params.get_global_param_with_uncertainty("fish_N_frac")
    prot_ret, u_ret = params.get_global_param_with_uncertainty("aquafeed_N_retention")
    feed_waste, u_waste = params.get_global_param_with_uncertainty("aquafeed_waste_fraction")    
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    u_fisk = float(dataset_unc.loc['Fiskeridirektoratet', 'uncertainty'])    
    aquaculture_production, _ = find_aquaculture_production(dataset_unc)
    uncertainty = combine_uncertainties_percent(u_fisk, u_import, u_fish_N, u_ret, u_waste)
    for year, value in aquaculture_production.items():
        collected_years.add(year)
        # according to Schäppl2025Ann, p254: N content in fish and shellfish: 2.8% according to UNECE Guidance, Annex 6 Table 12
        eaten_feed_N = value / prot_ret        # kt N in eaten feed
        total_feed_N = eaten_feed_N / (1 - feed_waste)
        domestic_feed_N = total_feed_N*(1-import_fraction)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': domestic_feed_N,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_food_export(results, dataset_unc, trade_mapping, trade_params, trade_data):
    flow_code = 'MP.FP-RW.RW-Food export-Nmix'
    types_to_keep = ['korn/planter', 'kjøtt/fisk/meieri/egg', 'mat']
    impeks = 2
    aggregated_data, uncertainty = find_trade_flow(trade_data, trade_mapping, trade_params, types_to_keep, impeks, dataset_unc, wide = False)
    collected_years = set()
    data_sources = 'SSB'    
    comment = 'ok'
    for year in expected_years:
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_feed_export(results, dataset_unc, trade_mapping, trade_params, trade_data):
    flow_code = 'MP.FP-RW.RW-Feed export-Nmix'
    collected_years = set()    
    types_to_keep = ['for', 'fiskefor', 'kjæledyrfor']    
    impeks = 2
    aggregated_data, uncertainty = find_trade_flow(trade_data, trade_mapping, trade_params, types_to_keep, impeks, dataset_unc, wide = False)
    collected_years = set()
    data_sources = 'SSB'    
    comment = 'ok'
    for year in expected_years:
        if aggregated_data['year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = aggregated_data[aggregated_data['year'] == year]
            value = n_amount_row['N_amount'].values[0]
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_ag_mineral_fertilizer(results, dataset_unc):
    flow_code = 'MP.OP-AG.SM-Mineral fertilizer-Nmix'
    collected_years = set()
    # domestic fertilizer use is both from import and from domestic use
    # assume all import goes to domestic use
    # means that this flow should be given as (total domestic use) - (import)
    # this should be ok because in the Fertilizer by nutrient data from FAO, use is always larger than import
    comment = 'ok'
    uncertainty = get_uncertainty(dataset_unc, 'Fertilizer by nutrient')
    data_sources = 'FAOSTAT Fertilizer by nutrient'
    data_use = pd.read_csv('data_files/FAOSTAT_data_en_11-21-2025.csv')
    data_use = data_use[['Year', 'Value']]
    data_imp = pd.read_csv('data_files/FAOSTAT_data_en_11-12-2025-2.csv')
    data_imp = data_imp[(data_imp['Element'] == 'Import quantity')]
    data_imp = data_imp[['Year', 'Value']]
    for year in expected_years:
        n_amount_use = data_use[data_use['Year'] == year]
        if n_amount_use.empty:
            # No use data → skip; report_missing_years will handle it
            continue
    
        n_amount_imp = data_imp[data_imp['Year'] == year]
        value_imp = n_amount_imp['Value'].values[0] if not n_amount_imp.empty else 0.0
    
        value = (n_amount_use['Value'].values[0] - value_imp) / 1000
        collected_years.add(year)
        results.append({
             'flow_name': flow_code,
             'year': year,
             'value': value, 
             'comment': comment,
             'data_sources': data_sources,
             'uncertainty': uncertainty
         })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    

def _add_industrial_waste_fuels(results, dataset_unc):
    flow_code = 'MP.OP-EF.IC-Industrial waste fuels-Nmix'
    collected_years = set()
    # SSB tabell 08205 har tall for "egentilvirket bioenergi". 
    # Net caloric value for conversion, taken from https://www.ipcc-nggip.iges.or.jp/public/2006gl/pdf/2_Volume2/V2_1_Ch1_Introduction.pdf table 1.2
    # using a typical value corresponding to the 9 rows of categories in the data sheet
    # first: NCV given as TJ/Gg (Gg = M kg = kt)
    # 1 TJ = 0.28 GWh
    comment = 'ok'
    data_sources = 'SSB'
    year_values, uncertainty = find_industrial_waste_fuels(dataset_unc)
    for year, value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

    
def _add_other_industry_waste(results, dataset_unc, OP_out):
    flow_code = 'MP.OP-PR.SO-Other industry waste-Nmix'
    collected_years = set()
    industry_waste, industry_waste_unc = find_other_industry_waste(dataset_unc,params)
    data_sources = 'SSB'
    for year, value in industry_waste.items():
        collected_years.add(year)   
        if year < 1995:
            comment = 'extrapolated'
        else:
            comment = 'ok'
        OP_out.loc[year,'value'] += value
        OP_out.loc[year,'entries'] += 1
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': industry_waste_unc[year]        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_other_industry_wastewater(results, dataset_unc, OP_out):
    flow_code = 'MP.OP-PR.WW-Other industry wastewater-Nmix'
    collected_years = set()
    # bruker data fra Miljødirektoratet, utslipp per virksomhet, manuelt kategorisert
    data_sources = 'Miljødirektoratet'
    comment = 'ok'
    year_values, uncertainty = find_other_industry_wastewater(dataset_unc)
    for year,value in year_values.items():
        collected_years.add(year)        
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,        
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_hs_mineral_fertilizer(results, dataset_unc):
    flow_code = 'MP.OP-HS.HS-Mineral fertilizer-Nmix'
    collected_years = set()
    # mineral fertilizer used outside of agriculture, no official statistics in Norway
    # Schäppi2025Ann suggests a default value of 2 % of total mineral fertilizer use if no other data is available
    # I have data from FAOSTAT for agricultural use. 
    # gives F_HS = 0,2/0,98*F_AG
    comment = 'ok'
    data_sources = 'FAOSTAT Fertilizer by nutrient'
    nonag_share, u_nonag = params.get_global_param_with_uncertainty("fert_nonag_share_of_total_use")
    ag_share = 1-nonag_share        # 0.98
    nonag_over_ag = nonag_share / ag_share
    u_dataset = get_uncertainty(dataset_unc, 'Fertilizer by nutrient')   
    uncertainty = combine_uncertainties_percent(u_nonag, u_dataset)
    data_use = pd.read_csv('data_files/FAOSTAT_data_en_11-21-2025.csv')
    data_use = data_use[['Year', 'Value']]
    for year in expected_years:
        n_amount_use = data_use[data_use['Year'] == year]
        if n_amount_use.empty:
            continue
        collected_years.add(year)
        value = n_amount_use['Value'].values[0] / 1000 * nonag_over_ag
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_fo_mineral_fertilizer(results, dataset_unc):
    flow_code = 'MP.OP-FS.FO-Mineral fertilizer-Nmix'
    collected_years = set()
    # using data for forest fertilization from SSB
    comment = 'ok'
    data_sources = 'SSB table 05543'
    forest_fert_N_per_da, u_forest = params.get_global_param_with_uncertainty("forest_fert_N_per_da")
    u_dataset = get_uncertainty(dataset_unc, '05543')
    uncertainty = combine_uncertainties_percent(u_forest, u_dataset)
    workbook = openpyxl.load_workbook('data_files/05543_20251217-111610.xlsx')
    sheet = workbook['Areal']
    for r in range(4, 32): 
        year = int(sheet.cell(row=r, column=2).value)
        collected_years.add(year)
        area = float(sheet.cell(row=r, column=3).value) # da fertilized
        value = area * forest_fert_N_per_da / 1e6# using 15 kg pr dekar, https://www.nibio.no/nyheter/lnnsomt--gjdsle-skog; tN -> ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # years before 1997
    workbook = openpyxl.load_workbook('data_files/skoggjødsling_før_1995.xlsx')
    sheet = workbook['Ark1']
    for r in range(2, 14): 
        year = int(sheet.cell(row=r, column=1).value)
        collected_years.add(year)
        area = float(sheet.cell(row=r, column=2).value) # 1000m ha fertilized
        value = area * forest_fert_N_per_da / 1e2# using 15 kg pr dekar, https://www.nibio.no/nyheter/lnnsomt--gjdsle-skog; tN -> ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        }) 
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    

def _add_op_NH3_emissions(results, dataset_unc):
    flow_code = 'MP.OP-AT.AT-Emissions-NH3'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    conv, u_conv = params.get_global_param_with_uncertainty("NH3_to_N_factor")
    u_dataset = get_uncertainty(dataset_unc, 'CRLTAP')
    uncertainty = combine_uncertainties_percent(u_conv, u_dataset)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '2B')|(data[2] == '2A')|(data[2] == '2C')|(data[2] == '2D')|(
        data[2] == '2G')|(data[2] == '2H')]
    data = data[data[3] == 'NH3']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_op_NOx_emissions(results, dataset_unc):
    flow_code = 'MP.OP-AT.AT-Emissions-NOx'
    collected_years = set()
    comment = 'ok'
    # use data from CRLTAP
    data_sources = 'CRLTAP Inventory Submissions'
    conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
    u_dataset = get_uncertainty(dataset_unc, 'CRLTAP')
    uncertainty = combine_uncertainties_percent(u_conv, u_dataset)
    data = pd.read_csv('data_files/webdabData1863365.txt', sep=';', header=None,skiprows = 4)
    data = data[(data[2] == '2B')|(data[2] == '2A')|(data[2] == '2C')|(data[2] == '2D')|(
        data[2] == '2G')|(data[2] == '2H')]
    data = data[data[3] == 'NOx']
    data[1] = data[1].astype(int)
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    sums = data.groupby(1)[5].sum() * conv  
    for year, val in sums.items():
        collected_years.add(int(year))
        results.append({
            'flow_name': flow_code,
            'year': int(year),
            'value': float(val),
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_op_N2O_emissions(results, dataset_unc):
    flow_code = 'MP.OP-AT.AT-Emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 3
    data_sources = 'UNFCCC CRT'
    conv, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    u_dataset = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    uncertainty = combine_uncertainties_percent(u_conv, u_dataset)
    comment = 'ok'
    data = pd.read_csv('data_files/N2O_NOx_OP.csv')
    data = data[['year', 'N2O', 'NOx']]
    for index, row in data.iterrows():
        year = int(row['year'])
        collected_years.add(year)
        value = row['N2O']*conv
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
def _add_op_untreated_wastewater(results, dataset_unc, OP_out):
    flow_code = 'MP.OP-HY.SW-Untreated wastewater-Nmix'
    collected_years = set()
    data_sources = 'Miljødirektoratet'
    comment = 'ok'
    year_values, uncertainty = find_op_untreatewd_wastewater(dataset_unc)
    for year,value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,        
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_mineral_fertilizer_export(results, dataset_unc):
    flow_code = 'MP.OP-RW.RW-Mineral fertilizer export-Nmix'
    collected_years = set()
    # using data from FAO
    comment = 'ok'
    data_sources = 'FAOSTAT Fertilizer by Nutrient'
    data = pd.read_csv('data_files/FAOSTAT_data_en_11-12-2025-2.csv')
    filtered_data = data[(data['Element'] == 'Export quantity') & (data['Value'] != 0)]
    final_data = filtered_data[['Year', 'Value']]
    uncertainty = get_uncertainty(dataset_unc,'Fertilizer by nutrient')
    for year in expected_years:
        if final_data['Year'].isin([year]).any():
            collected_years.add(year)
            n_amount_row = final_data[final_data['Year'] == year]
            if n_amount_row.empty:
                continue
            value = (n_amount_row['Value'].values[0])/1000
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_other_goods_export(results, dataset_unc, OP_out, trade_mapping, trade_params, trade_data):
    flow_code = 'MP.OP-RW.RW-Other goods export-Nmix'
    collected_years = set()    
    year_values, uncertainty = find_other_goods_export(dataset_unc, trade_mapping, trade_params, trade_data)
    data_sources = 'SSB'    
    comment = 'ok'
    for year,value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_consumer_goods(results, dataset_unc, params, OP_out, OP_in, trade_mapping, trade_params):
    """
    Compute consumer-goods flow with the same logic as _plot_consumer_goods_mass_balance,
    but only record a value for years where all inflows and all outflows exist.

    Inflows (6 expected):
      1) AG.SM-MP.OP-Crop products for industrial use-Nmix
      2) AG.MM-MP.OP-Non-edible animal products-Nmix
      3) PR.SO-MP.OP-Recycling-Nmix
      4) EF.EC-MP.OP-Fuel used as feedstock-Nmix
      5) FS.FO-MP.OP-Industrial round wood-Nmix
      6) RW.RW-MP.OP-Other goods import -Nmix

    Outflows (4 expected):
      1) MP.OP-PR.SO-Other industry waste-Nmix
      2) MP.OP-PR.WW-Other industry wastewater-Nmix
      3) MP.OP-HY.SW-Untreated wastewater-Nmix
      4) MP.OP-RW.RW-Other goods export-Nmix
      5) MP.OP-EF.IC-Industrial waste fuels-Nmix
    """

    flow_code = 'MP.OP-HS.HS-Consumer goods-Nmix'
    collected_years = set()
    uncertainty = get_uncertainty(dataset_unc, 'mass_balance')

    # Expected number of independent flows
    N_IN  = 6
    N_OUT = 5

    inflow_totals = {}
    outflow_totals = {}
    inflow_count = {}
    outflow_count = {}

    # Small helper to add to dicts and counts
    def add_flow(year, val, totals_dict, count_dict):
        if val == 0 or val is None:
            return
        totals_dict[year] = totals_dict.get(year, 0.0) + val
        count_dict[year] = count_dict.get(year, 0) + 1

    # --- Inflows -----------------------------------------------------------
    # 1) Crop products for industrial use
    year_values, _ = find_industrial_crop_products(dataset_unc,GNB_FILE='data_files/aei_pr_gnb__custom_18744910_spreadsheet.xlsx')
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # 2) Non-edible animal products
    year_values = find_non_edible_animal_products(params)
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # 3) Recycling
    year_values, _, _, _ = find_recycling(dataset_unc, trade_mapping, trade_params)
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # 4) Fuel used as feedstock
    year_values, _ = find_feedstock_fuel(params, dataset_unc)
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # 5) Industrial round wood
    year_values, _ = find_industrial_round_wood(params, dataset_unc)
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # 6) Other goods import
    trade_params = params.get_trade_params()
    trade_mapping = params.get_trade_mapping()
    year_values, _ = find_other_goods_import(dataset_unc, trade_mapping, trade_params, trade_data)
    for year, val in year_values.items():
        add_flow(year, val, inflow_totals, inflow_count)

    # --- Outflows ----------------------------------------------------------
    # 1) Other industry waste
    year_values, _ = find_other_industry_waste(dataset_unc, params)
    for year, val in year_values.items():
        add_flow(year, val, outflow_totals, outflow_count)

    # 2) Other industry wastewater (to WW)
    year_values, _ = find_other_industry_wastewater(dataset_unc)
    for year, val in year_values.items():
        add_flow(year, val, outflow_totals, outflow_count)

    # 3) Untreated wastewater
    year_values, _ = find_op_untreatewd_wastewater(dataset_unc)
    for year, val in year_values.items():
        add_flow(year, val, outflow_totals, outflow_count)

    # 4) Other goods export
    year_values, _ = find_other_goods_export(dataset_unc, trade_mapping, trade_params, trade_data)
    for year, val in year_values.items():
        add_flow(year, val, outflow_totals, outflow_count)

    # 5) Industrial waste fuels
    year_values, _ = find_industrial_waste_fuels(dataset_unc)
    for year, val in year_values.items():
        add_flow(year, val, outflow_totals, outflow_count)

    # --- Net consumer goods per year --------------------------------------
    all_years = sorted(expected_years)  # EXPECTED_YEARS from utils

    for year in all_years:
        in_val  = inflow_totals.get(year, 0.0)
        out_val = outflow_totals.get(year, 0.0)

        n_in  = inflow_count.get(year, 0)
        n_out = outflow_count.get(year, 0)

        collected_years.add(year)

        if n_in == N_IN and n_out == N_OUT:
            value = in_val - out_val
            comment = 'ok'
            data_sources = 'mass balance'
        else:
            # You can choose: store 0, or store NaN / skip. Here we store 0 with a flag.
            value = 'nan'
            comment = f'in progress: only {n_in}/{N_IN} inflows and {n_out}/{N_OUT} outflows'
            data_sources = 'entries missing for mass balance'

        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })

    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def add_flow_column(df, flow_name, year_values, direction):
    """
    df         : DataFrame with year as index (or empty initially)
    flow_name  : column name to add
    year_values: dict {year: value}
    direction  : 'in' or 'out'

    Returns updated df (does not modify in place).
    """
    s = pd.Series(year_values, name=flow_name, dtype=float)

    # Make sure years are numeric, sorted
    s.index = pd.to_numeric(s.index)
    s = s.sort_index()

    # If df is empty, create it with this index
    if df.empty:
        df = pd.DataFrame(index=s.index)

    # Align to union of years
    df = df.reindex(df.index.union(s.index)).sort_index()

    # Inflows positive, outflows negative
    if direction == 'out':
        s = -s

    df[flow_name] = s
    df = df.fillna(0.0)

    return df

def _plot_consumer_goods_mass_balance(fname = 'consumer_goods_mass_balance.png'):
    
    df = pd.DataFrame()
    
    # Inflow 1
    year_values, _ = find_industrial_crop_products(dataset_unc,GNB_FILE='data_files/aei_pr_gnb__custom_18744910_spreadsheet.xlsx')
    df = add_flow_column(df,
                         flow_name='AG.SM-MP.OP-Crop products for industrial use-Nmix',
                         year_values=year_values,
                         direction='in')
    
    # Inflow 2
    year_values = find_non_edible_animal_products(params)
    df = add_flow_column(df,
                         flow_name='AG.MM-MP.OP-Non-edible animal products-Nmix',
                         year_values=year_values,
                         direction='in')
    
    # Inflow 3
    year_values,_,_,_ = find_recycling(dataset_unc)
    df = add_flow_column(df,
                         flow_name='AG.MM-MP.OP-Non-edible animal products-Nmix',
                         year_values=year_values,
                         direction='in')
    
    # Inflow 4
    year_values, _ = find_feedstock_fuel(params, dataset_unc)
    df = add_flow_column(df,
                         flow_name='EF.EC-MP.OP-Fuel used as feedstock-Nmix',
                         year_values=year_values,
                         direction='in')
    
    # Inflow 5
    year_values, _ = find_industrial_round_wood(params, dataset_unc)
    df = add_flow_column(df,
                         flow_name='FS.FO-MP.OP-Industrial round wood-Nmix',
                         year_values=year_values,
                         direction='in')
    
    # Inflow 6
    trade_params = params.get_trade_params()      # index: param_id
    trade_mapping = params.get_trade_mapping()    # columns: type, konv, Varenr, ...
    year_values, _ = find_other_goods_import(dataset_unc, trade_mapping, trade_params, trade_data)
    df = add_flow_column(df,
                         flow_name='RW.RW-MP.OP-Other goods import -Nmix',
                         year_values=year_values,
                         direction='in')
    
    
    # Outflow 1
    year_values_waste,_ = find_other_industry_waste(dataset_unc,params)  
    df = add_flow_column(df,
                         flow_name='MP.OP-PR.SO-Other industry waste-Nmix',
                         year_values=year_values_waste,
                         direction='out')
    
    # Outflow 2
    year_values_waste,_ = find_other_industry_wastewater(dataset_unc)  
    df = add_flow_column(df,
                         flow_name='MP.OP-PR.WW-Other industry wastewater-Nmix',
                         year_values=year_values_waste,
                         direction='out')
    
    # Outflow 3
    year_values_waste,_ = find_op_untreatewd_wastewater(dataset_unc)  
    df = add_flow_column(df,
                         flow_name='MP.OP-HY.SW-Untreated wastewater-Nmix',
                         year_values=year_values_waste,
                         direction='out')
    
    # Outflow 4
    year_values,_ = find_other_goods_export(dataset_unc, trade_mapping, trade_params, trade_data)  
    df = add_flow_column(df,
                         flow_name='MP.OP-RW.RW-Other goods export-Nmix',
                         year_values=year_values,
                         direction='out')

    years = df.index.values
    col_sum = df.sum(axis=0)  # sum over years for each flow (column)
    in_flows  = col_sum[col_sum > 0].index.tolist()
    out_flows = col_sum[col_sum < 0].index.tolist()
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.tab20.colors
    color_map = {f: colors[i % len(colors)] for i, f in enumerate(df.columns)}

    # Stack inflows (positive)
    bottom_pos = np.zeros_like(years, dtype=float)
    for f in in_flows:
        y = df[f].values
        top = bottom_pos + y
        ax.fill_between(years, bottom_pos, top,
                        color=color_map.get(f, "grey"), alpha=0.8,
                        label=f)
        bottom_pos = top

    # Stack outflows (negative)
    bottom_neg = np.zeros_like(years, dtype=float)
    for f in out_flows:
        y = df[f].values
        top = bottom_neg + y
        ax.fill_between(years, bottom_neg, top,
                        color=color_map.get(f, "grey"), alpha=0.8,
                        label=f)
        bottom_neg = top

    # Net balance
    total = df.sum(axis=1).values  # inflows - outflows
    ax.plot(years, total, color="black", linewidth=2, label="Net (in - out)")

    ax.axhline(0, color="black", linewidth=1)
    ax.set_xlabel("Year")
    ax.set_ylabel("N (kt/y)")
    ax.set_title('Consumer goods mass balance')

    ax.legend(bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()

    if fname:
        plt.savefig(fname, dpi=300)
        plt.close(fig)
    else:
        plt.show()

def plot_other_industry_waste_categories(dataset_unc, params,
                                         outpath=None,
                                         title="Other industry waste by category"):
    """
    Leser inn data som i find_other_industry_waste, men splitter på avfallstyper
    (paper, plastic, wood, etc.) og lager en kumulativ linjeplot per år.

    - X-akse: år
    - Y-akse: N (kt/y) per kategori (kumulativt stablet)
    - Hver kategori vises som en linje; den øverste linjen er total mengde.
    """

    # --- 1. Init per-category dicts ----------------------------------------
    # One dict per category: {year: value_N}
    categories = [
        "paper", "plastic", "wood", "textiles",
        "wet_organic", "other_materials",
        "hazardous", "contaminated_masses", "mixed_waste"
    ]
    cat_data = {cat: {} for cat in categories}

    # N fractions from params
    paper_N     = params.waste_N_frac("paper")
    plastic_N   = params.waste_N_frac("plastic")
    wood_N      = params.waste_N_frac("wood")
    textiles_N  = params.waste_N_frac("textiles")
    wet_org_N   = params.waste_N_frac("wet_organic")
    other_mat_N = params.waste_N_frac("other_materials")
    hazardous_N = params.waste_N_frac("hazardous")
    # contam_N    = params.waste_N_frac("contaminated_masses")
    mixed_N     = params.waste_N_frac("mixed_waste")

    # --- 2. 1995–2011: file 05282 (structure as in your code) -------------
    wb = openpyxl.load_workbook("data_files/05282_20260211-091021.xlsx")
    sheet = wb["05282"]

    for col in range(2, 170, 10):
        year = int(sheet.cell(row=4, column=col).value)

        # Start category totals at 0 for this year
        paper_val     = 0.0
        plastic_val   = 0.0
        wood_val      = 0.0
        textiles_val  = 0.0
        wet_org_val   = 0.0
        other_mat_val = 0.0
        hazardous_val = 0.0
        contam_val    = 0.0
        mixed_val     = 0.0   # not present in this period, but keep for consistency

        # Bergverk og industri col+2; Industri col+3; Annen uspesifisert næring col+8
        # Papir (row 7)
        for c in [2, 3, 8]:
            paper_val += float(sheet.cell(row=7, column=col + c).value) * paper_N

        # Plast (row 9)
        for c in [2, 3, 8]:
            plastic_val += float(sheet.cell(row=9, column=col + c).value) * plastic_N

        # Treavfall (row 12)  NOTE: you had plastic_N here; assuming that was a typo
        for c in [2, 3, 8]:
            wood_val += float(sheet.cell(row=12, column=col + c).value) * wood_N

        # Tekstiler (row 13)   NOTE: you had plastic_N here; assuming textiles_N
        for c in [2, 3, 8]:
            textiles_val += float(sheet.cell(row=13, column=col + c).value) * textiles_N

        # Våtorganisk: bare fra bergverk (col+2) (row 14)
        for c in [2]:
            wet_org_val += float(sheet.cell(row=14, column=col + c).value) * wet_org_N

        # Andre materialer (row 17)
        for c in [2, 3, 8]:
            other_mat_val += float(sheet.cell(row=17, column=col + c).value) * other_mat_N

        # Farlig avfall (row 18)
        for c in [2, 3, 8]:
            hazardous_val += float(sheet.cell(row=18, column=col + c).value) * hazardous_N

        # # Forurensede masser (row 19)
        # for c in [2, 3, 8]:
        #     contam_val += float(sheet.cell(row=19, column=col + c).value) * contam_N

        # Store per category
        cat_data["paper"][year]              = paper_val
        cat_data["plastic"][year]            = plastic_val
        cat_data["wood"][year]               = wood_val
        cat_data["textiles"][year]           = textiles_val
        cat_data["wet_organic"][year]        = wet_org_val
        cat_data["other_materials"][year]    = other_mat_val
        cat_data["hazardous"][year]          = hazardous_val
        # cat_data["contaminated_masses"][year]= contam_val
        cat_data["mixed_waste"][year]        = mixed_val  # 0 here

        if year == 1995:
            value_1995_total = (paper_val + plastic_val + wood_val + textiles_val +
                                wet_org_val + other_mat_val + hazardous_val + contam_val)

    # --- 3. 2012–2023: file 10514 -----------------------------------------
    wb = openpyxl.load_workbook("data_files/10514_20260211-094101.xlsx")
    sheet = wb["10514"]

    for col in range(2, 115, 10):
        year = int(sheet.cell(row=4, column=col).value)

        paper_val     = 0.0
        plastic_val   = 0.0
        wood_val      = 0.0
        textiles_val  = 0.0
        wet_org_val   = 0.0
        other_mat_val = 0.0
        hazardous_val = 0.0
        contam_val    = 0.0
        mixed_val     = 0.0

        # Våtorganisk avfall (row 7), only industry/mining col+2
        for c in [2]:
            wet_org_val += float(sheet.cell(row=7, column=col + c).value) * wet_org_N

        # Treavfall (row 9)
        for c in [2, 3, 8]:
            wood_val += float(sheet.cell(row=9, column=col + c).value) * wood_N

        # Papir (row 11)
        for c in [2, 3, 8]:
            paper_val += float(sheet.cell(row=11, column=col + c).value) * paper_N

        # Plast (row 17)
        for c in [2, 3, 8]:
            plastic_val += float(sheet.cell(row=17, column=col + c).value) * plastic_N

        # Tekstiler (row 19)
        for c in [2, 3, 8]:
            textiles_val += float(sheet.cell(row=19, column=col + c).value) * textiles_N

        # Andre materialer (row 24)
        for c in [2, 3, 8]:
            other_mat_val += float(sheet.cell(row=24, column=col + c).value) * other_mat_N

        # Farlig avfall (row 22)
        for c in [2, 3, 8]:
            hazardous_val += float(sheet.cell(row=22, column=col + c).value) * hazardous_N

        # Blandet avfall (row 23)
        for c in [2, 3, 8]:
            mixed_val += float(sheet.cell(row=23, column=col + c).value) * mixed_N

        # # Forurensede masser (same row 23 in your code)
        # for c in [2, 3, 8]:
        #     contam_val += float(sheet.cell(row=23, column=col + c).value) * contam_N

        # Store per category
        cat_data["paper"][year]              = paper_val
        cat_data["plastic"][year]            = plastic_val
        cat_data["wood"][year]               = wood_val
        cat_data["textiles"][year]           = textiles_val
        cat_data["wet_organic"][year]        = wet_org_val
        cat_data["other_materials"][year]    = other_mat_val
        cat_data["hazardous"][year]          = hazardous_val
        cat_data["contaminated_masses"][year]= contam_val
        cat_data["mixed_waste"][year]        = mixed_val

    # --- 4. Extrapolate 1990–1994 as in your code (total) ------------------
    wb = openpyxl.load_workbook("data_files/kommunalt_avfall_1985_1995.xlsx")
    sheet = wb["avfallsmengder"]

    waste_kt_1992 = float(sheet.cell(row=2, column=3).value)
    waste_kt_1995 = float(sheet.cell(row=3, column=3).value)
    N_frac        = value_1995_total / waste_kt_1995
    value_1992    = waste_kt_1992 * N_frac
    change_per_year = (value_1995_total - value_1992) / 3

    # For 1990–1994 we only know total; distribute across categories
    # using the 1995 category shares
    total_1995_by_cat = {cat: cat_data[cat].get(1995, 0.0) for cat in categories}
    total_1995 = sum(total_1995_by_cat.values()) or 1.0  # avoid division by zero
    cat_shares_1995 = {cat: val / total_1995 for cat, val in total_1995_by_cat.items()}

    i = 0
    for year in range(1990, 1995):
        total_val = value_1992 + change_per_year * i
        i += 1
        for cat in categories:
            cat_data[cat][year] = total_val * cat_shares_1995[cat]

    # --- 5. Build wide DataFrame: Year × category --------------------------
    # Collect all years
    all_years = sorted({yr for d in cat_data.values() for yr in d.keys()})
    df = pd.DataFrame(index=all_years)

    for cat in categories:
        s = pd.Series(cat_data[cat], name=cat, dtype=float)
        df[cat] = s

    df = df.fillna(0.0).sort_index()

    # --- 6. Plot cumulative lines per year ---------------------------------
    years = df.index.values
    cats_for_plot = [
        "wet_organic", "paper", "plastic", "wood", "textiles",
        "other_materials", "hazardous", "contaminated_masses", "mixed_waste"
    ]
    cats_for_plot = [c for c in cats_for_plot if c in df.columns]

    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.tab20.colors
    color_map = {cat: colors[i % len(colors)] for i, cat in enumerate(cats_for_plot)}

    # for cat in cats_for_plot:
    #     ax.plot(years, df[cat].values,
    #             color=color_map[cat], linewidth=2, label=cat)
    # Cumulative stacking: each line is cumulative sum over categories
    cumulative = np.zeros_like(years, dtype=float)
    for cat in cats_for_plot:
        cumulative = cumulative + df[cat].values
        ax.plot(years, cumulative,
                color=color_map[cat], linewidth=2, label=cat)

    ax.set_xlabel("Year")
    ax.set_ylabel("N in waste (kt/y)")
    ax.set_title(title)

    ax.legend(bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()

    if outpath:
        plt.savefig(outpath, dpi=300)
        plt.close(fig)
    else:
        plt.show()

    return df  # handy if you want to inspect or reuse the data



# Example usage
if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)

