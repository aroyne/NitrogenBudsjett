#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 20 09:52:40 2025

@author: anja
"""
import openpyxl
import pandas as pd
import numpy as np

from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
)
from calculations.shared_flow_calculations import (
    find_industrial_round_wood
    )

expected_years = EXPECTED_YEARS


def execute_calculations():
    results = []
    params = NParameters("data_files/N_parameters.xlsx")
    dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')
    
    _add_fo_N2O_emissions(results, params, dataset_unc)
    _add_fo_N2_emissions(results, params, dataset_unc)
    _add_fo_leaching(results, params, dataset_unc)
    _add_industrial_round_wood(results, params, dataset_unc)
    _add_fuel_wood_for_households(results, params, dataset_unc)
    _add_ol_N2O_emissions(results, params, dataset_unc)
    _add_ol_N2_emissions(results, params, dataset_unc)
    # _add_ol_NOx_emissions(results, params, dataset_unc)
    _add_ol_leaching(results, params, dataset_unc)
    _add_ol_grazing(results,params)


    return results  # Returns a list of flow records

def _add_fo_N2O_emissions(results, params, dataset_unc):
    flow_code = 'FS.FO-AT.AT-Emissions-N2O'
    collected_years = set()
    # UNFCCC, common reporting Norway https://unfccc.int/documents/644485
    # Table 4
    # NOx given as NA
    data_sources = 'UNFCCC CRT'
    comment = 'ok'
    uncertainty = 10
    workbook = openpyxl.load_workbook('data_files/N2O_NOx_HS_FS.xlsx')
    sheet = workbook['Ark1']
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    N2O_to_N, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)
    for row in range(6, 39):  
        year = int(sheet.cell(row=row, column=1).value)  
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=4).value)*N2O_to_N # ktN2O to ktN
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty        
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_fo_N2_emissions(results, params, dataset_unc):
    flow_code = 'FS.FO-AT.AT-Emissions-N2'
    collected_years = set()
    data_sources = 'UNFCCC CRT + Butterbach-Bahl et al. (2013)'
    comment = 'ok'

    ratio, u_ratio = params.get_global_param_with_uncertainty("forest_N2_to_N2O_ratio")
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    N2O_to_N, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv, u_ratio)

    workbook = openpyxl.load_workbook('data_files/N2O_NOx_HS_FS.xlsx')
    sheet = workbook['Ark1']

    for row in range(6, 39):
        year = int(sheet.cell(row=row, column=1).value)
        collected_years.add(year)
        # same base as FO N2O: column 4 kt N2O → kt N, then multiply by N2:N2O ratio
        N2O_value = float(sheet.cell(row=row, column=4).value) * N2O_to_N
        value = N2O_value * ratio
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
    
    
def _add_fo_leaching(results, params, dataset_unc):
    flow_code = 'FS.FO-HY.SW-Leaching-Nmix'
    collected_years = set()
    # data from TEOTIL3
    data_sources = 'TEOTIL'
    comment = 'ok'
    # prior to 2013: use TEOTIL data published on    https://www.miljodirektoratet.no/ansvarsomrader/overvaking-arealplanlegging/miljoovervaking/overvakingsprogrammer/forurensning-og-klimagasser/teotil/
    # use fraction of 59% of "Bakgrunn" as found from recent teotil3 data
    workbook = openpyxl.load_workbook('data_files/Tilførsel av nitrogen til kystområdene fordelt på kilder.xlsx')
    sheet = workbook['Data fra Miljødirektoratet']
    comment = 'ok'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    frac, u_frac = params.get_global_param_with_uncertainty("FO_leaching_bg_fraction")
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    # For TEOTIL3 years (direct forest component)
    uncertainty_recent = u_teotil    
    # For pre‑2013 (using fraction of background and interpolation)
    uncertainty_early = combine_uncertainties_percent(u_teotil, u_frac, u_interp)
    for r in range(2, 25):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=4).value)/1000*frac
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_early
        })
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet = workbook['totn_by_source']
    comment = 'ok'
    for r in range(2, 13):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=11).value)/1000 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_recent
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

    
def _add_industrial_round_wood(results, params, dataset_unc):
    flow_code = 'FS.FO-MP.OP-Industrial round wood-Nmix'
    collected_years = set()   
    year_values, uncertainty = find_industrial_round_wood(params, dataset_unc)
    comment = 'ok'
    data_sources = 'FAOSTAT'
    for year, value in year_values.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,   # kt N/år
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

    
def _add_fuel_wood_for_households(results, params, dataset_unc):
    flow_code = 'FS.FO-EF.OE-Fuel wood for households-Nmix'
    collected_years = set()
    # using data from SSB
    comment = 'ok'
    data_sources = 'SSB'
    u_09702 = get_uncertainty(dataset_unc, '09702')
    N_content, u_N = params.get_global_param_with_uncertainty("firewood_N_frac")
    uncertainty = combine_uncertainties_percent(u_09702, u_N)
    workbook = openpyxl.load_workbook('data_files/09702_20251120-133716.xlsx')
    sheet = workbook['VedTonn'] # unit: 1000 ton firewood
    for r in range(4, 39):  
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)*N_content 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

    
def _add_ol_grazing(results,params):
    flow_code = 'FS.OL-AG.MM-Grazing-Nmix'
    collected_years = set()
    data_sources = 'NIBIO'
    comment = 'ok'
    # using data from "organisert beitebruk", https://www.nibio.no/tema/landskap/utmarksbeite/beitebruk/beitestatistikk
    # data on feed uptakes (fu) for different animals from table 1.2 in Hegrenes&Asheim 2006: Verdi av fôr frå utmarksbeide og sysselsetting i beitebasterte næringar
    # according to that source, the feed uptake of lambs is approximately equal to that of sheep.
    # values in mill FEm
    fu_sheep_1996 = 303*0.7
    fu_cattle_1996 = 303*0.27
    fu_goat_1996 = 303*0.02
    Jones, u_Jones = params.get_global_param_with_uncertainty("Jones_factor")
    uncertainty = 20
    sau = {}
    lam = {}
    storfe = {}
    geit = {}
    workbook = openpyxl.load_workbook('data_files/OBB_Fylke_1970-2025.xlsx')
    # sau og lam
    sheet = workbook['Sau1990-99'] # unit: 1000 ton firewood
    for col in range(7,100,10):
        year = int(sheet.cell(row=1, column=col).value)
        sau[year] = float(sheet.cell(row=22, column=col-3).value)
        lam[year] = float(sheet.cell(row=22, column=col-2).value)
    sheet = workbook['Sau2000-09'] # unit: 1000 ton firewood
    for col in range(7,100,10):
        year = int(sheet.cell(row=1, column=col).value)
        sau[year] = float(sheet.cell(row=23, column=col-3).value)
        lam[year] = float(sheet.cell(row=23, column=col-2).value)
    sheet = workbook['Sau2010-19'] # unit: 1000 ton firewood
    for col in range(7,100,10):
        year = int(sheet.cell(row=1, column=col).value)
        sau[year] = float(sheet.cell(row=23, column=col-3).value)
        lam[year] = float(sheet.cell(row=23, column=col-2).value)
    sheet = workbook['Sau2020-29'] # unit: 1000 ton firewood
    for col in range(7,60,10):
        year = int(sheet.cell(row=1, column=col).value)
        sau[year] = float(sheet.cell(row=14, column=col-3).value)
        lam[year] = float(sheet.cell(row=14, column=col-2).value)
    # storfe og geit
    sheet = workbook['Storfe og geit1993-2019'] # unit: 1000 ton firewood
    for col in range(5,60,6):
        year = int(sheet.cell(row=1, column=col).value)
        storfe[year] = float(sheet.cell(row=24, column=col-2).value)
        geit[year] = float(sheet.cell(row=24, column=col-1).value)
    for col in range(67,200,8):
        year = int(sheet.cell(row=1, column=col).value)
        storfe[year] = float(sheet.cell(row=24, column=col-2).value)
        geit[year] = float(sheet.cell(row=24, column=col-1).value)
    sheet = workbook['Storfe og geit2020-29'] # unit: 1000 ton firewood
    for col in range(7,50,8):
        year = int(sheet.cell(row=1, column=col).value)
        storfe[year] = float(sheet.cell(row=14, column=col-2).value)
        geit[year] = float(sheet.cell(row=14, column=col-1).value)
    # geit og storfe: ekstrapolere tilbake til 1990
    # storfe
    years = np.array(list(storfe.keys()), dtype=float)
    y = np.array(list(storfe.values()), dtype=float)    
    a, b = np.polyfit(years, y, 1)    
    years_back = np.array([1990, 1991, 1992], dtype=float)
    y_back = a * years_back + b    
    interp_back = dict(zip(years_back.astype(int), y_back))
    storfe.update(interp_back)
    storfe = dict(sorted(storfe.items()))
    # geit
    years = np.array(list(geit.keys()), dtype=float)
    y = np.array(list(geit.values()), dtype=float)    
    a, b = np.polyfit(years, y, 1)    
    years_back = np.array([1990, 1991, 1992], dtype=float)
    y_back = a * years_back + b    
    interp_back = dict(zip(years_back.astype(int), y_back))
    geit.update(interp_back)
    geit = dict(sorted(geit.items()))
    # find feed uptakes based on 1996 values
    fu_sheep = fu_sheep_1996/sau[1996]
    fu_lamb = fu_sheep_1996/lam[1996]
    fu_cattle = fu_cattle_1996/storfe[1996]
    fu_goat = fu_goat_1996/geit[1996]
    # feed unit to N
    # antar 150 g råprotein pr FEm og ordinær Jones faktor
    protein_cont = 150e-3
    for year in range (1990,2026):
        collected_years.add(year)
        value = (sau[year]*fu_sheep+lam[year]*fu_lamb+storfe[year]*fu_cattle+geit[year]*fu_goat)*protein_cont/Jones
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_ol_N2O_emissions(results, params, dataset_unc):
    flow_code = 'FS.OL-AT.AT-Emissions-N2O'
    collected_years = set()
    data_sources = 'UNFCCC CRT'
    comment = 'ok'

    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    N2O_to_N, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv)

    workbook = openpyxl.load_workbook('data_files/N2O_NOx_HS_FS.xlsx')
    sheet = workbook['Ark1']

    for row in range(6, 39):
        year = int(sheet.cell(row=row, column=1).value)
        collected_years.add(year)
        value = float(sheet.cell(row=row, column=8).value) * N2O_to_N  # kt N2O → kt N
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_ol_N2_emissions(results, params, dataset_unc):
    flow_code = 'FS.OL-AT.AT-Emissions-N2'
    collected_years = set()
    comment = 'ok'
    data_sources = 'UNFCCC CRT + Butterbach-Bahl et al. (2013)'

    ratio, u_ratio = params.get_global_param_with_uncertainty("forest_N2_to_N2O_ratio")
    u_unfccc = get_uncertainty(dataset_unc, 'UNFCCC_emissions')
    N2O_to_N, u_conv = params.get_global_param_with_uncertainty("N2O_to_N_factor")
    uncertainty = combine_uncertainties_percent(u_unfccc, u_conv, u_ratio)

    workbook = openpyxl.load_workbook('data_files/N2O_NOx_HS_FS.xlsx')
    sheet = workbook['Ark1']

    for row in range(6, 39):
        year = int(sheet.cell(row=row, column=1).value)
        collected_years.add(year)
        N2O_value = float(sheet.cell(row=row, column=8).value) * N2O_to_N
        value = N2O_value * ratio
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value,
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty,
        })

    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
    
    
# def _add_ol_NOx_emissions(results, params, dataset_unc):
#     flow_code = 'FS.OL-AT.AT-Emissions-NOx'
#     collected_years = set()
#     comment = 'ok'
#     data_sources = 'CRLTAP Inventory Submissions'

#     u_crltap = get_uncertainty(dataset_unc, 'CRLTAP')
#     conv, u_conv = params.get_global_param_with_uncertainty("NOx_to_N_factor")
#     uncertainty = combine_uncertainties_percent(u_crltap, u_conv)

#     categories = ['4F1', '4F2']
#     sums = load_crltap_emissions_to_N(
#         filename='data_files/webdabData1863365.txt',
#         categories=categories,
#         pollutant='NOx',
#         conv_to_N=conv,
#     )

#     for year, val in sums.items():
#         year = int(year)
#         collected_years.add(year)
#         results.append({
#             'flow_name': flow_code,
#             'year': year,
#             'value': float(val),
#             'comment': comment,
#             'data_sources': data_sources,
#             'uncertainty': uncertainty,
#         })

#     missing_years = expected_years - collected_years
#     report_missing_years(flow_code, missing_years, results)
    
def _add_ol_leaching(results, params, dataset_unc):
    flow_code = 'FS.OL-HY.SW-Leaching-Nmix'
    collected_years = set()
    # data from TEOTIL3
    data_sources = 'TEOTIL'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    frac, u_frac = params.get_global_param_with_uncertainty("OL_leaching_bg_fraction")
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')
    uncertainty_recent = u_teotil
    uncertainty_early = combine_uncertainties_percent(u_teotil, u_frac, u_interp)
    comment = 'ok'
    # prior to 2013: use TEOTIL data published on    https://www.miljodirektoratet.no/ansvarsomrader/overvaking-arealplanlegging/miljoovervaking/overvakingsprogrammer/forurensning-og-klimagasser/teotil/
    # use fraction of 42% of "Bakgrunn" as found from recent teotil3 data
    workbook = openpyxl.load_workbook('data_files/Tilførsel av nitrogen til kystområdene fordelt på kilder.xlsx')
    sheet = workbook['Data fra Miljødirektoratet']
    comment = 'ok'
    for r in range(2, 25):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=4).value)/1000*frac
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_early
        })
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet = workbook['totn_by_source']
    comment = 'ok'
    for r in range(2, 13):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=11).value)/1000 # Natural losses that would be expected from agricultural areas, even if there were no agricultural activities.
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_recent
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)
