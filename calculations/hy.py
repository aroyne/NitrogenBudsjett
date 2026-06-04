#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 11 13:23:00 2025

@author: anja
"""
import pandas as pd  # Ensure you have pandas installed
import openpyxl
from calculations.n_params import NParameters
from calculations.utils import (
    EXPECTED_YEARS,
    report_missing_years,
    combine_uncertainties_percent,
    get_uncertainty,
)
from calculations.shared_flow_calculations import (
    find_aquaculture_production
    )

expected_years = EXPECTED_YEARS

params = NParameters("data_files/N_parameters.xlsx")
dataset_unc = params.get_table('dataset_uncertainties').set_index('dataset_name')

def execute_calculations():
    results = []

    years = sorted(expected_years)
    outflow = pd.DataFrame({
        'year': years,
        'value': 0.0,               # float zeros; use 0 if you want ints
        'entries': 0    # need to count the right number of entries that year if comment to be 'ok'
    })
    outflow.set_index('year', inplace=True)
    
    ww = pd.DataFrame({
        'year': years,
        'value': 0.0,               # float zeros; use 0 if you want ints
        'entries': 0    # need to count the right number of entries that year if comment to be 'ok'
    })
    ww.set_index('year', inplace=True)
    
    _add_inflow_to_coastal_waters(results, dataset_unc, ww, outflow)
    _add_wild_shellfish(results, dataset_unc)
    _add_sw_N2_emissions(results, dataset_unc, outflow)
    _add_sw_N2O_emissions(results, dataset_unc, outflow)
    _add_cw_wild_fish(results, dataset_unc)
    _add_farmed_fish(results, dataset_unc)
    _add_waste_feed(results, dataset_unc)
    _add_aquaculture_excretia(results, dataset_unc)
    

    return results  # Returns a list of flow records

def _add_inflow_to_coastal_waters(results, dataset_unc, ww, outflow):
    flow_code = 'HY.SW-HY.CW-Inflow to coastal waters-Nmix'
    collected_years = set()
    # data from TEOTIL3. Subtract PR.WW-HY.CW to avoid double counting
    data = pd.read_excel('Report.xlsx', sheet_name='2a. Database N flows', header=None, skiprows=2)  
    flow_codes = ['PR.WW-HY.CW-Treated wastewater discharge-Nmix']
    # Accessing by index: Column C = 2, Year (Column P) = 15, Value (Column N) = 13
    data[2] = data[2].str.strip()  # Strip whitespace from flow codes
    filtered_data = data.loc[(data[2].isin(flow_codes)) & (data[17] == 'ok')]  # Column index 2 for Flow Code
    filtered_data = filtered_data.rename(columns={2: 'FlowCode', 15: 'Year', 13: 'Value', 17: 'Comment'})
    for index, row in filtered_data.iterrows():
        year = int(row['Year'])
        ww.loc[year,'entries'] = 1
        if row['Comment'] == 'ok':        
            value = row['Value']
            ww.loc[year,'value'] = value
    data_sources = 'TEOTIL'
    comment = 'ok'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    uncertainty = u_teotil    # prior to 2013: use TEOTIL data published on    https://www.miljodirektoratet.no/ansvarsomrader/overvaking-arealplanlegging/miljoovervaking/overvakingsprogrammer/forurensning-og-klimagasser/teotil/
    workbook = openpyxl.load_workbook('data_files/Tilførsel av nitrogen til kystområdene fordelt på kilder.xlsx')
    sheet = workbook['Data fra Miljødirektoratet']
    comment = 'ok'
    for r in range(2, 25):# not including aquaculture or wastewater from this sheet
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=4).value)/1000 
        value += float(sheet.cell(row=r, column=5).value)/1000 
        value += float(sheet.cell(row=r, column=6).value)/1000 
        value += float(sheet.cell(row=r, column=7).value)/1000 
        collected_years.add(year)
        outflow.loc[year,'entries'] = 1
        outflow.loc[year,'value'] = value
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet1 = workbook['totn_to_coast']
    sheet2 = workbook['totn_by_source'] # to subtract aquaculture
    comment = 'ok'
    for r in range(2, 13):
        year = int(sheet1.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet1.cell(row=r, column=2).value)/1000 - float(sheet2.cell(row=r, column=4).value)/1000 
        if ww.loc[year,'entries'] == 1: # subtract discharged wastewater
            collected_years.add(year)
            value -= ww.loc[year,'value']
            outflow.loc[year,'entries'] = 1
            outflow.loc[year,'value'] = value
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_wild_shellfish(results, dataset_unc):
    flow_code = 'HY.CW-MP.FP-Shellfish-Nmix'
    collected_years = set()
    # use data from Fiskeridirektoratet on total wild fish catch
    # including macroalgae here
    workbook = openpyxl.load_workbook('data_files/art.xlsx')
    sheet = workbook['Sheet 1']
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    fish_N_frac, u_fish_N = params.get_global_param_with_uncertainty("fish_N_frac")
    seaweed_N_frac, u_seaweed_N = params.get_global_param_with_uncertainty("seaweed_N_frac")
    u_fisk = get_uncertainty(dataset_unc, 'Fiskeridirektoratet')   
    uncertainty = combine_uncertainties_percent(u_fisk, u_fish_N)
    for col in range(3, sheet.max_column + 1):  
        year = int(sheet.cell(row=1, column=col).value) 
        collected_years.add(year)
        value = 0
        for r in [36]: # skalldyr
            if sheet.cell(row=r, column=col).value is not None:
                value += float(sheet.cell(row=r, column=col).value)/1000  * fish_N_frac # converting from t fish to kt N 
        for r in [40]: # makroalger
            if sheet.cell(row=r, column=col).value is not None:
                value += float(sheet.cell(row=r, column=col).value) /1000  * seaweed_N_frac    
        value = value
        # according to Schäppl2025Ann, p254: N content in fish and shellfish: 2.8% according to UNECE Guidance, Annex 6 Table 12
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # before 2000
    workbook = openpyxl.load_workbook('data_files/fiske_1990_2000.xlsx')
    sheet = workbook['Ark1']
    for r in range(2, 12):  
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=4).value) * fish_N_frac #crustaceans
        value += float(sheet.cell(row=r, column=5).value) * seaweed_N_frac # seaweed
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)
  
def _add_sw_N2_emissions(results, dataset_unc, outflow):
    flow_code = 'HY.SW-AT.AT-Emissions-N2'
    collected_years = set()
    # retention data from TEOTIL3, assuming fraction to N2O
    data_sources = 'TEOTIL'
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet = workbook['totn_retention']
    comment = 'ok'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')    
    fraction_N2O, u_frac_N2O = params.get_global_param_with_uncertainty("surface_water_fraction_to_N2O")
    ret_frac, u_ret = params.get_global_param_with_uncertainty("surface_water_retention_fraction")    
    # For 2013+ (direct TEOTIL retention, only split into N2 and N2O by fraction_N2O)
    uncertainty_recent = combine_uncertainties_percent(u_teotil, u_frac_N2O)    
    # For pre-2013 (using typical retention + outflow): TEOTIL (underlying) + retention + fraction + interpolation
    uncertainty_early = combine_uncertainties_percent(u_teotil, u_ret, u_frac_N2O, u_interp)
    for r in range(2, 13):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)/1000 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value*(1-fraction_N2O), 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_recent
        })
    # prior to 2013: use typical retention rate
    ret_frac = params.get("surface_water_retention_fraction")
    for year in range(1984,2013):
        if outflow.loc[year,'entries'] == 1:
            collected_years.add(year)
            value = outflow.loc[year,'value']*ret_frac/(1-ret_frac)*(1-fraction_N2O)
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty_early
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_sw_N2O_emissions(results, dataset_unc, outflow):
    flow_code = 'HY.SW-AT.AT-Emissions-N2O'
    collected_years = set()
    # retention data from TEOTIL3, assuming fraction to N2O
    data_sources = 'TEOTIL'
    comment = 'ok'
    workbook = openpyxl.load_workbook('data_files/teotil3_n_summary.xlsx')
    sheet = workbook['totn_retention']
    comment = 'ok'
    u_teotil = get_uncertainty(dataset_unc, 'TEOTIL')
    u_interp = get_uncertainty(dataset_unc, 'trend interpolation')    
    fraction_N2O, u_frac_N2O = params.get_global_param_with_uncertainty("surface_water_fraction_to_N2O")
    ret_frac, u_ret = params.get_global_param_with_uncertainty("surface_water_retention_fraction")    
    # For 2013+ (direct TEOTIL retention, only split into N2 and N2O by fraction_N2O)
    uncertainty_recent = combine_uncertainties_percent(u_teotil, u_frac_N2O)    
    # For pre-2013 (using typical retention + outflow): TEOTIL (underlying) + retention + fraction + interpolation
    uncertainty_early = combine_uncertainties_percent(u_teotil, u_ret, u_frac_N2O, u_interp)
    for r in range(2, 13):
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value)/1000 
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value*fraction_N2O, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty_recent
        })
    # prior to 2013: use typical retention rate
    ret_frac = params.get("surface_water_retention_fraction")
    for year in range(1984,2013):
        if outflow.loc[year,'entries'] == 1:
            collected_years.add(year)
            value = outflow.loc[year,'value']*ret_frac/(1-ret_frac)*fraction_N2O
            results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty_early
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_cw_wild_fish(results, dataset_unc):
    flow_code = 'HY.CW-MP.FP-Fish (wild catch)-Nmix'
    collected_years = set()
    # use data from Fiskeridirektoratet on total wild fish catch
    workbook = openpyxl.load_workbook('data_files/art.xlsx')
    sheet = workbook['Sheet 1']
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    fish_N_frac, u_fish_N = params.get_global_param_with_uncertainty("fish_N_frac")
    u_fisk = get_uncertainty(dataset_unc, 'Fiskeridirektoratet')    
    uncertainty = combine_uncertainties_percent(u_fisk, u_fish_N)
    for col in range(3, sheet.max_column + 1):  
        year = int(sheet.cell(row=1, column=col).value) 
        collected_years.add(year)
        value = 0
        for r in [16,21,27,39,42]: # inkluderer makroalger her også, men ikke skalldyr
            if sheet.cell(row=r, column=col).value is not None:
                value += float(sheet.cell(row=r, column=col).value)     
        value = value/1000  * fish_N_frac # converting from t fish to kt N
        # according to Schäppl2025Ann, p254: N content in fish and shellfish: 2.8% according to UNECE Guidance, Annex 6 Table 12
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    # before 2000
    workbook = openpyxl.load_workbook('data_files/fiske_1990_2000.xlsx')
    sheet = workbook['Ark1']
    for r in range(2, 12):  
        year = int(sheet.cell(row=r, column=1).value) 
        collected_years.add(year)
        value = float(sheet.cell(row=r, column=2).value) # pelagic fish
        value += float(sheet.cell(row=r, column=3).value) # bottom fish
        value = value  * fish_N_frac # converting from kt fish to kt N
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_farmed_fish(results, dataset_unc):
    flow_code = 'HY.AC-MP.FP-Coastal fish and seafood-Nmix'
    collected_years = set()
    # use data from Fiskeridirektoratet on sold farmed fish. 
    # statistics on sold fish. According to SSB, this also includes fish that is processed in house: https://www.ssb.no/statbank/table/07326/
    aquaculture_production, uncertainty = find_aquaculture_production(dataset_unc)
    # from 1994
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    for year, value in aquaculture_production.items():
        collected_years.add(year)
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)


def _add_waste_feed(results, dataset_unc):
    flow_code = 'HY.AC-HY.CW-Waste feed-Nmix'
    collected_years = set()
    # assume 3% feed waste according to Wang2013Che        
    # Aas2022Aqu: protein retention in whole salmon/fillet
    # 2010 - 34/26 %
    # 2012 - 38/27 %
    # 2016 - 37/26 %
    # 2020 - 34/25 % 
    # 1990: much higher supply of marine ingredients - different retention?
    # Start by assuming average protein (N) retention of 35,75 %
    # Can then find input feed from produced fish / retention rate
    # use data from Fiskeridirektoratet on sold farmed fish. 
    # statistics on sold fish. According to SSB, this also includes fish that is processed in house: https://www.ssb.no/statbank/table/07326/
    aquaculture_production, uncertainty = find_aquaculture_production(dataset_unc)
    prot_ret, u_ret = params.get_global_param_with_uncertainty("aquafeed_N_retention")
    feed_waste, u_waste = params.get_global_param_with_uncertainty("aquafeed_waste_fraction")
    u_fisk = get_uncertainty(dataset_unc, 'Fiskeridirektoratet')    
    uncertainty = combine_uncertainties_percent(uncertainty, u_ret, u_waste)
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    for year, value in aquaculture_production.items():
        collected_years.add(year)
        value = value / prot_ret              # total feed N
        value = value * feed_waste / (1 - feed_waste)  # waste feed N
        results.append({
            'flow_name': flow_code,
            'year': year,
            'value': value, 
            'comment': comment,
            'data_sources': data_sources,
            'uncertainty': uncertainty
        })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)

def _add_aquaculture_excretia(results, dataset_unc):
    flow_code = 'HY.AC-HY.CW-Excretia-Nmix'
    collected_years = set()
    # assume 3% feed waste according to Wang2013Che        
    # Aas2022Aqu: protein retention in whole salmon/fillet
    # 2010 - 34/26 %
    # 2012 - 38/27 %
    # 2016 - 37/26 %
    # 2020 - 34/25 % 
    # 1990: much higher supply of marine ingredients - different retention?
    # Start by assuming average protein (N) retention of 35,75 %
    # Can then find input feed from produced fish / retention rate
    # use data from Fiskeridirektoratet on sold farmed fish. 
    # statistics on sold fish. According to SSB, this also includes fish that is processed in house: https://www.ssb.no/statbank/table/07326/
    aquaculture_production, uncertainty = find_aquaculture_production(dataset_unc)
    prot_ret, u_ret = params.get_global_param_with_uncertainty("aquafeed_N_retention")
    feed_waste, u_waste = params.get_global_param_with_uncertainty("aquafeed_waste_fraction")
    uncertainty = combine_uncertainties_percent(uncertainty, u_ret, u_waste)
    comment = 'ok'
    data_sources = 'Fiskeridirektoratet'
    for year, value in aquaculture_production.items():
        collected_years.add(year)
        value = value / prot_ret            # feed N
        value = value * (1 - prot_ret - feed_waste)  # excreted N       
        results.append({
                'flow_name': flow_code,
                'year': year,
                'value': value, 
                'comment': comment,
                'data_sources': data_sources,
                'uncertainty': uncertainty
            })
    missing_years = expected_years - collected_years
    report_missing_years(flow_code, missing_years, results)




if __name__ == "__main__":
    calculations = execute_calculations()
    for calc in calculations:
        print(calc)

